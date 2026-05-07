from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE_SOURCE_ACORE_ONLY = "ACORE Only"
FILE_SOURCE_M61_ONLY = "M61 Only"
FILE_SOURCE_BOTH = "Both"

# --------------------------------------------------
# 1. FILE PATHS
# --------------------------------------------------
# ACP II: ACP II - Liquidity & Earnings Model - 2026-03-20 - v1.xlsm
# M61 ^: (In)  Liability_Relationship_4132026_1776102909429.xlsx

# ACP III: ACP III - Liquidity & Earnings Model - 2026-03-20 - v1.xlsm
# M61^: (In)  Liability_Relationship_4102026_1775858455341.xlsx

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


# m61
DEFAULT_FILE_A_PATH = os.path.join(
    BASE_DIR, "(In)  Liability_Relationship_4102026_1775858455341.xlsx"
)

DEFAULT_FILE_B_PATH = os.path.join(
    BASE_DIR, "ACP II - Liquidity & Earnings Model - 2026-03-20 - v1.xlsm"
)

# Fund behavior by selected primary type:
# - export_label: file naming label
# - scope_label: short label shown in scoped UI messages
# - fund_token: case-insensitive token to scope rows by Fund column
PRIMARY_TYPE_FUND_CONFIG = {
    "ACORE": {
        "export_label": "ACORE - ACP III",
        "scope_label": "ACP III",
        "fund_token": "credit partners iii",
        "fund_regex": r"\bcredit partners iii\b",
        # M61-style Fund column when Liability export has no Fund Name for this row.
        "recon_fund_display": "ACORE Credit Partners III",
    },
    "ACP II": {
        "export_label": "ACORE - ACP II",
        "scope_label": "ACP II",
        "fund_token": "credit partners ii",
        "fund_regex": r"\bcredit partners ii\b",
        "recon_fund_display": "ACORE Credit Partners II",
    },
    "ACP I": {
        "export_label": "ACORE - ACP I",
        "scope_label": "ACP I",
        "fund_token": "credit partners i",
        "fund_regex": r"\bcredit partners i\b",
        "recon_fund_display": "ACORE Credit Partners I",
    },
    "AOC II": {
        "export_label": "ACORE - AOC II",
        "scope_label": "AOC II",
        # Scoped rows: Fund must spell out Roman II (do not match Opportunistic Credit I).
        "fund_token": "opportunistic credit ii",
        "fund_regex": r"\bopportunistic\s+credit\s+ii\b",
        "recon_fund_display": "ACORE Opportunistic Credit II",
    },
    "AOC I": {
        "export_label": "ACORE - AOC I",
        "scope_label": "AOC I",
        "fund_token": "opportunistic credit i",
        "fund_regex": r"\bopportunistic\s+credit\s+i\b",
        "recon_fund_display": "ACORE Opportunistic Credit I",
    },
}


def _fund_shorthand_to_canonical_map() -> dict[str, str]:
    """Shorthand / internal labels -> M61-style Fund column (single source of truth with PRIMARY_TYPE_FUND_CONFIG)."""
    out: dict[str, str] = {}
    for _ptype, cfg in PRIMARY_TYPE_FUND_CONFIG.items():
        canon = cfg.get("recon_fund_display")
        if not canon:
            continue
        for key in (cfg.get("export_label"), cfg.get("scope_label")):
            if key and str(key).strip():
                out[str(key).strip()] = str(canon).strip()
    # Legacy / raw fund-name aliases from M61 exports -> canonical display labels.
    out["Opportunistic Credit II"] = PRIMARY_TYPE_FUND_CONFIG["AOC II"]["recon_fund_display"]
    out["Opportunistic Credit I"] = PRIMARY_TYPE_FUND_CONFIG["AOC I"]["recon_fund_display"]
    return out


# Used by ``normalize_recon_fund_for_output`` (case-insensitive match on keys).
FUND_SHORTHAND_TO_CANONICAL = _fund_shorthand_to_canonical_map()


def normalize_recon_fund_for_output(df: pd.DataFrame) -> pd.DataFrame:
    """
    Replace shorthand Fund labels (e.g. ``ACORE - ACP III``) with canonical M61-style names
    for display and Excel export. Does not alter ``reconcile()``; call on the recon dataframe
    after reconciliation.
    """
    if df is None or "Fund" not in getattr(df, "columns", ()):
        return df
    if df.empty:
        return df.reset_index(drop=True)
    lc = {k.lower(): v for k, v in FUND_SHORTHAND_TO_CANONICAL.items()}

    def _norm_one(v):
        if v is None:
            return v
        try:
            if pd.isna(v):
                return v
        except (TypeError, ValueError):
            return v
        s = str(v).strip()
        if not s or s.lower() in ("nan", "<na>", "nat", "none"):
            return v
        return lc.get(s.lower(), v)

    out = df.copy()
    out["Fund"] = out["Fund"].map(_norm_one)
    return out.reset_index(drop=True)


def _fund_cfg(primary_file_type: str) -> dict:
    return PRIMARY_TYPE_FUND_CONFIG.get(primary_file_type, {})


# Primary workbooks that share Fin Inpt + liability-note-driven financing match rules.
FIN_INPT_PRIMARY_TYPES = frozenset({"ACORE", "ACP II", "ACP I", "AOC II", "AOC I"})
# Funds where Sale rows should compare against M61 DealLevelAdvanceRate.
# ACP III/ACORE intentionally excluded for now (keeps Target Advance Rate behavior).
SALE_DEALLEVEL_PRIMARY_TYPES = frozenset({"ACP II", "ACP I", "AOC II", "AOC I"})

# Optional columns when reconcile(..., match_diagnostics=True). Does not affect matching.
MATCH_DIAGNOSTIC_COLUMNS = (
    "Match Method",
    "Match Confidence",
    "ACORE Deal ID (match)",
    "M61 Deal ID (match)",
    "ACORE Effective Date Key",
    "M61 Effective Date Key",
    "Facility Key (ACORE)",
    "Facility Key (M61)",
    "Match Diagnostics Warning",
)


def _stage_to_match_method_conf(stage: object) -> tuple[str, str]:
    """Business labels for staged matchers (pair_rows / _match_stage); confidence per product guidance."""
    try:
        if stage is None or (isinstance(stage, float) and pd.isna(stage)):
            s = ""
        else:
            s = str(stage).strip().lower()
    except (TypeError, ValueError):
        s = ""
    if s in ("", "none") or s == "nan":
        return "—", "—"
    table = {
        "financing_note": ("Financing Note Match", "High confidence"),
        "strict": ("Strict Match", "High confidence"),
        "deal_id": ("Deal ID Match", "Medium confidence"),
        "source_aware_facility": ("Source/Facility Match", "Medium confidence"),
        "source_aware": ("Source/Facility Match", "Medium confidence"),
        "fallback": ("Fallback Match", "Low confidence"),
        "note_deal_id_relaxed": ("Deal ID Match (relaxed)", "Low confidence"),
    }
    return table.get(s, (str(stage).strip() or "—", "—"))


def _match_diagnostic_fields(row: pd.Series, label_a: str, *, in_a: bool, in_b: bool) -> dict[str, object]:
    """Populate MATCH_DIAGNOSTIC_COLUMNS from merged reconcile row (additive only)."""
    stage = row.get("_match_stage")
    method, conf = _stage_to_match_method_conf(stage)

    acp_did = row.get("acp_match_key") if in_b else pd.NA
    m61_did = row.get(f"{label_a}_m61_match_key") if in_a else pd.NA
    acp_eff = row.get("effective_date_key") if in_b else pd.NA
    m61_eff = row.get(f"{label_a}_effective_date_key") if in_a else pd.NA
    fac_acp = row.get("fin_acp_key") if in_b else pd.NA
    fac_m61 = row.get(f"{label_a}_fin_m61_key") if in_a else pd.NA

    warns: list[str] = []
    if in_a and in_b:
        sa = "" if pd.isna(acp_did) else str(acp_did).strip()
        sm = "" if pd.isna(m61_did) else str(m61_did).strip()
        if sa and sm and sa != sm:
            warns.append("ACORE and M61 Deal ID keys differ for this matched row.")
        ea = "" if pd.isna(acp_eff) else str(acp_eff).strip()
        em = "" if pd.isna(m61_eff) else str(m61_eff).strip()
        if ea and em and ea != em:
            warns.append("ACORE and M61 Effective Date keys differ for this matched row.")

    return {
        "Match Method": method,
        "Match Confidence": conf,
        "ACORE Deal ID (match)": acp_did,
        "M61 Deal ID (match)": m61_did,
        "ACORE Effective Date Key": acp_eff,
        "M61 Effective Date Key": m61_eff,
        "Facility Key (ACORE)": fac_acp,
        "Facility Key (M61)": fac_m61,
        "Match Diagnostics Warning": " | ".join(warns) if warns else "",
    }


LAST_RECON_DIAGNOSTICS: dict[str, int | str] = {}
LAST_RECON_CONTEXT: dict[str, pd.DataFrame | str] = {}


def get_last_recon_diagnostics() -> dict[str, int | str]:
    """Most recent reconcile() row-count diagnostics (display-only helper)."""
    return dict(LAST_RECON_DIAGNOSTICS)


def get_last_recon_context() -> dict[str, pd.DataFrame | str]:
    """Most recent reconcile() source context for UI drill-down (already-loaded data only)."""
    out: dict[str, pd.DataFrame | str] = {}
    for k, v in LAST_RECON_CONTEXT.items():
        out[k] = v.copy() if isinstance(v, pd.DataFrame) else v
    return out


def _debug_rows(msg: str) -> None:
    """Temporary row-count diagnostics for reconciliation pipeline."""
    print(f"[RECON DEBUG] {msg}")


# --- Targeted trace: M61 row LN_Fin_AOCII_25-2820 / UnCommons PE (debug only; remove when resolved) ---
_TRACE_UNCOMMONS_LN_EXACT = "LN_Fin_AOCII_25-2820"


def _mask_trace_uncommons_m61(df: pd.DataFrame | None) -> pd.Series:
    """Boolean mask for the target M61 row: exact ``LN_Fin_AOCII_25-2820`` only."""
    if df is None or df.empty or "Liability Note" not in df.columns:
        return pd.Series(dtype=bool)
    ln = df["Liability Note"].fillna("").astype(str).str.strip()
    return ln.eq(_TRACE_UNCOMMONS_LN_EXACT)


def _debug_trace_uncommons_m61_load(
    df: pd.DataFrame,
    stage: str,
    *,
    primary_file_type: str,
    in_scope_mask: pd.Series | None = None,
) -> None:
    """Trace one M61 row through expanded filter (before/after in_scope_mask)."""
    m = _mask_trace_uncommons_m61(df)
    if not m.any():
        _debug_rows(
            f"TRACE UnCommons M61 [{stage}]: NOT IN DATAFRAME "
            f"(rows={len(df)} primary_type={primary_file_type!r})"
        )
        return
    hit = df.loc[m]
    for idx, r in hit.iterrows():
        note = str(r.get("Liability Note", "") or "").strip()
        lt_raw = r.get("Liability Type")
        lt_bucket = _liability_type_bucket(lt_raw)
        in_type_bucket = lt_bucket in M61_FINANCING_TYPE_BUCKETS
        fin_tok = bool(_fin_note_scope_mask(pd.Series([note]), primary_file_type).iloc[0])
        legacy_type_or_note = in_type_bucket or fin_tok
        scope_note = ""
        if in_scope_mask is not None and idx in in_scope_mask.index:
            scope_note = f" in_scope_mask={bool(in_scope_mask.loc[idx])}"
        parsed = parse_liability_note(note)
        _debug_rows(
            f"TRACE UnCommons M61 [{stage}]: FOUND idx={idx} Deal={r.get('Deal Name')!r} "
            f"Liability Type={lt_raw!r} -> bucket={lt_bucket!r} in_financing_buckets={in_type_bucket} "
            f"fin_note_token_match={fin_tok} legacy_type_OR_note={legacy_type_or_note}{scope_note}"
        )
        _debug_rows(f"TRACE UnCommons M61 [{stage}]: parse_liability_note -> {parsed!r}")
        _debug_rows(
            f"TRACE UnCommons M61 [{stage}]: Effective Date raw={r.get('Effective Date')!r} "
            f"Fund Name={r.get('Fund Name')!r}"
        )


def _debug_trace_uncommons_primary_fin_inpt(df_b: pd.DataFrame, *, primary_file_type: str) -> None:
    """Trace primary rows for UnCommons / deal id 25-2820 on Fin Inpt side."""
    if df_b is None or df_b.empty:
        _debug_rows(f"TRACE UnCommons PRIMARY [{primary_file_type}]: empty df_b")
        return
    dn = df_b["Deal Name"].fillna("").astype(str).str.lower()
    m = dn.str.contains("uncommons", na=False)
    if "deal_id_value" in df_b.columns:
        did = df_b["deal_id_value"].fillna("").astype(str).str.strip()
        m = m | did.str.contains("25-2820", na=False)
    if "acp_extracted_deal_id" in df_b.columns:
        did2 = df_b["acp_extracted_deal_id"].fillna("").astype(str).str.strip()
        m = m | did2.str.contains("25-2820", na=False)
    if not m.any():
        _debug_rows(
            f"TRACE UnCommons PRIMARY [{primary_file_type}]: NO ROW with UnCommons / 25-2820 "
            f"(primary_rows={len(df_b)})"
        )
        return
    cols = [
        c
        for c in (
            "Deal Name",
            "deal_id_value",
            "acp_extracted_deal_id",
            "acp_match_key",
            "effective_date_key",
            "match_key",
            "Note Name",
            "Facility",
            "Source",
            "Effective Date",
        )
        if c in df_b.columns
    ]
    sub = df_b.loc[m, cols]
    _debug_rows(
        f"TRACE UnCommons PRIMARY [{primary_file_type}]: {len(sub)} row(s) — "
        f"expect match vs M61 on Deal ID + eff date + facility for AOC II"
    )
    for i, (_, r) in enumerate(sub.iterrows(), start=1):
        _debug_rows(f"TRACE UnCommons PRIMARY [{primary_file_type}] #{i}: {r.to_dict()!r}")


def _debug_trace_uncommons_m61_match_state(
    a: pd.DataFrame,
    *,
    primary_file_type: str,
    matchable_a: set,
    fin_note_rows: set | None,
    id_anchored: set | None,
    scope_lbl: str,
    _scope_ok: pd.Series | None,
) -> None:
    """After Fin Inpt guards: show whether trace row is in matchable_a / fin_note / fund scope."""
    m = _mask_trace_uncommons_m61(a)
    if not m.any():
        _debug_rows(
            f"TRACE UnCommons M61 [match_state]: row absent from df_a (rows={len(a)})"
        )
        return
    for idx, r in a.loc[m].iterrows():
        rid = int(r["_row_id_a"]) if "_row_id_a" in r.index and pd.notna(r["_row_id_a"]) else -1
        note = str(r.get("Liability Note", "") or "")
        pc = parse_liability_note(note)
        nc = pc.get("note_category", "")
        in_fin = nc == "Fin"
        in_scope_fund = bool(r.get("_m61_in_scope")) if "_m61_in_scope" in r.index else True
        id_in_pri = bool(r.get("_id_in_primary")) if "_id_in_primary" in r.index else False
        fc = (pc.get("fund_code") or "").strip()
        ok_scope = True
        if _scope_ok is not None and idx in _scope_ok.index:
            ok_scope = bool(_scope_ok.loc[idx])
        in_matchable = rid in matchable_a if rid >= 0 else False
        in_fin_note_set = rid in fin_note_rows if fin_note_rows is not None and rid >= 0 else None
        in_id_anchor = rid in id_anchored if id_anchored is not None and rid >= 0 else None
        _debug_rows(
            f"TRACE UnCommons M61 [match_state]: _row_id_a={rid} scope_lbl={scope_lbl!r} "
            f"parse_note_category={nc!r} (need Fin for fin_note_rows) fund_code={fc!r} "
            f"_m61_in_scope={in_scope_fund} _scope_ok={ok_scope} _id_in_primary={id_in_pri}"
        )
        _debug_rows(
            f"TRACE UnCommons M61 [match_state]: in fin_note_rows={in_fin_note_set} "
            f"in id_anchored={in_id_anchor} in matchable_a={in_matchable} "
            f"| keys: id_match_key={r.get('id_match_key', '')!r} fin_m61_key={r.get('fin_m61_key', '')!r} "
            f"match_key={r.get('match_key', '')!r}"
        )


def _debug_trace_uncommons_pairing_keys(a: pd.DataFrame, b: pd.DataFrame) -> None:
    """After id_match_key / fin_* keys exist: compare UnCommons M61 row to primary Fin Inpt rows."""
    m61_m = _mask_trace_uncommons_m61(a)
    if not m61_m.any():
        return
    for _, ar in a.loc[m61_m].iterrows():
        _debug_rows(
            "TRACE UnCommons [pairing_keys] M61: "
            f"id_match_key={ar.get('id_match_key')!r} fin_m61_key={ar.get('fin_m61_key')!r} "
            f"strict_key={ar.get('strict_key')!r} eff_key={ar.get('effective_date_key')!r} "
            f"facility_norm={ar.get('facility_norm')!r} source_bucket={ar.get('source_bucket')!r}"
        )
    pb = b["Deal Name"].fillna("").astype(str).str.lower().str.contains("uncommons", na=False)
    if "acp_match_key" in b.columns:
        pb = pb | b["acp_match_key"].astype(str).str.contains("25-2820", na=False)
    if not pb.any():
        _debug_rows(
            "TRACE UnCommons [pairing_keys] PRIMARY: no row with UnCommons / 25-2820 in deal_id"
        )
        return
    for _, br in b.loc[pb].iterrows():
        _debug_rows(
            "TRACE UnCommons [pairing_keys] PRIMARY: "
            f"id_match_key={br.get('id_match_key')!r} fin_acp_key={br.get('fin_acp_key')!r} "
            f"strict_key={br.get('strict_key')!r} deal={br.get('Deal Name')!r} "
            f"eff_key={br.get('effective_date_key')!r} Source={br.get('Source')!r} "
            f"facility_norm={br.get('facility_norm')!r}"
        )


def _debug_trace_uncommons_merged(merged: pd.DataFrame, *, label_a: str) -> None:
    """Locate trace row in assembled merged frame (pre-record loop)."""
    col = f"{label_a}_Liability Note"
    if merged.empty or col not in merged.columns:
        _debug_rows("TRACE UnCommons M61 [merged]: merged empty or no prefixed Liability Note")
        return
    ln = merged[col].fillna("").astype(str).str.strip()
    m = ln.eq(_TRACE_UNCOMMONS_LN_EXACT)
    if not m.any():
        _debug_rows(
            "TRACE UnCommons M61 [merged]: NOT FOUND in merged "
            f"(merged_rows={len(merged)})"
        )
        return
    hit = merged.loc[m, [c for c in ("_merge", "_match_stage", col, f"{label_a}_Deal Name") if c in merged.columns]]
    _debug_rows(f"TRACE UnCommons M61 [merged]: FOUND {len(hit)} row(s)")
    for i, (_, r) in enumerate(hit.iterrows(), start=1):
        _debug_rows(f"TRACE UnCommons M61 [merged] #{i}: {r.to_dict()!r}")


def _debug_m61_load_preview(df: pd.DataFrame, n: int = 5) -> None:
    """TEMP: first n M61 rows after load_file_a (remove after verification)."""
    want = ("Deal Name", "Liability Note", "Effective Date", "Liability Type")
    cols = [c for c in want if c in df.columns]
    if not cols:
        _debug_rows("TEMP DEBUG: M61 load preview skipped (expected columns missing)")
        return
    _debug_rows(f"TEMP DEBUG: M61 load preview (head {n}) cols={cols}")
    snippet = df[cols].head(n).to_string(index=True, max_colwidth=60)
    for ln in snippet.splitlines():
        _debug_rows(f"TEMP DEBUG:   {ln}")


def _debug_match_key_sample_rows(
    label: str,
    df: pd.DataFrame,
    *,
    deal_col: str,
    facility_col: str,
    note_col: str,
    eff_col: str,
    n: int = 10,
) -> None:
    """TEMP: print raw key components + match_key for head(n) rows (remove after diagnosis)."""
    need = (deal_col, facility_col, note_col, eff_col, "match_key")
    missing = [c for c in need if c not in df.columns]
    _debug_rows(
        f"TEMP DEBUG: match_key components — {label} (head {n}); "
        f"source cols Deal={deal_col!r} Facility={facility_col!r} "
        f"Note={note_col!r} Date={eff_col!r}"
    )
    if missing:
        _debug_rows(f"TEMP DEBUG:   skip sample (missing columns: {missing})")
        return
    sub = df.loc[:, list(need)].head(n)
    for i, (_, row) in enumerate(sub.iterrows(), start=1):
        d, fac, nte, eff, mk = (
            row[deal_col],
            row[facility_col],
            row[note_col],
            row[eff_col],
            row["match_key"],
        )
        mk_s = "" if pd.isna(mk) else str(mk)
        if len(mk_s) > 220:
            mk_s = mk_s[:220] + "..."
        _debug_rows(
            f"TEMP DEBUG:   #{i} deal={d!r} facility={fac!r} note={nte!r} "
            f"eff={eff!r} match_key={mk_s!r}"
        )


def _debug_match_key_overlap_diagnosis(df_b: pd.DataFrame, df_a: pd.DataFrame, *, n: int = 10) -> None:
    """TEMP: for head(n) primary rows, show whether any M61 row shares deal+eff date and if facility_norm aligns."""
    req = {"deal_norm", "facility_norm", "effective_date_key", "match_key"}
    if not req.issubset(df_b.columns) or not req.issubset(df_a.columns):
        _debug_rows("TEMP DEBUG: match_key overlap diagnosis skipped (norm columns missing)")
        return
    _debug_rows(
        "TEMP DEBUG: match_key diagnosis — join key is deal + facility + eff date only; "
        "note is shown above for context but is NOT in match_key. "
        "Below: same deal_norm+date_key as primary? If yes, compare facility_norm lists."
    )
    for _, r in df_b.head(n).iterrows():
        d = r["deal_norm"]
        ed = r["effective_date_key"]
        fk = r["facility_norm"]
        mk = r["match_key"]
        cand = df_a[(df_a["deal_norm"] == d) & (df_a["effective_date_key"] == ed)]
        if cand.empty:
            _debug_rows(
                "TEMP DEBUG:   "
                f"no M61 row with same deal_norm+date_key | "
                f"deal_norm={d!r} date_key={ed!r} primary_facility_norm={fk!r} match_key={mk!r}"
            )
        else:
            m61_facs = sorted({str(x) for x in cand["facility_norm"].tolist()})
            hit = mk in set(cand["match_key"].tolist())
            _debug_rows(
                "TEMP DEBUG:   "
                f"deal_norm={d!r} date_key={ed!r} | primary_facility_norm={fk!r} | "
                f"m61_facility_norm_candidates={m61_facs} | full_key_match={hit}"
            )


def _debug_unmatched_after_merge(merged: pd.DataFrame, *, label_a: str, n: int = 10) -> None:
    """TEMP: isolate rows that survive filtering but still do not align on match_key."""
    if merged.empty or "_merge" not in merged.columns:
        return
    left = merged[merged["_merge"] == "left_only"].copy()
    right = merged[merged["_merge"] == "right_only"].copy()
    _debug_rows(
        "TEMP DEBUG: unmatched survivors after outer merge — "
        f"left_only_primary={len(left)} right_only_m61={len(right)}"
    )
    left_deals = left["Deal Name"].dropna().astype(str).str.strip()
    right_deals = right[f"{label_a}_Deal Name"].dropna().astype(str).str.strip()
    overlap_deals = sorted(set(left_deals).intersection(set(right_deals)))
    _debug_rows(
        "TEMP DEBUG: unmatched deal-name overlap (indicates key-component mismatch, not filter loss) — "
        f"overlap_deals_count={len(overlap_deals)} sample={overlap_deals[:n]}"
    )
    if not left.empty:
        _debug_rows("TEMP DEBUG: sample unmatched PRIMARY rows (head 10)")
        show_cols_l = [c for c in ("Deal Name", "Facility", "Note Name", "Effective Date", "match_key") if c in left.columns]
        for i, (_, r) in enumerate(left[show_cols_l].head(n).iterrows(), start=1):
            _debug_rows(
                "TEMP DEBUG:   "
                f"#{i} deal={r.get('Deal Name')!r} facility={r.get('Facility')!r} "
                f"note={r.get('Note Name')!r} eff={r.get('Effective Date')!r} key={r.get('match_key')!r}"
            )
    if not right.empty:
        _debug_rows("TEMP DEBUG: sample unmatched M61 rows (head 10)")
        show_cols_r = [
            c
            for c in (
                f"{label_a}_Deal Name",
                f"{label_a}_Liability Name",
                f"{label_a}_Liability Note",
                f"{label_a}_Effective Date",
                f"{label_a}_match_key",
            )
            if c in right.columns
        ]
        for i, (_, r) in enumerate(right[show_cols_r].head(n).iterrows(), start=1):
            _debug_rows(
                "TEMP DEBUG:   "
                f"#{i} deal={r.get(f'{label_a}_Deal Name')!r} facility={r.get(f'{label_a}_Liability Name')!r} "
                f"note={r.get(f'{label_a}_Liability Note')!r} eff={r.get(f'{label_a}_Effective Date')!r} "
                f"key={r.get(f'{label_a}_match_key')!r}"
            )


def detect_fund_label(uploaded_filename: str | None, primary_file_type: str) -> str:
    """Infer fund label from uploaded filename, else use selected primary fund label."""
    if uploaded_filename:
        name = uploaded_filename.upper()
        if re.search(r"\bACP\s+III\b", name):
            return PRIMARY_TYPE_FUND_CONFIG["ACORE"]["export_label"]
        if re.search(r"\bACP\s+II\b", name):
            return PRIMARY_TYPE_FUND_CONFIG["ACP II"]["export_label"]
        if re.search(r"\bACP\s+I\b(?!\s*I)", name):
            return PRIMARY_TYPE_FUND_CONFIG["ACP I"]["export_label"]
        if re.search(r"\bAOC\s+II\b", name):
            return PRIMARY_TYPE_FUND_CONFIG["AOC II"]["export_label"]
        if re.search(r"\bAOC\s+I\b(?!\s*I)", name):
            return PRIMARY_TYPE_FUND_CONFIG["AOC I"]["export_label"]

    cfg = _fund_cfg(primary_file_type)
    return cfg.get("export_label", primary_file_type)


def scope_label_for_primary_type(primary_file_type: str) -> str:
    cfg = _fund_cfg(primary_file_type)
    return cfg.get("scope_label", primary_file_type)


def _fund_series_matches_regex(s: pd.Series, pattern: str | None) -> pd.Series:
    if not pattern:
        return pd.Series(True, index=s.index)
    return s.fillna("").astype(str).str.contains(pattern, case=False, regex=True, na=False)


def _fund_series_contains_label_token(s: pd.Series, lab: str | None) -> pd.Series:
    """Match primary fund labels without substring false positives (e.g. ``AOC I`` vs ``AOC II``)."""
    if not lab or not str(lab).strip():
        return pd.Series(False, index=s.index)
    esc = re.escape(str(lab).strip())
    pat = rf"(?<![A-Za-z0-9]){esc}(?![A-Za-z0-9])"
    return s.fillna("").astype(str).str.contains(pat, case=False, regex=True, na=False)


def filter_recon_to_selected_fund(df_recon: pd.DataFrame, primary_file_type: str) -> pd.DataFrame:
    """Filter reconciliation rows by selected fund scope (display/export helper).

    Uses the same fund regex as ``_m61_in_scope`` / ``PRIMARY_TYPE_FUND_CONFIG``, token-safe
    primary labels, optional ``Fund (M61)`` alignment, and (for ACP II / AOC II / AOC I) the same
    liability-note ``fund_code`` guard as staged matching—without changing merge or pairing.

    Returns a row subset only; does not change any column values (including ``Fund``).
    """
    cfg = _fund_cfg(primary_file_type)
    token = cfg.get("fund_token")
    if not token:
        return df_recon.copy()
    if "Fund" not in df_recon.columns:
        return df_recon.copy()
    pattern = cfg.get("fund_regex") or re.escape(token)
    fund_series = df_recon["Fund"].fillna("").astype(str)

    mask_regex_main = _fund_series_matches_regex(df_recon["Fund"], pattern)
    if "Fund (M61)" in df_recon.columns:
        m61f = df_recon["Fund (M61)"]
        has_m61_fund = m61f.notna() & m61f.astype(str).str.strip().ne("")
        m61_col_ok = _fund_series_matches_regex(m61f, pattern)
        mask_regex_main = mask_regex_main & (~has_m61_fund | m61_col_ok)

    scope_lbl = scope_label_for_primary_type(primary_file_type)
    if primary_file_type in ("ACP II", "AOC II", "AOC I") and "Liability Note (M61)" in df_recon.columns:
        fc = df_recon["Liability Note (M61)"].map(
            lambda n: (parse_liability_note(n).get("fund_code") or "").strip()
        )
        note_scope_ok = fc.eq("") | fc.eq(scope_lbl)
    else:
        note_scope_ok = pd.Series(True, index=df_recon.index)

    mask_primary = pd.Series(False, index=df_recon.index)
    for lab in (cfg.get("export_label"), cfg.get("scope_label"), cfg.get("recon_fund_display")):
        if not lab:
            continue
        mask_primary |= _fund_series_contains_label_token(fund_series, lab)

    combined = (mask_regex_main | mask_primary) & note_scope_ok
    out = df_recon[combined].copy()
    _debug_rows(
        f"Scoped filter ({primary_file_type}) rows: in={len(df_recon)} out={len(out)} "
        f"mask_regex={int(mask_regex_main.sum())} mask_primary={int(mask_primary.sum())} "
        f"note_scope_ok={int(note_scope_ok.sum())}"
    )
    return out


def build_output_filename(
    primary_file_type: str,
    ext: str,
    uploaded_filename: str | None = None,
) -> str:
    """Download/save name, e.g. 'ACP II Finance Recon - 2026-04-13.xlsx'."""
    date_str = datetime.today().strftime("%Y-%m-%d")
    e = ext[1:] if ext.startswith(".") else ext
    fund_label = detect_fund_label(uploaded_filename, primary_file_type)
    return f"{fund_label} Finance Recon - {date_str}.{e}"


def default_recon_output_path(
    primary_file_type: str = "ACORE",
    uploaded_filename: str | None = None,
) -> str:
    return os.path.join(
        BASE_DIR,
        build_output_filename(primary_file_type, "xlsx", uploaded_filename),
    )
# --------------------------------------------------
# 2. CONSTANTS
# --------------------------------------------------
M61_FINANCING_TYPES = ["Repo", "Non", "Subline"]
M61_FINANCING_TYPE_BUCKETS = frozenset({"repo", "non", "sale", "clo", "subline"})

TARGET_FUNDS = {
    "acore credit partners iii",
    "acore credit partners ii",
    "acore credit partners i",
    "aoc ii",
    "aoc i",
    "mcp",
    "api",
    "acore",
}

# M61 "Fund Name" for AOC: short "aoc ii" / full "… Opportunistic Credit II" (Roman II required;
# do not match Opportunistic Credit I). Roman I uses a separate pattern so "… Credit II" is not matched as I.
AOC_M61_FUND_NAME_RE = re.compile(
    r"\b(?:aoc\s+ii|opportunistic\s+credit\s+ii)\b",
    re.IGNORECASE,
)
AOC_M61_FUND_NAME_I_RE = re.compile(
    r"\b(?:aoc\s+i|opportunistic\s+credit\s+i)\b",
    re.IGNORECASE,
)

FLOAT_TOLERANCE = 1e-6
# Status-only tolerance for rate-like numeric comparisons.
# Slightly wider than FLOAT_TOLERANCE to avoid false mismatches after percent rounding/serialization.
RATE_STATUS_TOLERANCE = 5e-5
ENABLE_DEAL_ID_SUFFIX_FALLBACK = False

# Field-level reconciliation statuses (single vocabulary for business-facing output).
FIELD_STATUS_MATCH = "MATCH"
FIELD_STATUS_MISMATCH = "MISMATCH"
FIELD_STATUS_MISSING_M61 = "MISSING FROM M61"
FIELD_STATUS_MISSING_ACORE = "MISSING FROM ACORE"
FIELD_STATUS_MISSING_BOTH = "MISSING FROM BOTH"


def _file_source_label_from_sides(*, in_a: bool, in_b: bool) -> str:
    """Which file(s) contributed this reconciliation row (ACORE primary vs M61 liability export)."""
    if in_a and in_b:
        return FILE_SOURCE_BOTH
    if in_b:
        return FILE_SOURCE_ACORE_ONLY
    return FILE_SOURCE_M61_ONLY


# Temporary: targeted debug for visibility surfacing (TBK / AOCI note / Block 21 San Mateo).
_RELATED_M61_DEBUG_MARKERS = (
    "LN_Fin_AOCI_22-2203",
    "TBK",
    "block 21",
    "san mateo",
)


def _acore_row_matches_related_debug_trace(row: pd.Series) -> bool:
    parts = [
        str(row.get("Deal Name", "") or ""),
        str(row.get("Note Name", "") or ""),
        str(row.get("Facility", "") or ""),
        str(row.get("Deal ID (ACP)", "") or ""),
    ]
    blob = " | ".join(parts).lower()
    return any(m.lower() in blob for m in _RELATED_M61_DEBUG_MARKERS)


def _related_m61_lookup_keys_from_acore_output_row(row: pd.Series) -> list[str]:
    """Build normalized deal-id keys aligned with ``df_a['m61_match_key']`` (liability note suffix)."""
    keys: list[str] = []

    def _add_raw(token: object) -> None:
        if token is None or (isinstance(token, float) and pd.isna(token)):
            return
        k = normalise_deal_id_key(str(token).strip())
        if k and k not in keys:
            keys.append(k)

    _add_raw(row.get("Deal ID Match Key (ACP)"))
    _add_raw(row.get("M61 Extracted Deal ID"))
    nn = row.get("Note Name")
    _add_raw(extract_deal_id_token(nn))
    _add_raw(extract_liability_note_suffix(nn))
    _add_raw(extract_deal_id_token(row.get("Deal ID (ACP)")))
    # Parsed deal id from Note Name (same rules as M61 LN_Fin_* notes).
    try:
        pn = parse_liability_note(nn).get("deal_id") or ""
    except (TypeError, ValueError):
        pn = ""
    _add_raw(pn)
    return keys


def _pick_deal_level_advance_rate(pick: pd.Series):
    """Deal-level advance chain from an unprefixed M61 liability row (same as ``_m61_deal_level_advance_rate``)."""
    for col in ("DealLevelAdvanceRate", "Deal Level Advance Rate", "Advance Rate", "Advance"):
        if col not in pick.index:
            continue
        v = pick.get(col)
        if not _is_blank_for_compare(v):
            return v
    return pd.NA


def _pick_first_recourse_value(pick: pd.Series):
    for src in ("Recourse %", "Recourse", "RecoursePct"):
        if src not in pick.index:
            continue
        v = pick.get(src)
        if _is_blank_for_compare(v):
            continue
        return v
    return pd.NA


def _pick_undrawn_capacity(pick: pd.Series):
    for col in ("Undrawn Capacity", "Current Undrawn Capacity"):
        if col not in pick.index:
            continue
        v = pick.get(col)
        if not _is_blank_for_compare(v):
            return v
    return pd.NA


def _m61_spread_raw_from_liab_pick(pick: pd.Series):
    """Same source as matched rows: ``in_liability_Spread`` / unprefixed ``Spread`` on the liability row (no coercion)."""
    if "Spread" not in pick.index:
        return pd.NA
    return pick.get("Spread")


def _m61_index_floor_raw_from_liab_pick(pick: pd.Series):
    """Same semantics as ``_merged_liab_col(..., 'IndexFloor')`` plus M61 ``Floor`` when IndexFloor is absent."""
    for col in ("IndexFloor", "Floor"):
        if col not in pick.index:
            continue
        v = pick.get(col)
        if not _is_blank_for_compare(v):
            return v
    return pd.NA


def _related_m61_liability_type_rank_for_source(*, acore_source, m61_liability_type) -> tuple[int, str]:
    """Display-only compatibility rank for related M61 candidate selection.

    For ACORE Sale rows:
    - 0: explicit Sale/Sold liability types
    - 1: other financing-like liability types (Repo/Non/CLO/Subline/Whole Loan/Sub Debt)
    - 2: Fund/Equity
    - 3: other/blank
    For non-Sale ACORE sources: all candidates rank equally (0).
    """
    src = normalise_text(acore_source)
    lt = normalise_text(m61_liability_type)
    is_sale_source = "sale" in src
    if not is_sale_source:
        return 0, "acore_source_not_sale"

    if "sale" in lt or "sold" in lt:
        return 0, "m61_liability_type_sale"
    financing_tokens = ("repo", "non", "clo", "subline", "whole loan", "wholeloan", "sub debt", "subdebt")
    if any(tok in lt for tok in financing_tokens):
        return 1, "m61_liability_type_financing_non_sale"
    if "fund" in lt or "equity" in lt or "eq" == lt:
        return 2, "m61_liability_type_fund_or_equity"
    if not lt:
        return 3, "m61_liability_type_blank"
    return 3, "m61_liability_type_other"


def _compute_m61_advance_rate_and_source_for_display(
    *,
    primary_file_type: str,
    target_advance_rate,
    deal_level_advance_rate,
    use_deal_level_adv: bool,
) -> tuple[object, object, object]:
    """Mirror ``reconcile`` Advance Rate (M61) selection for the ``Advance Rate`` compare field.

    Same branches as the main loop: ``SALE_DEALLEVEL_PRIMARY_TYPES`` → normalized deal-level vs target;
    ``FIN_INPT_PRIMARY_TYPES`` (non–sale-deallevel funds) → target only; else legacy deal-level vs target.
    """
    compare_val = pd.NA
    adv_m61 = pd.NA
    adv_src: object = pd.NA

    if primary_file_type in SALE_DEALLEVEL_PRIMARY_TYPES:
        if use_deal_level_adv:
            compare_val = deal_level_advance_rate
            adv_m61 = _normalize_aoc_ii_m61_adv_value(compare_val)
            adv_src = "Deal Level Advance Rate"
        else:
            compare_val = target_advance_rate
            adv_m61 = _normalize_aoc_ii_m61_adv_value(compare_val)
            adv_src = "Target Advance Rate"
    elif primary_file_type in FIN_INPT_PRIMARY_TYPES:
        compare_val = target_advance_rate
        adv_m61 = compare_val
        adv_src = "Target Advance Rate"
    else:
        if use_deal_level_adv:
            compare_val = deal_level_advance_rate
            adv_m61 = compare_val
            adv_src = "Deal Level Advance Rate"
        else:
            compare_val = target_advance_rate
            adv_m61 = compare_val
            adv_src = "Target Advance Rate"

    return adv_m61, adv_src, compare_val


def _related_m61_display_updates_from_pick(
    pick: pd.Series,
    *,
    primary_file_type: str,
    out_columns: pd.Index,
    acore_row: pd.Series | None = None,
) -> tuple[dict[str, object], list[str], dict[str, object]]:
    """Map a related ``df_a`` row into M61-side **comparison/display** columns (numeric/date only where appropriate)."""
    skip_reasons: list[str] = []
    upd: dict[str, object] = {}

    tgt = pick.get("Target Advance Rate") if "Target Advance Rate" in pick.index else pd.NA
    cur_adv = pick.get("Current Advance Rate") if "Current Advance Rate" in pick.index else pd.NA
    dlr = _pick_deal_level_advance_rate(pick)

    if "Target Advance Rate (M61)" in out_columns:
        upd["Target Advance Rate (M61)"] = tgt
    if "Current Advance Rate (M61 Raw)" in out_columns:
        upd["Current Advance Rate (M61 Raw)"] = cur_adv
    if "Deal Level Advance Rate (M61 Raw)" in out_columns:
        upd["Deal Level Advance Rate (M61 Raw)"] = dlr
    if "Raw Target Advance Rate from M61" in out_columns:
        upd["Raw Target Advance Rate from M61"] = tgt
    if "Raw Current Advance Rate from M61" in out_columns:
        upd["Raw Current Advance Rate from M61"] = cur_adv
    if "Raw Deal Level Advance Rate from M61" in out_columns:
        upd["Raw Deal Level Advance Rate from M61"] = dlr

    fund_name = pick.get("Fund Name")
    liab_type = pick.get("Liability Type")
    use_deal_level_adv = _is_sale_type_fund_or_deal(fund_name=fund_name, liability_type=liab_type)

    adv_m61, adv_src, compare_val = _compute_m61_advance_rate_and_source_for_display(
        primary_file_type=primary_file_type,
        target_advance_rate=tgt,
        deal_level_advance_rate=dlr,
        use_deal_level_adv=use_deal_level_adv,
    )

    if primary_file_type in SALE_DEALLEVEL_PRIMARY_TYPES:
        if use_deal_level_adv:
            if _is_blank_for_compare(compare_val):
                skip_reasons.append("advance_rate: deal-level chain blank")
        else:
            if _is_blank_for_compare(compare_val):
                skip_reasons.append("advance_rate: Target Advance Rate blank")
    elif primary_file_type in FIN_INPT_PRIMARY_TYPES:
        if _is_blank_for_compare(compare_val):
            skip_reasons.append("advance_rate: Target Advance Rate blank (Fin Inpt)")
    else:
        if use_deal_level_adv:
            if _is_blank_for_compare(compare_val):
                skip_reasons.append("advance_rate: deal-level chain blank")
        else:
            if _is_blank_for_compare(compare_val):
                skip_reasons.append("advance_rate: Target Advance Rate blank")

    if _is_blank_for_compare(adv_m61) and not any(
        x.startswith("advance_rate:") for x in skip_reasons
    ):
        skip_reasons.append("advance_rate: resolved Advance Rate (M61) is blank")

    if "Advance Rate (M61)" in out_columns:
        upd["Advance Rate (M61)"] = adv_m61
    if "Advance Rate Source (M61)" in out_columns:
        upd["Advance Rate Source (M61)"] = adv_src if not _is_blank_for_compare(adv_m61) else pd.NA
    if "Final Advance Rate (M61)" in out_columns:
        upd["Final Advance Rate (M61)"] = adv_m61

    raw_dlr_col = (
        pick.get("DealLevelAdvanceRate") if "DealLevelAdvanceRate" in pick.index else pd.NA
    )
    ar_meta = {
        "source_acore": acore_row.get("Source") if acore_row is not None else pd.NA,
        "m61_target_advance": tgt,
        "m61_raw_deallevel_advance_rate": raw_dlr_col,
        "m61_deal_level_chain": dlr,
        "use_deal_level_adv": use_deal_level_adv,
        "selected_source": adv_src,
        "advance_rate_m61": adv_m61,
    }

    sp_raw = _m61_spread_raw_from_liab_pick(pick)
    if "Spread (M61)" in out_columns:
        # Matched rows use raw ``spread_m61`` (see ``record["Spread (M61)"]``); do not coerce here.
        upd["Spread (M61)"] = pd.NA if _is_blank_for_compare(sp_raw) else sp_raw
    if _is_blank_for_compare(sp_raw):
        skip_reasons.append("spread: Spread blank or missing on M61 row")

    ix_floor_raw = _m61_index_floor_raw_from_liab_pick(pick)
    ix_norm = _normalize_index_floor_value(ix_floor_raw)
    if "Index Floor (M61)" in out_columns:
        upd["Index Floor (M61)"] = ix_norm
    try:
        ix_na = ix_norm is pd.NA or pd.isna(ix_norm)
    except (TypeError, ValueError):
        ix_na = ix_norm is pd.NA
    if _is_blank_for_compare(ix_floor_raw):
        skip_reasons.append("index_floor: IndexFloor/Floor blank on M61 row")
    elif ix_na:
        skip_reasons.append("index_floor: normalized empty/zero-like")

    ix_nm_out: object = pd.NA
    if "IndexName" in pick.index:
        ix_nm_raw = pick.get("IndexName")
        if ix_nm_raw is None or (isinstance(ix_nm_raw, float) and pd.isna(ix_nm_raw)):
            ix_nm_out = pd.NA
            skip_reasons.append("index_name: IndexName blank")
        else:
            s = str(ix_nm_raw).strip()
            if not s:
                ix_nm_out = pd.NA
                skip_reasons.append("index_name: whitespace-only")
            else:
                ix_nm_out = s
        if "Index Name (M61)" in out_columns:
            upd["Index Name (M61)"] = ix_nm_out
    else:
        skip_reasons.append("index_name: IndexName column absent on M61 row")

    rc = _pick_first_recourse_value(pick)
    rc_num = pd.NA
    if not _is_blank_for_compare(rc):
        rcc = _coerce_numeric_value(rc)
        if rcc is None:
            skip_reasons.append("recourse: not coercible to number (left blank)")
        else:
            rc_num = rcc
    else:
        skip_reasons.append("recourse: no Recourse % / Recourse / RecoursePct populated")
    if "Recourse % (M61)" in out_columns:
        upd["Recourse % (M61)"] = rc_num

    uc = _pick_undrawn_capacity(pick)
    uc_num = pd.NA
    if not _is_blank_for_compare(uc):
        ucc = _coerce_numeric_value(uc)
        if ucc is None:
            skip_reasons.append("undrawn: not coercible to number (left blank)")
        else:
            uc_num = ucc
    else:
        skip_reasons.append("undrawn: Undrawn Capacity / Current Undrawn Capacity blank")
    if "Undrawn Capacity (M61)" in out_columns:
        upd["Undrawn Capacity (M61)"] = uc_num

    return upd, skip_reasons, ar_meta


def _surface_related_m61_for_acore_only_rows(
    df_out: pd.DataFrame,
    df_a: pd.DataFrame,
    *,
    paired_row_ids_m61: set[int],
    primary_file_type: str = "ACORE",
) -> pd.DataFrame:
    """Populate M61-facing columns on **ACORE-only** rows from a related ``df_a`` line.

    Candidate pool: all M61 rows sharing the same deal / line keys (matched or unmatched).
    **Ranking:** (1) liability-type compatibility with ACORE Source, (2) same Pledge Date as the ACORE row
    (closest pledge-date distance first), (3) earliest M61 Effective Date, (4) unmatched rows preferred
    over already-paired rows as a final tiebreaker.
    Values copied from the chosen M61 row only (``Effective Date``, ``Pledge Date``, etc. are not taken from ACORE).

    Does **not** modify pairing, ``recon_status``, ``File Source``, or any ``* Status`` column.

    **Important:** Do not write related M61 text into ``Liability Note (M61)`` — that column feeds
    ``filter_recon_to_selected_fund`` (note fund_code guard).
    """
    n_before = len(df_out) if df_out is not None else 0
    _debug_rows(f"Related M61 enhancement: row count BEFORE={n_before}")

    if df_out is None or df_out.empty or df_a is None or df_a.empty:
        _debug_rows(
            f"Related M61 enhancement: row count AFTER={n_before} removed_or_merged=none "
            "(skipped: empty df_out or df_a)"
        )
        return df_out
    need = {"m61_match_key", "facility_norm", "effective_date_key", "Effective Date", "_row_id_a"}
    if not need.issubset(df_a.columns):
        _debug_rows(
            "TEMP DEBUG: related M61 surfacing skipped — df_a missing required columns "
            f"(have={set(df_a.columns)!r})"
        )
        _debug_rows(
            f"Related M61 enhancement: row count AFTER={n_before} removed_or_merged=none "
            "(skipped: df_a columns)"
        )
        return df_out
    if "File Source" not in df_out.columns:
        _debug_rows(
            f"Related M61 enhancement: row count AFTER={n_before} removed_or_merged=none "
            "(skipped: no File Source)"
        )
        return df_out

    ac_mask = df_out["File Source"].eq(FILE_SOURCE_ACORE_ONLY)
    if not ac_mask.any():
        _debug_rows(
            f"Related M61 enhancement: row count AFTER={n_before} removed_or_merged=none "
            "(skipped: no ACORE-only rows)"
        )
        return df_out

    out = df_out.copy()

    n_done = 0
    for idx in out.index[ac_mask]:
        row = out.loc[idx]
        if _acore_row_matches_related_debug_trace(row):
            _debug_rows(
                "TEMP DEBUG: related M61 surfacing RECOMPUTE row "
                f"deal={row.get('Deal Name')!r} eff_acp={row.get('Effective Date (ACP)')!r} "
                f"existing_eff_m61={row.get('Effective Date (M61)')!r}"
            )

        keys = _related_m61_lookup_keys_from_acore_output_row(row)
        dn = normalise_text(row.get("Deal Name") or "")
        fn_pri = normalise_facility(row.get("Facility"))

        _trace = _acore_row_matches_related_debug_trace(row)
        if _trace:
            _debug_rows(
                "TEMP DEBUG: related M61 surfacing TRACE — ACORE keys="
                f"{keys!r} deal={row.get('Deal Name')!r} facility={row.get('Facility')!r} "
                f"note={str(row.get('Note Name') or '')[:120]!r} eff_acp={row.get('Effective Date (ACP)')!r}"
            )

        def _pool_from_df(prefer_unmatched: bool) -> pd.DataFrame:
            base = df_a
            mask = pd.Series(False, index=base.index)
            for k in keys:
                if not k:
                    continue
                mask |= base["m61_match_key"].astype(str).str.strip().eq(k)
            cand = base.loc[mask].copy()
            if cand.empty and dn and "deal_norm" in base.columns:
                if fn_pri and "facility_norm" in base.columns:
                    cand = base.loc[
                        base["deal_norm"].astype(str).eq(dn)
                        & base["facility_norm"].astype(str).eq(fn_pri)
                    ].copy()
                else:
                    cand = base.loc[base["deal_norm"].astype(str).eq(dn)].copy()
            if cand.empty:
                return cand
            if prefer_unmatched and paired_row_ids_m61:
                um = cand.loc[~cand["_row_id_a"].astype(int).isin(paired_row_ids_m61)]
                if not um.empty:
                    return um
            return cand

        # Always build the pool from ALL candidates (matched or not); ranking handles the preference.
        # Using prefer_unmatched=True would silently drop the correct M61 row when it is already
        # paired with a different ACORE row, causing a lower-quality (wrong-date) row to be selected.
        pool = _pool_from_df(prefer_unmatched=False)
        pick_reason = "all_candidates_ranked"
        if pool.empty and keys:
            if _trace:
                sample = (
                    df_a.loc[:, ["m61_match_key", "Deal Name", "Liability Note", "effective_date_key"]]
                    .head(25)
                    .to_dict("records")
                )
                _debug_rows(
                    "TEMP DEBUG: related M61 surfacing — NO pool after key/deal fallbacks; "
                    f"keys_tried={keys!r} sample_m61_keys={sample!r}"
                )
            continue

        if pool.empty:
            continue

        narrowed = pool.copy()
        if fn_pri and "facility_norm" in narrowed.columns:
            sub = narrowed.loc[narrowed["facility_norm"].astype(str).eq(fn_pri)]
            if not sub.empty:
                narrowed = sub
        nn_pri = normalise_text(row.get("Note Name") or "")
        if nn_pri and "strict_note_norm" in narrowed.columns:
            sub = narrowed.loc[narrowed["strict_note_norm"].astype(str).eq(nn_pri)]
            if not sub.empty:
                narrowed = sub
        if narrowed.empty:
            narrowed = pool.copy()

        # Candidate ranking (display-only):
        # 1) same deal/liability key (pool/narrowed above)
        # 2) liability-type compatibility with ACORE Source (Sale source: Sale > Financing > Fund/Equity)
        # 3) same/closest pledge date
        # 4) earliest M61 effective date (NOT closeness to ACORE date — that would pick wrong rows)
        # 5) unmatched M61 rows preferred over already-paired rows as tiebreaker
        # Uses actual M61 ``Effective Date`` / ``Pledge Date`` columns — never ACORE dates for keys on ``df_a``.
        acore_source = row.get("Source")
        narrowed = narrowed.copy()
        _lt_rank_reason = narrowed.apply(
            lambda rr: _related_m61_liability_type_rank_for_source(
                acore_source=acore_source,
                m61_liability_type=rr.get("Liability Type"),
            ),
            axis=1,
        )
        narrowed["__lt_rank"] = _lt_rank_reason.map(lambda t: int(t[0]))
        narrowed["__lt_reason"] = _lt_rank_reason.map(lambda t: str(t[1]))
        # For ACORE Sale rows, only fallback to Fund/Equity when no Sale/Financing candidate exists.
        src_norm = normalise_text(acore_source)
        if "sale" in src_norm and (narrowed["__lt_rank"] <= 1).any():
            narrowed = narrowed.loc[narrowed["__lt_rank"] <= 1].copy()

        pri_ed = row.get("Effective Date (ACP)")
        if pri_ed is None or (isinstance(pri_ed, float) and pd.isna(pri_ed)):
            pri_ed = row.get("Effective Date")
        pri_ts = pd.to_datetime(pri_ed, errors="coerce")
        pri_pl = row.get("Pledge Date (ACP)")
        if pri_pl is None or (isinstance(pri_pl, float) and pd.isna(pri_pl)):
            pri_pl = row.get("Pledge Date")
        pri_pl_ts = pd.to_datetime(pri_pl, errors="coerce")

        work = narrowed.assign(
            __ed_ts=pd.to_datetime(narrowed["Effective Date"], errors="coerce")
        )
        if "Pledge Date" in narrowed.columns:
            work["__pl_ts"] = pd.to_datetime(narrowed["Pledge Date"], errors="coerce")
        else:
            work["__pl_ts"] = pd.NaT

        if pd.notna(pri_pl_ts):
            work["__pl_dist"] = (work["__pl_ts"] - pri_pl_ts).abs().dt.total_seconds()
        else:
            work["__pl_dist"] = pd.NA

        if pd.notna(pri_ts):
            work["__ed_dist"] = (work["__ed_ts"] - pri_ts).abs().dt.total_seconds()
        else:
            work["__ed_dist"] = pd.NA

        # __is_paired = 0 for unmatched M61 rows, 1 for already-paired rows.
        # Sorting ascending puts unmatched rows last in priority (preferred as tiebreaker).
        work["__is_paired"] = work["_row_id_a"].astype(int).isin(paired_row_ids_m61).astype(int)
        # Sort: (1) liability-type rank, (2) pledge-date proximity, (3) earliest M61 effective date,
        # (4) unmatched-before-paired tiebreaker.
        # __ed_dist (distance from ACORE effective date) is intentionally NOT used: for ACORE-only rows
        # that date proximity is circular — the correct M61 row often has a DIFFERENT effective date,
        # and favouring the row closest to ACORE's date selects the wrong candidate.
        work = work.sort_values(
            ["__lt_rank", "__pl_dist", "effective_date_key", "__is_paired"],
            ascending=[True, True, True, True],
            na_position="last",
        )

        pick = work.iloc[0]
        rid_pick = int(pick["_row_id_a"])
        _sel_reason = (
            f"pick_reason={pick_reason}; lt_reason={pick.get('__lt_reason')!r}; "
            f"lt_rank={pick.get('__lt_rank')!r}; pledge_dist={pick.get('__pl_dist')!r}; "
            f"eff_date_key={pick.get('effective_date_key')!r}; is_paired={pick.get('__is_paired')!r}"
        )
        cand_debug = []
        for _, rr in work.head(30).iterrows():
            cand_debug.append(
                {
                    "row_id_a": int(rr.get("_row_id_a")) if not pd.isna(rr.get("_row_id_a")) else None,
                    "liability_note": rr.get("Liability Note"),
                    "liability_type": rr.get("Liability Type"),
                    "effective_date": rr.get("Effective Date"),
                    "pledge_date": rr.get("Pledge Date"),
                    "spread": _m61_spread_raw_from_liab_pick(rr),
                    "index_floor": _m61_index_floor_raw_from_liab_pick(rr),
                    "lt_reason": rr.get("__lt_reason"),
                    "lt_rank": rr.get("__lt_rank"),
                    "pledge_dist": rr.get("__pl_dist"),
                    "is_paired": bool(rr.get("__is_paired", 0)),
                }
            )
        _debug_rows(
            "Related M61 candidates: "
            f"deal={row.get('Deal Name')!r} source_acore={acore_source!r} "
            f"eff_acore={row.get('Effective Date (ACP)')!r} pledge_acore={row.get('Pledge Date (ACP)')!r} "
            f"candidate_count={len(work)} candidates={cand_debug!r} selected_reason={_sel_reason}"
        )

        display_updates, skip_reasons, ar_meta = _related_m61_display_updates_from_pick(
            pick,
            primary_file_type=primary_file_type,
            out_columns=out.columns,
            acore_row=row,
        )
        _upd: dict[str, object] = {
            "Effective Date (M61)": pick.get("Effective Date"),
            "Liability Name (M61)": pick.get("Liability Name"),
            # Do not set Liability Note (M61): fund scoping reads it via parse_liability_note().
            "Liability Type (M61 Raw)": pick.get("Liability Type"),
        }
        _upd.update(display_updates)
        # Hard display-only guard: for related M61 attach, all M61-side values come from THIS selected pick row.
        # Never take ACORE dates/keys/fallback dates for these columns.
        _upd["Effective Date (M61)"] = pick.get("Effective Date")
        if "Pledge Date" in pick.index:
            _upd["Pledge Date (M61)"] = pick.get("Pledge Date")
        _upd["Liability Type (M61 Raw)"] = pick.get("Liability Type")
        _upd["Spread (M61)"] = (
            pd.NA if _is_blank_for_compare(_m61_spread_raw_from_liab_pick(pick)) else _m61_spread_raw_from_liab_pick(pick)
        )
        _upd["Index Floor (M61)"] = _normalize_index_floor_value(_m61_index_floor_raw_from_liab_pick(pick))
        if "IndexName" in pick.index:
            _ixn = pick.get("IndexName")
            _upd["Index Name (M61)"] = (
                pd.NA
                if (_ixn is None or (isinstance(_ixn, float) and pd.isna(_ixn)) or not str(_ixn).strip())
                else str(_ixn).strip()
            )
        _rc_pick = _pick_first_recourse_value(pick)
        _rc_num = _coerce_numeric_value(_rc_pick)
        _upd["Recourse % (M61)"] = pd.NA if _rc_num is None else _rc_num
        _uc_pick = _pick_undrawn_capacity(pick)
        _uc_num = _coerce_numeric_value(_uc_pick)
        _upd["Undrawn Capacity (M61)"] = pd.NA if _uc_num is None else _uc_num
        if "M61 Extracted Deal ID" in out.columns:
            _v = pick.get("liability_note_suffix_key")
            if _v is None or (isinstance(_v, float) and pd.isna(_v)):
                _v = pick.get("m61_extracted_deal_id")
            _upd["M61 Extracted Deal ID"] = _v
        if "Pledge Date (M61)" in out.columns and "Pledge Date" in pick.index:
            _upd["Pledge Date (M61)"] = pick.get("Pledge Date")

        for k, v in _upd.items():
            if k in out.columns:
                out.at[idx, k] = v

        _is_target = (
            str(row.get("Deal Name") or "").strip().lower() == "block 21 san mateo"
            and str(row.get("Facility") or "").strip().lower() == "tbk bank"
            and str(row.get("Source") or "").strip().lower() == "sale"
            and str(pd.to_datetime(row.get("Effective Date (ACP)"), errors="coerce").strftime("%Y-%m-%d"))
            == "2022-08-22"
        )
        if _is_target:
            _debug_rows(
                "UNDRAWN TRACE 3 (after related attach): "
                f"selected_pick_raw_undrawn={pick.get('Undrawn Capacity')!r} "
                f"selected_pick_raw_current_undrawn={pick.get('Current Undrawn Capacity')!r} "
                f"stored_undrawn_m61={out.at[idx, 'Undrawn Capacity (M61)'] if 'Undrawn Capacity (M61)' in out.columns else pd.NA!r}"
            )

        # For related-M61 context rows (still File Source = ACORE Only), evaluate field statuses
        # against displayed M61-side values instead of defaulting to "MISSING FROM M61".
        _r = out.loc[idx]
        if "Effective Date Status" in out.columns:
            out.at[idx, "Effective Date Status"] = compare_effective_date_status(
                _r.get("Effective Date (M61)"),
                _r.get("Effective Date (ACP)"),
            )
        if "Pledge Date Status" in out.columns:
            out.at[idx, "Pledge Date Status"] = compare_pledge_date_status(
                val_liability=_r.get("Pledge Date (M61)"),
                val_acp=_r.get("Pledge Date (ACP)"),
            )
        if "Advance Rate Status" in out.columns:
            _adv_m61 = _r.get("Advance Rate (M61)")
            _adv_acp = _r.get("Advance Rate (ACP)")
            _na = _coerce_rate_fraction(_adv_m61)
            _nb = _coerce_rate_fraction(_adv_acp)
            if _na is not None and _nb is not None:
                out.at[idx, "Advance Rate Status"] = (
                    FIELD_STATUS_MATCH if abs(_na - _nb) <= RATE_STATUS_TOLERANCE else FIELD_STATUS_MISMATCH
                )
            else:
                out.at[idx, "Advance Rate Status"] = compare_liability_primary_status(
                    _adv_m61, _adv_acp, "numeric"
                )
        if "Spread Status" in out.columns:
            _sp_m61 = _r.get("Spread (M61)")
            _sp_acp = _r.get("Spread (ACP)")
            _l_ok = _has_compare_value(_sp_m61)
            _p_ok = _has_compare_value(_sp_acp)
            if not _l_ok and not _p_ok:
                _sp_status = FIELD_STATUS_MISSING_BOTH
            elif not _p_ok and _l_ok:
                _sp_status = FIELD_STATUS_MISSING_ACORE
            elif _p_ok and not _l_ok:
                _sp_status = FIELD_STATUS_MISSING_M61
            else:
                _qa = _spread_percent_quantized_m61_compare(_sp_m61)
                _qb = _spread_percent_quantized_m61_compare(_sp_acp)
                if _qa is not None and _qb is not None:
                    _sp_status = FIELD_STATUS_MATCH if _qa == _qb else FIELD_STATUS_MISMATCH
                else:
                    _sp_status = compare_liability_primary_status(_sp_m61, _sp_acp, "numeric")
            out.at[idx, "Spread Status"] = _sp_status
        if "Undrawn Capacity Status" in out.columns:
            out.at[idx, "Undrawn Capacity Status"] = compare_liability_primary_status(
                _r.get("Undrawn Capacity (M61)"),
                _r.get("Undrawn Capacity (ACP)"),
                "numeric",
            )
        if "Index Floor Status" in out.columns:
            out.at[idx, "Index Floor Status"] = compare_optional(
                _r.get("Index Floor (M61)"),
                _r.get("Index Floor (ACP)"),
                "text",
            )
        if "Index Name Status" in out.columns:
            out.at[idx, "Index Name Status"] = compare_optional(
                _r.get("Index Name (M61)"),
                _r.get("Index Name (ACP)"),
                "text",
            )
        if "Recourse % Status" in out.columns:
            out.at[idx, "Recourse % Status"] = compare_optional(
                _r.get("Recourse % (M61)"),
                _r.get("Recourse % (ACP)"),
                "numeric",
            )

        _final_liab_type = (
            out.at[idx, "Liability Type (M61 Raw)"]
            if "Liability Type (M61 Raw)" in out.columns
            else pd.NA
        )
        _final_note_cat = out.at[idx, "M61 Note Category"] if "M61 Note Category" in out.columns else pd.NA
        _debug_rows(
            "Related M61 liability-type mapping: "
            f"raw_selected_m61_liability_type={pick.get('Liability Type')!r} "
            f"raw_selected_m61_liability_note={pick.get('Liability Note')!r} "
            f"raw_selected_m61_effective_date={pick.get('Effective Date')!r} "
            f"raw_selected_m61_pledge_date={pick.get('Pledge Date')!r} "
            f"final_output_liability_type_m61={_final_liab_type!r} "
            f"final_output_m61_note_category={_final_note_cat!r} "
            f"final_output_eff_date_m61={out.at[idx, 'Effective Date (M61)'] if 'Effective Date (M61)' in out.columns else pd.NA!r} "
            f"final_output_pledge_date_m61={out.at[idx, 'Pledge Date (M61)'] if 'Pledge Date (M61)' in out.columns else pd.NA!r} "
            f"final_output_spread_m61={out.at[idx, 'Spread (M61)'] if 'Spread (M61)' in out.columns else pd.NA!r} "
            f"final_output_index_floor_m61={out.at[idx, 'Index Floor (M61)'] if 'Index Floor (M61)' in out.columns else pd.NA!r}"
        )

        n_done += 1
        _debug_rows(
            "Related M61 advance-rate display: "
            f"Source_Type_ACORE={ar_meta['source_acore']!r} "
            f"M61_TargetAdvanceRate={ar_meta['m61_target_advance']!r} "
            f"M61_DealLevelAdvanceRate_raw={ar_meta['m61_raw_deallevel_advance_rate']!r} "
            f"M61_DealLevelAdvanceRate_chain={ar_meta['m61_deal_level_chain']!r} "
            f"use_deal_level={ar_meta['use_deal_level_adv']!r} "
            f"selected_source={ar_meta['selected_source']!r} "
            f"Advance_Rate_M61={ar_meta['advance_rate_m61']!r}"
        )
        if primary_file_type == "AOC I" and _trace:
            tgt_r = pick.get("Target Advance Rate") if "Target Advance Rate" in pick.index else pd.NA
            cur_r = pick.get("Current Advance Rate") if "Current Advance Rate" in pick.index else pd.NA
            dlr_r = _pick_deal_level_advance_rate(pick)
            raw_sp = _m61_spread_raw_from_liab_pick(pick)
            raw_ixf = _m61_index_floor_raw_from_liab_pick(pick)
            _debug_rows(
                "Related M61 AOC I display trace: related_m61_row_found=yes "
                f"m61_row_id={rid_pick} pick_reason={pick_reason} "
                f"deal={row.get('Deal Name')!r} eff_acp={row.get('Effective Date (ACP)')!r} "
                f"raw_m61 TargetAdv={tgt_r!r} CurrAdv={cur_r!r} DealLevelAdv={dlr_r!r} "
                f"raw_m61 Spread={raw_sp!r} IndexFloor={raw_ixf!r} "
                f"written Advance Rate (M61)={_upd.get('Advance Rate (M61)')!r} "
                f"Spread (M61)={_upd.get('Spread (M61)')!r} Index Floor (M61)={_upd.get('Index Floor (M61)')!r} "
                f"skip_or_blank_reasons={skip_reasons if skip_reasons else 'none'}"
            )
        elif _trace or "LN_Fin_AOCI" in str(pick.get("Liability Note", "")):
            _debug_rows(
                "TEMP DEBUG: related M61 ATTACHED — "
                f"acp_row_idx={idx} pick_reason={pick_reason} m61_row_id={rid_pick} "
                f"m61_match_key={pick.get('m61_match_key')!r} m61_eff={pick.get('Effective Date')!r} "
                f"m61_ln={str(pick.get('Liability Note') or '')[:100]!r}"
            )

    if n_done:
        _debug_rows(
            f"TEMP DEBUG: related M61 display surfacing applied to {n_done} ACORE-only row(s)"
        )
    n_after = len(out)
    removed_or_merged: list[str] = []
    if n_after != n_before:
        removed_or_merged.append(f"ROW_COUNT_CHANGED:{n_before}->{n_after}")
        _debug_rows(
            "Related M61 enhancement: unexpected row count change (should be display-only)"
        )
    _debug_rows(
        f"Related M61 enhancement: row count AFTER={n_after} "
        f"removed_or_merged={removed_or_merged if removed_or_merged else 'none'}"
    )
    return out


def _ensure_file_source_populated(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure ``File Source`` is never NA/blank before returning from ``reconcile()``."""
    if df is None or df.empty or "File Source" not in df.columns:
        return df
    cur = df["File Source"]
    missing = cur.isna() | (
        cur.astype(str).str.strip().isin(("", "nan", "NaN", "None", "<NA>", "<na>"))
    )
    if not missing.any():
        return df
    if "ID Match Result" not in df.columns:
        return df
    imr = df["ID Match Result"].astype(str).str.strip().str.lower()
    mapping: tuple[tuple[str, str], ...] = (
        ("both", FILE_SOURCE_BOTH),
        ("left_only", FILE_SOURCE_ACORE_ONLY),
        ("right_only", FILE_SOURCE_M61_ONLY),
    )
    for key, label in mapping:
        mask = missing & imr.eq(key)
        if mask.any():
            df.loc[mask, "File Source"] = label
    return df


ACP_SHEET_COLS = [
    "Deal Name",
    "Note Name",
    "Source",
    "Facility",
    "Advance Rate",
    "Spread",
    "Pledge",
    "Pledge Date",
    "Effective Date",
    "Current Undrawn Capacity",
    "Maturity Date",
    "Floor",
    "Recourse %",
]

# Pledge / Pledge Date handled separately (explicit statuses); not in this list.
COMPARE_FIELDS = [
    ("Advance Rate", "Current Advance Rate", "numeric"),
    ("Spread", "Spread", "numeric"),
    ("Effective Date", "Effective Date", "date"),
    ("Current Undrawn Capacity", "Undrawn Capacity", "numeric"),
]

RECON_STATUS_FIELDS = frozenset({"Advance Rate", "Spread", "Effective Date"})
ROW_LEVEL_STATUS_LABELS = (
    ("Advance Rate Status", "Advance Rate"),
    ("Spread Status", "Spread"),
    ("Effective Date Status", "Effective Date"),
    ("Undrawn Capacity Status", "Undrawn Capacity"),
    ("Index Floor Status", "Index Floor"),
    ("Index Name Status", "Index Name"),
    ("Recourse % Status", "Recourse %"),
    ("Pledge Date Status", "Pledge Date"),
)

LIABILITY_ADVANCE_RATE_COLUMNS = ("Current Advance Rate", "Advance Rate", "Advance")


def _status_issue_kind(v: object) -> str:
    s = "" if v is None else str(v).strip().upper()
    if not s:
        return ""
    # "MISSING FROM BOTH" is noise — absent on both sides, not a real discrepancy.
    # Must be checked before the general "MISSING" branch.
    if "MISSING FROM BOTH" in s:
        return "missing_both"
    if "MISSING" in s:
        return "missing"
    if "MISMATCH" in s or "DIFFERENCE" in s or "NO MATCH" in s:
        return "mismatch"
    if "MATCH" in s:
        return "match"
    return ""


def _derive_business_recon_status(
    record: dict[str, object],
    *,
    in_a: bool,
    in_b: bool,
    primary_file_type: str,
) -> str:
    """Business-facing row status label (summary only; does not change matching/pairing)."""
    if in_b and not in_a:
        return "MISSING IN M61"
    if in_a and not in_b:
        return f"MISSING IN {scope_label_for_primary_type(primary_file_type)}"
    if not in_a and not in_b:
        return "MISSING"

    mismatch_fields: list[str] = []
    missing_fields: list[str] = []
    for status_col, field_label in ROW_LEVEL_STATUS_LABELS:
        if status_col not in record:
            continue
        kind = _status_issue_kind(record.get(status_col))
        if kind == "mismatch":
            mismatch_fields.append(field_label)
        elif kind == "missing":
            # Only one-sided missing values are meaningful issues.
            # "missing_both" is intentionally ignored — not flagged to the user.
            missing_fields.append(field_label)

    if not mismatch_fields and not missing_fields:
        return "MATCH"
    if missing_fields and not mismatch_fields:
        return "MATCH WITH MISSING FIELDS: " + ", ".join(missing_fields)
    if mismatch_fields and not missing_fields:
        return "MATCH WITH DIFFERENCES: " + ", ".join(mismatch_fields)
    # Both mismatches and one-sided missing fields present — call out both.
    return (
        "MATCH WITH DIFFERENCES: "
        + ", ".join(mismatch_fields)
        + "; MISSING FIELDS: "
        + ", ".join(missing_fields)
    )

LIABILITY_VALUE_LABELS = {
    "Current Advance Rate": "Advance Rate (M61)",
    "Spread": "Spread (M61)",
    "Pledge": "Pledge (M61)",
    "Pledge Date": "Pledge Date (M61)",
    "Effective Date": "Effective Date (M61)",
    "Undrawn Capacity": "Current Undrawn Capacity (M61)",
    "Current Balance": "Current Balance (M61)",
    "Liability Note": "Liability Note (M61)",
    "Financial Institution": "Financial Institution (M61)",
    "Maturity Date": "Maturity Date (M61)",
    "Fund Name": "Fund (M61)",
    "Liability Name": "Liability Name (M61)",
    "target": "Target (M61)",
    "Status": "Status (M61)",
}

# Extra Liability_Relationship columns to load (NA if absent).
LIABILITY_SOURCE_EXTRA_COLS = [
    "IndexFloor",
    # Some M61 exports use ``Floor`` for the index floor (same role as ``IndexFloor`` on matched rows).
    "Floor",
    "IndexName",
    "Recourse %",
    "Recourse",
    "RecoursePct",
]

RECON_ORDERED_COLS = [
    "Fund",
    "Deal Name",
    "Facility",
    "Financial Line",
    "Note Name",
    "Liability Note (M61)",
    "M61 Note Category",
    "Deal ID (ACP)",
    "Deal ID Match Key (ACP)",
    "Liability Note Suffix (M61)",
    "M61 Extracted Deal ID",
    "ID Match Result",
    "Match Stage",
    "Source",
    "File Source",
    "Liability Name (M61)",
    "Facility Norm (ACP)",
    "Facility Norm (M61)",
    "match_key",
    "Effective Date (ACP)",
    "Effective Date (M61)",
    "Pledge Date (ACP)",
    "Pledge Date (M61)",
    "Advance Rate (ACP)",
    "Advance Rate (M61)",
    "Advance Rate Source (M61)",
    "Current Advance Rate (M61 Raw)",
    "Deal Level Advance Rate (M61 Raw)",
    "Raw Target Advance Rate from M61",
    "Raw Current Advance Rate from M61",
    "Raw Deal Level Advance Rate from M61",
    "Final Advance Rate (M61)",
    "Liability Type (M61 Raw)",
    "Target Advance Rate (M61)",
    "Spread (ACP)",
    "Spread (M61)",
    "Undrawn Capacity (ACP)",
    "Undrawn Capacity (M61)",
    "Index Floor (ACP)",
    "Index Floor (M61)",
    "Index Name (ACP)",
    "Index Name (M61)",
    "Recourse % (ACP)",
    "Recourse % (M61)",
    "Advance Rate (ACP) Debug",
    "Effective Date Status",
    "Pledge Date Status",
    "Advance Rate Status",
    "Spread Status",
    "Undrawn Capacity Status",
    "Index Floor Status",
    "Index Name Status",
    "Recourse % Status",
    "recon_status",
]

FACILITY_NORM_MAP = {
    "jpm repo": "jpm",
    "gs repo": "gs",
    "ms repo": "ms",
    "boa repo": "boa",
    "tbd": "tbd",
    "acpiii-jpm-repo": "jpm",
    "acpiii-gs-repo": "gs",
    "acpiii-ms-repo": "ms",
    "acpiii-boa-repo": "boa",
    "acpii-jpm-repo": "jpm",
    "acpii-gs-repo": "gs",
    "acpii-ms-repo": "ms",
    "acpii-boa-repo": "boa",
    "aocii-jpm-repo": "jpm",
    "aocii-gs-repo": "gs",
    "aocii-ms-repo": "ms",
    "aocii-boa-repo": "boa",
    "aoci-jpm-repo": "jpm",
    "aoci-gs-repo": "gs",
    "aoci-ms-repo": "ms",
    "aoci-boa-repo": "boa",
    "acpi-jpm-repo": "jpm",
    "acpi-gs-repo": "gs",
    "acpi-ms-repo": "ms",
    "acpi-boa-repo": "boa",
    # Plain financial-institution name aliases (no fund prefix / no "-Repo" suffix).
    # Required so ACORE Facility strings like "JP Morgan" produce the same token ("jpm")
    # as the M61 note suffix "_JPM", enabling financing_note stage matching.
    "jp morgan": "jpm",
    "j.p. morgan": "jpm",
    "jpmorgan": "jpm",
    "jpm": "jpm",
    "jp morgan repo": "jpm",
    "j.p. morgan repo": "jpm",
    "goldman sachs": "gs",
    "goldman sachs repo": "gs",
    "bank of america": "boa",
    "bofa": "boa",
    "morgan stanley": "ms",
    "morgan stanley repo": "ms",
}

PRIMARY_INTERNAL_FIELDS = (
    "deal_name",
    "note_name",
    "source",
    "facility",
    "advance_rate",
    "spread",
    "pledge",
    "pledge_date",
    "effective_date",
    "undrawn_capacity",
    "maturity_date",
    "floor",
    "recourse_pct",
)

INTERNAL_FIELD_TO_OUTPUT_COL = {
    "deal_name": "Deal Name",
    "note_name": "Note Name",
    "source": "Source",
    "facility": "Facility",
    "advance_rate": "Advance Rate",
    "spread": "Spread",
    "pledge": "Pledge",
    "pledge_date": "Pledge Date",
    "effective_date": "Effective Date",
    "undrawn_capacity": "Current Undrawn Capacity",
    "maturity_date": "Maturity Date",
    "floor": "Floor",
    "recourse_pct": "Recourse %",
}


class PrimaryFileSchemaError(ValueError):
    def __init__(self, primary_type: str, missing: list[str]):
        self.primary_type = primary_type
        self.missing = missing
        detail = "; ".join(missing) if missing else "unknown"
        super().__init__(f"[{primary_type}] Missing required column(s): {detail}")


PRIMARY_FILE_CONFIG: dict[str, dict] = {
    "ACORE": {
        "sheet_name": "10) Fin Inpt",
        "header_row": 6,
        "column_map": {
            "deal_name": "Deal Name",
            "note_name": "Note Name",
            "source": "Source",
            "facility": "Facility",
            "advance_rate": "Advance Rate",
            "spread": "Spread",
            "pledge": "Pledge",
            "pledge_date": "Pledge Date",
            "effective_date": "Effective Date",
            "undrawn_capacity": "Current Undrawn Capacity",
            "maturity_date": "Maturity Date",
            "floor": "Floor",
            "recourse_pct": "Recourse %",
        },
        "display_name": "ACORE",
        "ui_display_label": "ACORE - ACP III",
        "model_descriptor": "ACORE Liquidity & Earnings Model",
        "source_indicator_primary_only": "ACORE",
        "missing_in_primary_label": "ACORE",
        "excel_primary_column_suffix": "ACORE",
        "primary_only_legend_label": "ACORE Only",
        "primary_group_header": "ACORE — PRIMARY DATA",
    },
    "AOC II": {
        # Same tab/layout family as ACP II / ACORE; advance is typically labeled "Advance"
        # (load_primary_file synthesizes "Advance Rate" when needed).
        "sheet_name": "10) Fin Inpt",
        "header_row": 6,
        "sanitize_fin_inpt_headers": True,
        "column_map": {
            "deal_name": "Deal Name",
            "note_name": "Note Name",
            "source": "Source",
            "facility": "Facility",
            "advance_rate": "Advance",
            "spread": "Spread",
            "pledge": "Pledge",
            "pledge_date": "Pledge Date",
            "effective_date": "Effective Date",
            "undrawn_capacity": "Current Undrawn Capacity",
            "maturity_date": "Maturity Date",
            "floor": "Floor",
            "recourse_pct": "Recourse %",
        },
        "display_name": "AOC II",
        "ui_display_label": "ACORE - AOC II",
        "model_descriptor": "AOC II Liquidity & Earnings Model",
        "source_indicator_primary_only": "AOC II",
        "missing_in_primary_label": "AOC II",
        "excel_primary_column_suffix": "AOC II",
        "primary_only_legend_label": "AOC II Only",
        "primary_group_header": "AOC II — PRIMARY DATA",
    },
    "ACP II": {
        # Layout verified on "ACP II - Liquidity & Earnings Model - 2026-03-20 - v1.xlsm".
        # Data tab: "10) Fin Inpt" (same name as ACORE/ACP III). pandas header_row=6 = Excel row 7.
        # Present on that sheet: Deal Name, Note Name, Effective Date, Pledge Date, Advance,
        # Source, Facility, Spread, Floor, Recourse % (plus fees / IDs). Advance → Advance Rate
        # is handled in load_primary_file. Not on this sheet: Pledge, Current Undrawn Capacity,
        # Maturity Date — those are added as NA via ensure_columns (recon still runs).
        "sheet_name": "10) Fin Inpt",
        "header_row": 6,
        "sanitize_fin_inpt_headers": True,
        "column_map": {
            "deal_name": "Deal Name",
            "note_name": "Note Name",
            "source": "Source",
            "facility": "Facility",
            "advance_rate": "Advance",
            "spread": "Spread",
            "pledge": "Pledge",
            "pledge_date": "Pledge Date",
            "effective_date": "Effective Date",
            "undrawn_capacity": "Current Undrawn Capacity",
            "maturity_date": "Maturity Date",
            "floor": "Floor",
            "recourse_pct": "Recourse %",
        },
        "display_name": "ACP II",
        "ui_display_label": "ACORE - ACP II",
        "model_descriptor": "ACP II Liquidity & Earnings Model",
        "source_indicator_primary_only": "ACP II",
        "missing_in_primary_label": "ACP II",
        "excel_primary_column_suffix": "ACP II",
        "primary_only_legend_label": "ACP II Only",
        "primary_group_header": "ACP II — PRIMARY DATA",
    },
    "ACP I": {
        "sheet_name": "10) Fin Inpt",
        "header_row": 6,
        "sanitize_fin_inpt_headers": True,
        "column_map": {
            "deal_name": "Deal Name",
            "note_name": "Note Name",
            "source": "Source",
            "facility": "Facility",
            "advance_rate": "Advance",
            "spread": "Spread",
            "pledge": "Pledge",
            "pledge_date": "Pledge Date",
            "effective_date": "Effective Date",
            "undrawn_capacity": "Current Undrawn Capacity",
            "maturity_date": "Maturity Date",
            "floor": "Floor",
            "recourse_pct": "Recourse %",
        },
        "display_name": "ACP I",
        "ui_display_label": "ACORE - ACP I",
        "model_descriptor": "ACP I Liquidity & Earnings Model",
        "source_indicator_primary_only": "ACP I",
        "missing_in_primary_label": "ACP I",
        "excel_primary_column_suffix": "ACP I",
        "primary_only_legend_label": "ACP I Only",
        "primary_group_header": "ACP I — PRIMARY DATA",
    },
    "AOC I": {
        "sheet_name": "9) Fin Inpt",
        "header_row": 6,
        "sanitize_fin_inpt_headers": True,
        "column_map": {
            "deal_name": "Deal Name",
            "note_name": "Note Name",
            "source": "Source",
            "facility": "Facility",
            "advance_rate": "Advance",
            "spread": "Spread",
            "pledge": "Pledge",
            "pledge_date": "Pledge Date",
            "effective_date": "Effective Date",
            "undrawn_capacity": "Current Undrawn Capacity",
            "maturity_date": "Maturity Date",
            "floor": "Floor",
            "recourse_pct": "Recourse %",
        },
        "display_name": "AOC I",
        "ui_display_label": "ACORE - AOC I",
        "model_descriptor": "AOC I Liquidity & Earnings Model",
        "source_indicator_primary_only": "AOC I",
        "missing_in_primary_label": "AOC I",
        "excel_primary_column_suffix": "AOC I",
        "primary_only_legend_label": "AOC I Only",
        "primary_group_header": "AOC I — PRIMARY DATA",
    },
}

STREAMLIT_PRIMARY_FILE_TYPES = ("ACP I", "ACP II", "ACORE", "AOC I", "AOC II")

PRIMARY_REQUIRED_IN_SHEET = ("deal_name", "note_name", "facility", "effective_date")


def get_primary_config(primary_file_type: str) -> dict:
    if primary_file_type not in PRIMARY_FILE_CONFIG:
        known = ", ".join(sorted(PRIMARY_FILE_CONFIG))
        raise ValueError(
            f"Unknown primary file type {primary_file_type!r}. Expected one of: {known}"
        )
    return PRIMARY_FILE_CONFIG[primary_file_type]


def ensure_columns(df: pd.DataFrame, columns) -> pd.DataFrame:
    for col in columns:
        if col not in df.columns:
            df[col] = pd.NA
    return df


def _norm_colname_key(name) -> str:
    if name is None:
        return ""
    s = str(name)
    s = s.replace("\n", " ").replace("\r", " ").replace("\t", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def _normalize_missing_placeholders(series: pd.Series) -> pd.Series:
    """Treat common source placeholders for missing values as NA (display/compare should not see them as 0)."""
    s = series.copy()
    s = s.where(~s.isna(), pd.NA)
    s_str = s.astype(str).str.strip()
    missing_tokens = {"", "-", "—", "–", "−", "nan", "none", "<na>", "nat", "null"}
    return s.where(~s_str.str.lower().isin(missing_tokens), pd.NA)


def _canonicalize_m61_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize M61 export column headers to expected canonical names."""
    out = df.copy()
    alias_to_canonical = {
        "fund name": "Fund Name",
        "liability name": "Liability Name",
        "liability type": "Liability Type",
        "financial institution": "Financial Institution",
        "deal name": "Deal Name",
        "liability note": "Liability Note",
        "status": "Status",
        "pledge": "Pledge",
        "pledge date": "Pledge Date",
        "effective date": "Effective Date",
        "maturity date": "Maturity Date",
        "current advance rate": "Current Advance Rate",
        "target advance rate": "Target Advance Rate",
        "dealleveladvancerate": "DealLevelAdvanceRate",
        "deal level advance rate": "DealLevelAdvanceRate",
        "current balance": "Current Balance",
        "undrawn capacity": "Undrawn Capacity",
        "spread": "Spread",
        "index floor": "IndexFloor",
        "indexfloor": "IndexFloor",
        "target": "target",
        "in_liability": "in_liability",
    }
    ren = {}
    seen_targets = set(out.columns)
    for c in out.columns:
        k = _norm_colname_key(c)
        tgt = alias_to_canonical.get(k)
        if tgt and c != tgt and tgt not in seen_targets:
            ren[c] = tgt
            seen_targets.add(tgt)
    if ren:
        out = out.rename(columns=ren)
    return out


def _liability_type_bucket(value) -> str:
    s = normalise_text(value)
    if not s:
        return ""
    if "subline" in s:
        return "subline"
    if "repo" in s:
        if "non" in s:
            return "non"
        return "repo"
    if s in ("non", "non-repo", "non repo") or s.startswith("non "):
        return "non"
    return s


# ACP II "10) Fin Inpt" can pick up merged-area junk columns (float headers, "0.0025.1", etc.).
_FIN_INPT_NUMERIC_LIKE_HEADER_RE = re.compile(r"^[\d.]+$")


def _sanitize_fin_inpt_raw_df(df: pd.DataFrame, cfg: dict) -> pd.DataFrame:
    if not cfg.get("sanitize_fin_inpt_headers"):
        return df
    keep = []
    for c in df.columns:
        if not isinstance(c, str):
            continue
        s = str(c).strip()
        if _FIN_INPT_NUMERIC_LIKE_HEADER_RE.fullmatch(s):
            continue
        keep.append(c)
    return df.loc[:, keep].copy()


def inspect_primary_workbook(path: str, primary_file_type: str) -> None:
    """
    Print the configured sheet name, header row, and raw column headers as read by pandas.
    Use: python recon_enhanced_output.py --inspect-primary --file-b <workbook.xlsm> --primary-type ACP II
    """
    cfg = get_primary_config(primary_file_type)
    sheet = cfg["sheet_name"]
    hdr = cfg["header_row"]
    xl = pd.ExcelFile(path)
    print(f"primary_file_type={primary_file_type!r}")
    print(f"workbook={path!r}")
    print(f"all_sheet_names ({len(xl.sheet_names)}):")
    for i, name in enumerate(xl.sheet_names):
        mark = " <-- configured" if name == sheet else ""
        print(f"  [{i:3d}] {name!r}{mark}")
    if sheet not in xl.sheet_names:
        print(f"ERROR: configured sheet_name {sheet!r} not found.", file=sys.stderr)
        return
    df0 = pd.read_excel(path, sheet_name=sheet, header=hdr, nrows=0)
    print(f"\nUsing sheet_name={sheet!r}  header_row={hdr} (pandas read_excel header=)")
    print(f"raw column count: {len(df0.columns)}")
    for i, c in enumerate(df0.columns):
        print(f"  [{i:2d}] {c!r}  ({type(c).__name__})")
    if cfg.get("sanitize_fin_inpt_headers"):
        df1 = _sanitize_fin_inpt_raw_df(df0, cfg)
        print(f"\nAfter sanitize_fin_inpt_headers: {len(df1.columns)} columns")
        for i, c in enumerate(df1.columns):
            print(f"  [{i:2d}] {c!r}")


def _missing_primary_columns(df: pd.DataFrame, cfg: dict) -> list[str]:
    missing = []
    cmap = cfg["column_map"]
    cols = set(df.columns.astype(str))
    for key in PRIMARY_REQUIRED_IN_SHEET:
        src = cmap.get(key)
        if not src:
            missing.append(f"{key}: no mapping configured")
        elif src not in cols:
            missing.append(f"{key}: expected column {src!r} (not found in sheet)")
    return missing


def safe_parse_date(series):
    s = pd.Series(series)
    out = pd.to_datetime(s, errors="coerce", dayfirst=False).dt.normalize()

    missing = out.isna() & s.notna()
    if missing.any():
        sub = s[missing].astype(str).str.strip()
        for fmt in ("%m/%d/%y", "%m-%d-%y", "%m/%d/%Y", "%m-%d-%Y"):
            retry = pd.to_datetime(sub, format=fmt, errors="coerce").dt.normalize()
            take = retry.notna()
            if take.any():
                out.loc[sub.index[take]] = retry.loc[sub.index[take]]
            missing = out.isna() & s.notna()
            if not missing.any():
                break

    missing = out.isna() & s.notna()
    if missing.any():
        num = pd.to_numeric(s[missing], errors="coerce")
        ok = num.notna() & (num > 20000) & (num < 1_000_000)
        if ok.any():
            ex = pd.to_datetime(num[ok], unit="D", origin="1899-12-30", errors="coerce").dt.normalize()
            out.loc[ex.index] = ex

    return out


def _drop_hidden_fin_inpt_rows(df_raw: pd.DataFrame, path: str, cfg: dict, primary_file_type: str) -> pd.DataFrame:
    """AOC Fin Inpt guard: exclude Excel rows that are hidden/filtered in the source sheet."""
    if primary_file_type not in ("AOC II", "AOC I"):
        return df_raw
    try:
        wb = load_workbook(path, read_only=False, data_only=True)
        ws = wb[cfg["sheet_name"]]
        header_row = int(cfg["header_row"])  # 0-based pandas header row
        first_data_excel_row = header_row + 2  # 1-based Excel row number
        keep_mask = []
        for i in range(len(df_raw)):
            excel_row = first_data_excel_row + i
            hidden = bool(ws.row_dimensions[excel_row].hidden)
            keep_mask.append(not hidden)
        kept = int(sum(1 for x in keep_mask if x))
        if kept != len(df_raw):
            _debug_rows(
                "Primary hidden-row filter applied "
                f"primary_type={primary_file_type} kept_rows={kept}/{len(df_raw)} sheet={cfg['sheet_name']!r}"
            )
        return df_raw.loc[keep_mask].reset_index(drop=True)
    except Exception as e:
        _debug_rows(f"Primary hidden-row filter skipped ({primary_file_type}): {e}")
        return df_raw


def load_primary_file(path: str, primary_file_type: str) -> pd.DataFrame:
    cfg = get_primary_config(primary_file_type)
    df_raw = pd.read_excel(path, sheet_name=cfg["sheet_name"], header=cfg["header_row"])
    df_raw = _sanitize_fin_inpt_raw_df(df_raw, cfg)
    df_raw = _drop_hidden_fin_inpt_rows(df_raw, path, cfg, primary_file_type)

    if "Advance" in df_raw.columns and "Advance Rate" not in df_raw.columns:
        df_raw["Advance Rate"] = df_raw["Advance"]
    if "Undrawn Capacity" in df_raw.columns and "Current Undrawn Capacity" not in df_raw.columns:
        df_raw["Current Undrawn Capacity"] = df_raw["Undrawn Capacity"]

    miss = _missing_primary_columns(df_raw, cfg)
    if miss:
        raise PrimaryFileSchemaError(primary_file_type, miss)

    cmap = cfg["column_map"]
    all_src = list(dict.fromkeys(cmap[k] for k in PRIMARY_INTERNAL_FIELDS))
    df_raw = ensure_columns(df_raw, all_src)

    use_cols = [cmap[k] for k in PRIMARY_INTERNAL_FIELDS]
    renamer = {cmap[k]: INTERNAL_FIELD_TO_OUTPUT_COL[k] for k in PRIMARY_INTERNAL_FIELDS}
    df = df_raw.loc[:, use_cols].rename(columns=renamer)
    deal_id_col = None
    if "Deal ID" in df_raw.columns:
        deal_id_col = "Deal ID"
    else:
        # Header-tolerant fallback (spaces/case variants).
        deal_id_aliases = {"deal id", "dealid"}
        for c in df_raw.columns:
            if _norm_colname_key(c).replace(" ", "") in {"dealid"} or _norm_colname_key(c) in deal_id_aliases:
                deal_id_col = c
                break
    if deal_id_col is not None:
        df["Deal ID"] = df_raw[deal_id_col]

    df = ensure_columns(df, ACP_SHEET_COLS)
    keep_cols = list(ACP_SHEET_COLS) + (["Deal ID"] if "Deal ID" in df.columns else [])
    df = df[keep_cols].copy().dropna(subset=["Deal Name"])

    for col in ["Deal Name", "Note Name", "Source", "Facility", "Pledge"]:
        df[col] = df[col].astype(str).str.strip()
    if "Deal ID" in df.columns:
        df["Deal ID"] = df["Deal ID"].apply(
            lambda v: pd.NA if pd.isna(v) or not str(v).strip() else str(v).strip()
        )

    df["Pledge Date"] = safe_parse_date(df["Pledge Date"])
    df["Effective Date"] = safe_parse_date(df["Effective Date"])
    df["Maturity Date"] = safe_parse_date(df["Maturity Date"])
    return df.reset_index(drop=True)


def _fin_note_tokens_for_primary(primary_file_type: str) -> tuple[str, ...]:
    """Liability-note substrings that flag financing lines for the selected fund (M61 load filter OR)."""
    p = str(primary_file_type or "").strip().upper()
    if p == "ACORE":
        return (
            "LN_FIN_ACPIII",
            "LN_FIN_ACP_III",
            "LN_FIN_ACP3",
            "LN_FIN_ACP_3",
            "ACPIII",
        )
    if p == "ACP II":
        return (
            "LN_FIN_ACPII",
            "LN_FIN_ACP_II",
            "LN_FIN_ACP2",
            "LN_FIN_ACP_2",
            "ACPII",
        )
    if p == "ACP I":
        return (
            "LN_FIN_ACP_I_",
            "LN_FIN_ACP_I",
            "LN_FIN_ACP1",
            "LN_FIN_ACP_1",
            "ACP I",
        )
    if p == "AOC II":
        return (
            "LN_FIN_AOCII",
            "LN_FIN_AOC_II",
            "LN_FIN_AOC2",
            "LN_FIN_AOC_2",
            "AOCII",
            "AOC II",
        )
    if p == "AOC I":
        return (
            "LN_FIN_AOCI_",
            "LN_FIN_AOC_I",
            "LN_FIN_AOC1",
            "LN_FIN_AOC_1",
            "AOCI",
            "AOC I",
        )
    return ()


def _fin_note_scope_mask(liab_note_up: pd.Series, primary_file_type: str) -> pd.Series:
    toks = _fin_note_tokens_for_primary(primary_file_type)
    if not toks:
        return pd.Series(False, index=liab_note_up.index)
    pat = "|".join(re.escape(t) for t in toks)
    return liab_note_up.astype(str).str.contains(pat, case=False, regex=True, na=False)


def load_file_a(
    path: str,
    *,
    primary_file_type: str = "ACORE",
    return_excluded: bool = False,
    return_diagnostics: bool = False,
):
    preferred_sheet = "Liability_Relationship"
    xl = pd.ExcelFile(path)
    sheet_name = preferred_sheet
    if preferred_sheet not in xl.sheet_names:
        # Fallback for exports that use spaces/casing variants.
        fallback = None
        for sn in xl.sheet_names:
            sn_key = _norm_colname_key(sn).replace(" ", "_")
            if "liability" in sn_key and "relationship" in sn_key:
                fallback = sn
                break
        if fallback:
            sheet_name = fallback
        else:
            sheet_name = xl.sheet_names[0]
    df = pd.read_excel(path, sheet_name=sheet_name)
    df = _canonicalize_m61_columns(df)
    m61_raw_rows_loaded = int(len(df))
    _debug_rows(
        f"M61 read_excel: sheet={sheet_name!r} rows={len(df)} cols={len(df.columns)} "
        f"(full sheet read; no usecols/nrows subset)"
    )
    keep = [
        "Fund Name",
        "Liability Name",
        "Liability Type",
        "Financial Institution",
        "Deal Name",
        "Liability Note",
        "Status",
        "Pledge",
        "Pledge Date",
        "Effective Date",
        "Maturity Date",
        "Current Advance Rate",
        "Target Advance Rate",
        "Current Balance",
        "Undrawn Capacity",
        "Spread",
        "target",
        "in_liability",
    ]
    keep = list(dict.fromkeys(keep + list(LIABILITY_ADVANCE_RATE_COLUMNS) + LIABILITY_SOURCE_EXTRA_COLS))
    df = ensure_columns(df, keep)
    _debug_rows(f"M61 after ensure_columns(keep): rows={len(df)}")

    df["Fund Name"] = df["Fund Name"].astype(str).str.strip()
    fund_lower = df["Fund Name"].str.lower()
    in_target_set = fund_lower.isin(TARGET_FUNDS)
    aoc_style = df["Fund Name"].str.contains(AOC_M61_FUND_NAME_RE, regex=True, na=False) | df[
        "Fund Name"
    ].str.contains(AOC_M61_FUND_NAME_I_RE, regex=True, na=False)
    fund_mask = in_target_set | aoc_style
    df["Liability Type Bucket"] = df["Liability Type"].apply(_liability_type_bucket)
    liab_type_mask = df["Liability Type Bucket"].isin(M61_FINANCING_TYPE_BUCKETS)
    liab_note_up = df["Liability Note"].fillna("").astype(str).str.upper()
    fin_note_scope_mask = _fin_note_scope_mask(liab_note_up, primary_file_type)
    # Rows are dropped only when ALL of the following are false:
    # (1) Liability Type bucket in financing set (repo/non/sale/clo/subline), OR
    # (2) Liability Note contains primary-specific financing tokens (e.g. LN_FIN_AOCII), OR
    # (3) Fallback: Fund Name matches this run's primary fund scope (same regex as reconcile
    #     _m61_in_scope) AND Deal Name is non-blank — avoids silently dropping real fund rows
    #     that lack both a financing bucket and a parsed note prefix.
    fcfg = _fund_cfg(primary_file_type)
    fpattern = fcfg.get("fund_regex") or (
        re.escape(fcfg["fund_token"]) if fcfg.get("fund_token") else None
    )
    if fpattern:
        primary_fund_row_mask = df["Fund Name"].fillna("").astype(str).str.contains(
            fpattern, case=False, regex=True, na=False
        )
    else:
        primary_fund_row_mask = pd.Series(False, index=df.index)
    deal_name_nonempty = df["Deal Name"].fillna("").astype(str).str.strip().ne("")
    fund_deal_context_keep = primary_fund_row_mask & deal_name_nonempty
    in_scope_mask = liab_type_mask | fin_note_scope_mask | fund_deal_context_keep
    _debug_trace_uncommons_m61_load(
        df,
        "load_file_a:before_expanded_filter",
        primary_file_type=primary_file_type,
        in_scope_mask=in_scope_mask,
    )
    _fin_toks = _fin_note_tokens_for_primary(primary_file_type)
    _fallback_only = fund_deal_context_keep & ~(liab_type_mask | fin_note_scope_mask)
    _debug_rows(
        "M61 pre-filter diagnostics: "
        f"rows_matching_target_fund_labels={int(fund_mask.sum())}/{len(df)} "
        f"rows_matching_liability_types={int(liab_type_mask.sum())}/{len(df)} "
        f"rows_matching_financing_note_token={int(fin_note_scope_mask.sum())}/{len(df)} tokens={_fin_toks!r} "
        f"rows_primary_fund_regex_match={int(primary_fund_row_mask.sum())}/{len(df)} "
        f"rows_fund_deal_fallback={int(fund_deal_context_keep.sum())}/{len(df)} "
        f"rows_retained_only_by_fund_deal_fallback={int(_fallback_only.sum())} "
        f"rows_in_scope_after_expanded_filter={int(in_scope_mask.sum())}/{len(df)} "
        f"blank_deal_name={int(df['Deal Name'].isna().sum())} "
        f"blank_liability_note={int(df['Liability Note'].isna().sum())}"
    )
    excluded_by_type = df.loc[
        ~in_scope_mask,
        [
            c
            for c in (
                "Fund Name",
                "Deal Name",
                "Liability Type",
                "Liability Name",
                "Liability Note",
                "Effective Date",
            )
            if c in df.columns
        ],
    ].copy()
    excluded_by_type["Exclusion Reason"] = "Excluded by expanded M61 in-scope filter"
    ex_type_counts = (
        excluded_by_type["Liability Type"].fillna("<NA>").astype(str).value_counts().to_dict()
        if "Liability Type" in excluded_by_type.columns
        else {}
    )
    _debug_rows(
        "M61 rows excluded by expanded in-scope filter: "
        f"excluded_rows={len(excluded_by_type)} excluded_type_counts={ex_type_counts}"
    )
    if not excluded_by_type.empty:
        _debug_rows("TEMP DEBUG: sample rows excluded by expanded in-scope filter (head 10)")
        for i, (_, er) in enumerate(excluded_by_type.head(10).iterrows(), start=1):
            _debug_rows(
                "TEMP DEBUG:   "
                f"#{i} deal={er.get('Deal Name')!r} type={er.get('Liability Type')!r} "
                f"liability_name={er.get('Liability Name')!r} note={er.get('Liability Note')!r} "
                f"eff={er.get('Effective Date')!r}"
            )
    df = df[in_scope_mask].copy()
    _debug_trace_uncommons_m61_load(
        df,
        "load_file_a:after_expanded_filter",
        primary_file_type=primary_file_type,
    )
    _debug_rows(
        "M61 after expanded in-scope filter: "
        f"rows={len(df)} "
        f"(kept liability buckets={sorted(M61_FINANCING_TYPE_BUCKETS)} OR note tokens={_fin_toks!r} "
        f"OR primary-fund+nonblank-deal fallback pattern={fpattern!r})"
    )
    # Keep DealLevelAdvanceRate when the export provides it (otherwise ensure_columns would add NA).
    if "DealLevelAdvanceRate" in df.columns and "DealLevelAdvanceRate" not in keep:
        keep = list(dict.fromkeys(list(keep) + ["DealLevelAdvanceRate"]))
    df = df[keep].copy()

    # Preserve missing-vs-zero semantics for M61 numeric fields (e.g., '-' must stay missing, not 0).
    _m61_numeric_like_cols = [
        "Undrawn Capacity",
        "Current Balance",
        "Current Advance Rate",
        "Target Advance Rate",
        "DealLevelAdvanceRate",
        "Deal Level Advance Rate",
        "Advance Rate",
        "Advance",
        "Spread",
        "target",
        "IndexFloor",
        "Floor",
        "Recourse %",
        "Recourse",
        "RecoursePct",
    ]
    _raw_undrawn_preview = (
        df["Undrawn Capacity"].head(25).tolist() if "Undrawn Capacity" in df.columns else []
    )
    for _c in _m61_numeric_like_cols:
        if _c in df.columns:
            df[_c] = _normalize_missing_placeholders(df[_c])
    _stored_undrawn_preview = (
        df["Undrawn Capacity"].head(25).tolist() if "Undrawn Capacity" in df.columns else []
    )
    _debug_rows(
        "M61 Undrawn debug: "
        f"raw_before_coercion_head25={_raw_undrawn_preview!r} "
        f"stored_after_parsing_head25={_stored_undrawn_preview!r}"
    )
    if "Undrawn Capacity" in df.columns:
        # Final hard guard: coerce true numerics, preserve missing placeholders as NA (never 0).
        _uc_raw_before = df["Undrawn Capacity"].head(25).tolist()
        _uc_coerced = df["Undrawn Capacity"].map(_coerce_numeric_value)
        df["Undrawn Capacity"] = pd.Series(
            [pd.NA if v is None else v for v in _uc_coerced], index=df.index, dtype="object"
        )
        _uc_after = df["Undrawn Capacity"].head(25).tolist()
        _debug_rows(
            "M61 Undrawn debug (post hard guard): "
            f"raw_before={_uc_raw_before!r} stored_after={_uc_after!r}"
        )

    for col in [
        "Deal Name",
        "Liability Note",
        "Financial Institution",
        "Liability Name",
        "Fund Name",
        "Pledge",
    ]:
        df[col] = df[col].astype(str).str.strip()
    _debug_rows(
        "M61 post-cleaning diagnostics: "
        f"rows={len(df)} blank_deal_name={(df['Deal Name'].str.strip() == '').sum()} "
        f"blank_liability_note={(df['Liability Note'].str.strip() == '').sum()}"
    )

    df["Pledge Date"] = safe_parse_date(df["Pledge Date"])
    df["Effective Date"] = safe_parse_date(df["Effective Date"])
    df["Maturity Date"] = safe_parse_date(df["Maturity Date"])
    df = df.reset_index(drop=True)
    m61_rows_after_liability_type_filter = int(len(df))
    excluded_by_type = excluded_by_type.reset_index(drop=True)
    diagnostics = {
        "m61_sheet_used": sheet_name,
        "m61_raw_rows_loaded": m61_raw_rows_loaded,
        "m61_rows_matching_target_fund_labels": int(fund_mask.sum()),
        "m61_rows_matching_financing_note_token": int(fin_note_scope_mask.sum()),
        "m61_rows_fund_deal_context_keep": int(fund_deal_context_keep.sum()),
        "m61_rows_retained_only_by_fund_deal_fallback": int(_fallback_only.sum()),
        "m61_rows_after_liability_type_filter": m61_rows_after_liability_type_filter,
        "m61_rows_excluded_by_liability_type_filter": int(len(excluded_by_type)),
    }
    if return_excluded and return_diagnostics:
        return df, excluded_by_type, diagnostics
    if return_excluded:
        return df, excluded_by_type
    if return_diagnostics:
        return df, diagnostics
    return df


def load_file_b(path: str) -> pd.DataFrame:
    return load_primary_file(path, "ACORE")


def normalise_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip().lower()


def _alnum_compact_upper(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())


def _compact_alnum_is_fund_prefixed_tbd_only(u: str) -> bool:
    """True when compact form is optional fund slug + one or more ``TBD`` (e.g. ``ACPIIITBDTBD``)."""
    if not u or "TBD" not in u:
        return False
    rest = u
    for pref in ("ACPIII", "ACPII", "AOCII", "AOCI", "ACP3", "ACP2", "AOC2", "ACPI"):
        if rest.startswith(pref):
            rest = rest[len(pref) :]
            break
    return bool(rest) and bool(re.fullmatch(r"(TBD)+", rest))


def _normalise_dashes_for_facility(s: str) -> str:
    """Map unicode / typographic dashes to ASCII hyphen so fund-prefixed facility patterns match."""
    if not s:
        return s
    return (
        str(s)
        .replace("\u2013", "-")
        .replace("\u2014", "-")
        .replace("\u2212", "-")
        .replace("–", "-")
        .replace("—", "-")
    )


def _facility_source_collapses_to_tbd(raw) -> bool:
    """True when value is only TBD and/or ACP/AOC fund slug tokens (for facility/source matching).

    Normalizes cases such as ``TBD``, ``ACPIII-TBD``, ``ACPIII-TBD-TBD``, and compact ``ACPIIITBDTBD``
    so they align with plain ``tbd`` from Fin Inpt.
    """
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return False
    s = _normalise_dashes_for_facility(str(raw).strip())
    if not s or s.lower() in ("nan", "none", "<na>", "nat"):
        return False
    nt = normalise_text(s).replace("_", "-")
    if nt == "tbd":
        return True
    parts = [p for p in re.split(r"[\s_\-|,/]+", nt) if p]
    has_tbd_token = any(p.lower() == "tbd" for p in parts)
    if has_tbd_token:
        fund_re = re.compile(
            r"^(?:acpiii|acp\s*iii|acp3|acpii|acp\s*ii|acp2|aocii|aoc\s*ii|aoc2|aoci|aoc\s*i\b|acpi)$",
            re.IGNORECASE,
        )
        for i, p in enumerate(parts):
            pl = p.lower()
            if pl == "tbd":
                continue
            if fund_re.match(p):
                continue
            # ``AOC II - TBD`` / ``ACP I - TBD`` split as ``aoc`` ``ii`` / ``acp`` ``i`` …
            if pl in ("ii", "iii", "i") and i > 0 and parts[i - 1].lower() in ("aoc", "acp"):
                continue
            if pl in ("aoc", "acp") and i + 1 < len(parts) and parts[i + 1].lower() in ("ii", "iii", "i"):
                continue
            return False
        return True
    return _compact_alnum_is_fund_prefixed_tbd_only(_alnum_compact_upper(s))


# Strip M61 export / generated facility prefixes (``ACPIII-MS-Repo`` / ``ACPIII – MS - Repo`` → ``ms repo``).
# Allow unicode dashes (normalized upstream) and whitespace between fund slug and bank/repo tokens.
_M61_FUND_FACILITY_PREFIX = re.compile(
    r"^(?:acpiii|acp\s*iii|acp3|acpii|acp\s*ii|acp2|acpi|acp\s*i\b|aocii|aoc\s*ii|aoc2|aoci|aoc\s*i\b)"
    r"(?:\s*[-_/]+\s*|\s+)",
    re.IGNORECASE,
)


def _normalise_m61_fund_prefixed_facility(raw) -> str | None:
    """If ``raw`` is ``ACPIII-MS-Repo`` / ``ACPIII-GS-Repo`` style, return ``FACILITY_NORM_MAP`` token; else None."""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    s = _normalise_dashes_for_facility(str(raw).strip())
    if not s or s.lower() in ("nan", "none", "<na>", "nat"):
        return None
    t = normalise_text(s).replace("_", "-")
    rest = t
    while True:
        m = _M61_FUND_FACILITY_PREFIX.match(rest)
        if not m:
            break
        rest = rest[m.end() :].lstrip("-_/")
        if not rest:
            return None
    if rest == t:
        return None
    readable = " ".join(rest.replace("-", " ").replace("_", " ").split())
    if _facility_source_collapses_to_tbd(readable):
        return "tbd"
    nt2 = normalise_text(readable)
    if nt2 in FACILITY_NORM_MAP:
        return FACILITY_NORM_MAP[nt2]
    if "repo" not in nt2:
        cand = normalise_text(f"{readable} repo")
        if cand in FACILITY_NORM_MAP:
            return FACILITY_NORM_MAP[cand]
    return None


def normalise_facility(raw):
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    s = _normalise_dashes_for_facility(str(raw).strip())
    if not s or s.lower() in ("nan", "none", "<na>", "nat"):
        return ""
    pipe_parts = [p.strip() for p in re.split(r"\||,|/", s) if p.strip()]
    if len(pipe_parts) > 1:
        for p in pipe_parts:
            if _facility_source_collapses_to_tbd(p):
                return "tbd"
    if _facility_source_collapses_to_tbd(s):
        return "tbd"
    bank = _normalise_m61_fund_prefixed_facility(s)
    if bank is not None:
        return bank
    nt = normalise_text(s)
    return FACILITY_NORM_MAP.get(nt, nt)


def extract_deal_id_token(value) -> str:
    """Extract canonical deal-id token ``NN-NNNN`` from arbitrary text/value."""
    if pd.isna(value):
        return ""
    s = str(value).strip()
    if not s:
        return ""
    # Normalize common dash variants / spacing around the dash before matching.
    s_norm = (
        s.replace("–", "-")
        .replace("—", "-")
        .replace("−", "-")
    )
    # Do not rely on ``\\b`` before the first digit: deal ids can follow letters (e.g. ``…ACPIII_25-2852``).
    m = re.search(r"(?<![0-9])(\d{2})\s*-\s*(\d{4})(?![0-9])", s_norm)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    # Fallback for compact six-digit exports (e.g., 252852 -> 25-2852).
    m2 = re.search(r"(?<![0-9])(\d{2})(\d{4})(?![0-9])", s_norm)
    if m2:
        return f"{m2.group(1)}-{m2.group(2)}"
    return ""


def normalise_deal_id_key(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip().lower()


def _ln_compact_alnum(note: str) -> str:
    """Uppercase alphanumeric run (fund codes like ACPIII / ACP3 survive as one token)."""
    return _alnum_compact_upper(note)


_LN_FIN_RE = re.compile(r"LN[_\s]?FIN", re.IGNORECASE)
_LN_SUB_RE = re.compile(r"LN[_\s]?SUB", re.IGNORECASE)
_LN_EQ_RE = re.compile(r"LN[_\s]?EQ", re.IGNORECASE)

def _parse_fund_scope_from_note(note: str) -> str:
    """Return fund scope label when a fund code is present in the note, else ``\"\"``.

    Order matters: longer tokens (``ACPIII``, ``ACPII``, ``AOCII``) before Roman ``I`` forms.
    """
    u = _ln_compact_alnum(note)
    if "ACPIII" in u or "ACP3" in u:
        return "ACP III"
    if "ACPII" in u or "ACP2" in u:
        return "ACP II"
    if ("ACPI" in u or "ACP1" in u) and "ACPII" not in u and "ACPIII" not in u:
        return "ACP I"
    if "AOCII" in u or "AOC2" in u:
        return "AOC II"
    if ("AOCI" in u or "AOC1" in u) and "AOCII" not in u:
        return "AOC I"
    s = str(note or "").strip()
    if re.search(r"\bACP\s+III\b", s, re.IGNORECASE):
        return "ACP III"
    if re.search(r"\bACP\s+II\b", s, re.IGNORECASE):
        return "ACP II"
    if re.search(r"\bACP\s+I\b(?!\s*I)", s, re.IGNORECASE):
        return "ACP I"
    if re.search(r"\bAOC\s+II\b", s, re.IGNORECASE):
        return "AOC II"
    if re.search(r"\bAOC\s+I\b(?!\s*I)", s, re.IGNORECASE):
        return "AOC I"
    return ""


def _parse_suffix_after_deal_id(note: str, deal_id: str) -> str:
    """Trailing bank / facility token after the deal id (e.g. ``GS``, ``JPM``, ``BOA``, ``TBD``)."""
    if not deal_id or not str(note).strip():
        return ""
    s = str(note).strip()
    did = deal_id.strip()
    compact = did.replace("-", "")
    for token in (did, compact):
        if not token:
            continue
        pos = s.upper().find(token.upper())
        if pos < 0:
            continue
        tail = s[pos + len(token) :].strip(" _-|/\t")
        if not tail:
            return ""
        if _facility_source_collapses_to_tbd(tail):
            return "TBD"
        for part in re.split(r"[_\s|,\-/]+", tail):
            t = part.strip()
            if t and re.fullmatch(r"[A-Za-z]{2,6}", t):
                return t.upper()
        return ""
    return ""


def parse_liability_note(note) -> dict[str, str]:
    """Parse M61 **Liability Note** into category, fund, deal id, and optional source/facility suffix.

    Returns a dict with keys:

    - ``note_category``: ``Fin`` | ``Sub`` | ``Eq/Fund`` | ``Other`` (from ``LN_FIN`` / ``LN_SUB`` / ``LN_EQ``…)
    - ``fund_code``: ``ACP III`` | ``ACP II`` | ``ACP I`` | ``AOC II`` | ``AOC I`` | ``\"\"`` when absent
    - ``deal_id``: ``NN-NNNN`` token when present (same rules as ``extract_deal_id_token``)
    - ``source_suffix``: bank / facility shorthand when present (e.g. ``GS``, ``BOA``, ``TBD``)

    Works across ACP III / II / I and AOC II / I note styles (``ACPIII``, ``ACPI``, ``ACP3``, ``AOCII``, ``AOCI``, etc.).
    """
    empty = {
        "note_category": "Other",
        "fund_code": "",
        "deal_id": "",
        "source_suffix": "",
    }
    if note is None:
        return dict(empty)
    try:
        if pd.isna(note):
            return dict(empty)
    except (TypeError, ValueError):
        pass
    s = str(note).strip()
    if not s or s.lower() in ("nan", "<na>", "nat", "none"):
        return dict(empty)

    if _LN_FIN_RE.search(s):
        cat = "Fin"
    elif _LN_SUB_RE.search(s):
        cat = "Sub"
    elif _LN_EQ_RE.search(s):
        cat = "Eq/Fund"
    else:
        cat = "Other"

    deal_id = extract_deal_id_token(s)
    fund_code = _parse_fund_scope_from_note(s)
    source_suffix = _parse_suffix_after_deal_id(s, deal_id) if deal_id else ""

    return {
        "note_category": cat,
        "fund_code": fund_code,
        "deal_id": deal_id,
        "source_suffix": source_suffix,
    }


def extract_liability_note_suffix(value) -> str:
    """Extract deal-id token from Liability Note (e.g., 25-2852 anywhere in text)."""
    return parse_liability_note(value).get("deal_id", "")


# Ordered category list used by both backend column derivation and UI filter defaults.
M61_NOTE_CATEGORIES: list[str] = ["Financing", "Subline", "Equity/Fund", "Other"]

# Liability-type substrings / tokens for ``categorize_m61_note_type`` (maintenance-friendly).
_M61_LT_SUBLINE = "subline"
_M61_LT_TBD_RE = re.compile(r"\btbd\b", re.IGNORECASE)
_M61_LT_CLO_RE = re.compile(r"\bclo\b", re.IGNORECASE)
_M61_LT_EQUITY_RE = re.compile(r"\bequity\b", re.IGNORECASE)


def categorize_m61_note_type(liability_type_raw) -> str:
    """Map M61 **Liability Type** (raw export value) to sidebar / export note categories.

    Display and filter only; does not affect merge or reconciliation row counts.

    - **Financing**: Repo, Sale, Non, CLO (and TBD for legacy exports)
    - **Subline**: Subline
    - **Equity/Fund**: Equity
    - **Other**: blank, missing M61 side, or anything else

    Evaluation order: Subline → TBD / CLO / Sale / Non / Repo → Equity/Fund → Other.
    """
    if liability_type_raw is None:
        return "Other"
    try:
        if pd.isna(liability_type_raw):
            return "Other"
    except (TypeError, ValueError):
        pass
    s = normalise_text(liability_type_raw)
    if not s:
        return "Other"

    if _M61_LT_SUBLINE in s:
        return "Subline"

    if _M61_LT_TBD_RE.search(s) or _M61_LT_CLO_RE.search(s):
        return "Financing"

    if "sale" in s:
        return "Financing"

    if s in ("non", "non-repo", "non repo") or " non" in f" {s}":
        return "Financing"

    if "repo" in s:
        return "Financing"

    if _M61_LT_EQUITY_RE.search(s):
        return "Equity/Fund"

    return "Other"


def categorize_m61_note_category(
    liability_note_raw,
    liability_type_raw,
    primary_source_raw=None,
    primary_file_type: str = "",
) -> str:
    """Primary M61 Note Category source = Liability Note parser; fallback = type, then primary source.

    ``primary_file_type`` is used for fund-specific overrides:
    - AOC II: Whole Loan is NOT Financing (falls through to "Other").  All other funds
      retain the legacy behaviour where Whole Loan maps to Financing.
    """
    pft = str(primary_file_type or "").strip()
    s = normalise_text(primary_source_raw)
    # AOC I / AOC II override: Whole Loan (including WL-CPACE variants) is an "Other"
    # bucket for filtering/categorization, even when M61 Liability Type is Sale.
    if pft in ("AOC II", "AOC I") and (
        "whole loan" in s or "wholeloan" in s or "wl-cpace" in s or "wlcpace" in s
    ):
        return "Other"
    p = parse_liability_note(liability_note_raw)
    by_note = (p.get("note_category") or "").strip()
    if pft == "ACP II" and by_note == "Eq/Fund":
        # ACP II override: LN-Eq liabilities map to Fund/Equity source family but
        # are categorized as "Other" (not Financing / Equity-Fund).
        return "Other"
    if by_note == "Fin":
        return "Financing"
    if by_note == "Sub":
        return "Subline"
    if by_note == "Eq/Fund":
        return "Equity/Fund"
    by_type = categorize_m61_note_type(liability_type_raw)
    if by_type != "Other":
        return by_type
    if not s:
        return "Other"
    if "subline" in s:
        return "Subline"
    if pft == "ACP II":
        # ACP II "Other" bucket explicitly includes these Source/Type families.
        if "whole loan" in s or "wholeloan" in s or "wl-cpace" in s or "fund" in s or "equity" in s:
            return "Other"
    # For AOC II / AOC I, Whole Loan is a distinct asset class, not a financing facility —
    # skip it here so it falls through to "Other" and is excluded from Financing filter.
    if primary_file_type in ("AOC II", "AOC I") and ("whole loan" in s or "wholeloan" in s):
        return "Other"
    financing_src_tokens = ("repo", "sale", "non", "clo", "sub debt", "subdebt", "whole loan", "wholeloan")
    if any(tok in s for tok in financing_src_tokens):
        return "Financing"
    return "Other"


def _source_bucket(value, *, primary_file_type: str = "", liability_note_raw=None, is_m61: bool = False) -> str:
    s = normalise_text(value)
    pft = str(primary_file_type or "").strip()
    if is_m61 and pft == "ACP II":
        # ACP II connector override driven by M61 Liability Note prefix:
        # - LN-Sub -> Subline
        # - LN-Eq  -> Fund/Equity source family
        # - LN-Fin -> financing source family from Liability Type token
        p = parse_liability_note(liability_note_raw)
        by_note = (p.get("note_category") or "").strip()
        if by_note == "Sub":
            return "subline"
        if by_note == "Eq/Fund":
            return "fund"
        if by_note == "Fin":
            if "repo" in s:
                return "repo"
            if s in ("non", "non-repo", "non repo") or " non" in f" {s}":
                return "non"
            if "sale" in s:
                return "sale"
            if "clo" in s:
                return "clo"
            if "tbd" in s:
                return "tbd"
            # Keep non-financing/noise tokens out of ACP II financing connector matches.
            return ""
    if not s:
        return ""
    if pft == "ACP II" and ("fund" in s or "equity" in s):
        return "fund"
    if "subline" in s:
        return "subline"
    if "repo" in s:
        return "repo"
    if s in ("non", "non-repo", "non repo") or " non" in f" {s}":
        return "non"
    return s


def _is_sale_type_fund_or_deal(*, fund_name, liability_type) -> bool:
    """Business override: sale-type funds compare against deal-level advance rate."""
    fn = normalise_text(fund_name)
    lt = normalise_text(liability_type)
    # Use whole-word sale markers only (avoid false positives like "wholesale").
    sale_re = re.compile(r"\b(sale|sold)\b", re.IGNORECASE)
    return bool(sale_re.search(fn or "")) or bool(sale_re.search(lt or ""))


def _m61_deal_level_advance_rate(row: pd.Series, label_a: str):
    """Deal-level advance rate fallback chain from M61 export fields.

    Prefer ``DealLevelAdvanceRate`` (M61 export) before legacy ``Advance Rate`` / ``Advance`` names.
    """
    for col in ("DealLevelAdvanceRate", "Deal Level Advance Rate", "Advance Rate", "Advance"):
        v = row.get(f"{label_a}_{col}")
        if not _is_blank_for_compare(v):
            return v
    return pd.NA


def date_key(series):
    return safe_parse_date(series).dt.strftime("%Y-%m-%d").fillna("")


def add_m61_facility_for_match_key(df: pd.DataFrame) -> pd.DataFrame:
    """Build a facility-like string on M61 rows for ``build_match_key`` (not display).

    Primary Fin Inpt uses **Facility** (e.g. ``GS Repo``). M61 **Liability Name** is not a stable join
    key (repeats across rows). Combine **Financial Institution** + **Liability Type** so
    ``normalise_facility`` can align with the model (e.g. ``JPM`` + ``Repo`` → ``jpm repo``).
    Falls back to **Liability Name** only when FI + type would be blank.
    """
    out = df.copy()
    fi = out["Financial Institution"].fillna("").astype(str).str.strip()
    lt = out["Liability Type"].fillna("").astype(str).str.strip()
    combo = (fi + " " + lt).str.strip()
    liab_name = out["Liability Name"].fillna("").astype(str).str.strip()
    out["_m61_facility_for_match"] = combo.where(combo.ne(""), liab_name)
    return out


def _has_compare_value(v):
    if pd.isna(v):
        return False
    return str(v).strip().lower() not in ("", "nan", "none")


def _coerce_numeric_value(v) -> float | None:
    """Best-effort numeric coercion; accepts percent strings like ``2.250%``."""
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none", "<na>", "nat"):
        return None
    # Missing placeholders from source files (must remain missing, never coerced to 0).
    if s in ("-", "—", "–", "−"):
        return None
    pct = "%" in s
    s_num = s.replace("%", "").replace(",", "").strip()
    if not s_num or s_num in ("-", "—", "–", "−"):
        return None
    try:
        out = float(s_num)
    except ValueError:
        return None
    return out / 100.0 if pct else out


def _coerce_rate_fraction(v) -> float | None:
    """Normalize a rate-like value to fraction form for status comparison.

    Accepts either percent-point inputs (e.g. 7.43, 78.14) or fraction inputs (0.0743, 0.7814),
    including strings with '%' already handled by ``_coerce_numeric_value``.
    """
    n = _coerce_numeric_value(v)
    if n is None:
        return None
    # Heuristic: values beyond +/-1 are percent-points, convert to fraction.
    return (n / 100.0) if abs(n) > 1.0 else n


_SPREAD_STATUS_PCT_QUANT = Decimal("0.01")


def _spread_percent_quantized_m61_compare(v) -> Decimal | None:
    """Spread as a percentage number quantized to 2 decimal places (M61 / display precision).

    Normalizes via ``_coerce_rate_fraction`` (percent strings, percent-points, fractions),
    then rounds half-up so e.g. ``4.875%`` vs ``4.88%`` compares equal.
    """
    frac = _coerce_rate_fraction(v)
    if frac is None:
        return None
    try:
        return (Decimal(str(frac)) * Decimal("100")).quantize(
            _SPREAD_STATUS_PCT_QUANT, rounding=ROUND_HALF_UP
        )
    except InvalidOperation:
        return None


def _normalize_aoc_ii_m61_adv_value(v):
    """Null/empty M61 advance cells → NA for display (—); preserve real numeric values."""
    if _is_blank_for_compare(v):
        return pd.NA
    return v


def _normalize_index_floor_value(v):
    """Normalize index-floor-like values; blank/zero-ish values are treated as missing."""
    n = _coerce_numeric_value(v)
    if n is not None:
        return pd.NA if abs(n) <= FLOAT_TOLERANCE else n
    if v is None:
        return pd.NA
    try:
        if pd.isna(v):
            return pd.NA
    except (TypeError, ValueError):
        return pd.NA
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none", "<na>", "nat"):
        return pd.NA
    return s


def compare_values(val_a, val_b, comparison_type):
    if pd.isna(val_a) and pd.isna(val_b):
        return FIELD_STATUS_MATCH
    if pd.isna(val_a) or pd.isna(val_b):
        return FIELD_STATUS_MISMATCH

    if comparison_type == "numeric":
        try:
            n_a = _coerce_numeric_value(val_a)
            n_b = _coerce_numeric_value(val_b)
            if n_a is None or n_b is None:
                return FIELD_STATUS_MISMATCH
            return (
                FIELD_STATUS_MATCH
                if abs(n_a - n_b) <= FLOAT_TOLERANCE
                else FIELD_STATUS_MISMATCH
            )
        except Exception:
            return FIELD_STATUS_MISMATCH

    if comparison_type == "date":
        s_a = safe_parse_date(pd.Series([val_a])).iloc[0]
        s_b = safe_parse_date(pd.Series([val_b])).iloc[0]
        if pd.isna(s_a) and pd.isna(s_b):
            return FIELD_STATUS_MATCH
        if pd.isna(s_a) or pd.isna(s_b):
            return FIELD_STATUS_MISMATCH
        return (
            FIELD_STATUS_MATCH
            if s_a.normalize().date() == s_b.normalize().date()
            else FIELD_STATUS_MISMATCH
        )

    return (
        FIELD_STATUS_MATCH
        if normalise_text(val_a) == normalise_text(val_b)
        else FIELD_STATUS_MISMATCH
    )


def compare_liability_primary_status(val_liab, val_acp, comparison_type):
    """
    Liability (export) vs primary model: MATCH / MISMATCH / MISSING FROM M61 / MISSING FROM ACORE / MISSING FROM BOTH.
    """
    l_ok = _has_compare_value(val_liab)
    p_ok = _has_compare_value(val_acp)
    if not l_ok and not p_ok:
        return FIELD_STATUS_MISSING_BOTH
    if not p_ok and l_ok:
        return FIELD_STATUS_MISSING_ACORE
    if p_ok and not l_ok:
        return FIELD_STATUS_MISSING_M61
    return compare_values(val_liab, val_acp, comparison_type)


def compare_optional(val_liab, val_acp, kind="text"):
    """
    Optional cross-file fields (index floor/name, recourse): same vocabulary as other field statuses.
    """
    l_ok = not _is_blank_for_compare(val_liab)
    a_ok = not _is_blank_for_compare(val_acp)
    if not l_ok and not a_ok:
        return FIELD_STATUS_MISSING_BOTH
    if l_ok and not a_ok:
        return FIELD_STATUS_MISSING_ACORE
    if a_ok and not l_ok:
        return FIELD_STATUS_MISSING_M61
    eff_kind = kind
    if kind == "numeric":
        try:
            float(val_liab)
            float(val_acp)
        except (TypeError, ValueError):
            eff_kind = "text"
    return compare_values(val_liab, val_acp, eff_kind)


def _effective_date_cell_populated(v):
    """
    True when the cell resolves to a real calendar date (datetime, string date, Excel serial, etc.).
    Uses the same parsing path as the rest of the tool so primary pledge dates are not treated as blank.
    """
    if v is None or isinstance(v, bool):
        return False
    try:
        if pd.isna(v):
            return False
    except (TypeError, ValueError):
        return False

    if isinstance(v, str):
        s = v.strip()
        if not s or s.lower() in ("nan", "none", "nat", "<na>", "-", "—", "n/a"):
            return False

    dt = safe_parse_date(pd.Series([v])).iloc[0]
    if pd.isna(dt):
        return False

    # Small integers often mis-parse as 1970-01-01 via ns-epoch heuristics in to_datetime.
    try:
        fv = float(v)
    except (TypeError, ValueError):
        fv = None
    if fv is not None and abs(fv) < 1000:
        if dt.normalize() == pd.Timestamp("1970-01-01").normalize():
            return False

    return True


def compare_effective_date_status(val_liability, val_acp):
    has_acp = _effective_date_cell_populated(val_acp)
    has_liab = _effective_date_cell_populated(val_liability)

    if not has_acp and not has_liab:
        return FIELD_STATUS_MISSING_BOTH
    if not has_acp:
        return FIELD_STATUS_MISSING_ACORE
    if not has_liab:
        return FIELD_STATUS_MISSING_M61

    dt_liab = safe_parse_date(pd.Series([val_liability])).iloc[0]
    dt_acp = safe_parse_date(pd.Series([val_acp])).iloc[0]
    if pd.isna(dt_acp):
        return FIELD_STATUS_MISSING_ACORE
    if pd.isna(dt_liab):
        return FIELD_STATUS_MISSING_M61
    return (
        FIELD_STATUS_MATCH
        if dt_liab.normalize().date() == dt_acp.normalize().date()
        else FIELD_STATUS_MISMATCH
    )


def _recon_token_for_effective_date_status(display):
    if display == FIELD_STATUS_MATCH:
        return "MATCH"
    if display in (FIELD_STATUS_MISMATCH, "NO MATCH"):
        return "MISMATCH"
    return "MISSING"


def _recon_token_for_compare_status(display):
    if display == FIELD_STATUS_MATCH:
        return "MATCH"
    if display == FIELD_STATUS_MISMATCH:
        return "MISMATCH"
    if isinstance(display, str) and display.startswith("MISSING FROM"):
        return "MISSING"
    return "MISSING"


def _is_blank_for_compare(v):
    if v is None or pd.isna(v):
        return True
    s = str(v).strip().lower()
    return s in ("", "nan", "none", "nat", "<na>")


def compare_pledge_date_status(*, val_liability, val_acp):
    has_acp = _effective_date_cell_populated(val_acp)
    has_liab = _effective_date_cell_populated(val_liability)

    if not has_acp and not has_liab:
        return FIELD_STATUS_MISSING_BOTH
    if not has_acp and has_liab:
        return FIELD_STATUS_MISSING_ACORE
    if has_acp and not has_liab:
        return FIELD_STATUS_MISSING_M61

    dt_l = safe_parse_date(pd.Series([val_liability])).iloc[0]
    dt_a = safe_parse_date(pd.Series([val_acp])).iloc[0]
    if pd.isna(dt_a):
        return FIELD_STATUS_MISSING_ACORE
    if pd.isna(dt_l):
        return FIELD_STATUS_MISSING_M61
    if dt_l.normalize().date() == dt_a.normalize().date():
        return FIELD_STATUS_MATCH
    return FIELD_STATUS_MISMATCH


def _primary_facility_match_token(source, facility) -> str:
    """Single normalized facility/source token for financing match keys (aligns with ``normalise_facility``)."""
    for raw in (facility, source):
        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
            continue
        s = _normalise_dashes_for_facility(str(raw).strip())
        if not s or s.lower() in ("nan", "none", "<na>", "nat"):
            continue
        parts = [p.strip() for p in re.split(r"\||,|/", s) if p.strip()]
        # Prefer TBD composite segments (e.g. ``Repo | ACPIII-TBD-TBD``) before whole-string norm.
        if len(parts) > 1:
            for p in parts:
                if _facility_source_collapses_to_tbd(p):
                    return "tbd"
        if _facility_source_collapses_to_tbd(s):
            return "tbd"
        t = normalise_facility(s)
        if t:
            return t
        for p in parts:
            t2 = normalise_facility(p)
            if t2:
                return t2
    return ""


def _normalize_ln_suffix_token(raw) -> str:
    """Map parsed liability-note suffix (``GS``, ``BOA``, …) to the same token space as ``normalise_facility``."""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    s = str(raw).strip()
    if not s or s.lower() in ("nan", "none", "<na>", "nat"):
        return ""
    if _facility_source_collapses_to_tbd(s):
        return "tbd"
    bank = _normalise_m61_fund_prefixed_facility(s)
    if bank is not None:
        return bank
    u = s.upper()
    alias = {"BAML": "boa", "BAC": "boa", "GS": "gs", "JPM": "jpm", "MS": "ms", "BOA": "boa", "TBD": "tbd"}
    if u in alias:
        return alias[u]
    return normalise_facility(f"{s} repo") or normalise_text(s)


def _facility_norm_for_debug_cell(v) -> str:
    """Normalized facility token for recon output / debug columns (blank → empty string)."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    t = str(v).strip()
    if not t or t.lower() in ("nan", "none", "<na>", "nat"):
        return ""
    return normalise_facility(t)


def _fin_m61_key_from_row(row: pd.Series, scope_label: str) -> str:
    """Composite financing key on M61 rows: deal id + eff date + optional suffix; ``Fin`` + fund scope only.

    ACP II extension: ``Sub`` (subline) rows also produce a bare deal-id + effective-date key so
    they can match ACORE "Sub Debt" rows via the financing_note stage.  LN_Sub notes carry no
    source suffix, and ACORE "Sub Debt" rows with unrecognised facilities (e.g. Bank OZK) also
    fall back to the bare key — so the two sides naturally align.
    """
    p = parse_liability_note(row.get("Liability Note"))
    nc = p.get("note_category")
    # ACP II-only: LN_Sub rows produce a keyed ``|sub`` token so they exclusively match
    # ACORE "Sub Debt" source rows (which also carry the ``|sub`` discriminator on the
    # ACORE side).  Using a bare deal+date key would cause LN_Sub rows to steal matches
    # from Repo/NoN/CLO ACORE rows that share the same bare key (unrecognised facility).
    if nc == "Sub" and scope_label == "ACP II":
        did = normalise_deal_id_key(p.get("deal_id") or "")
        if not did:
            return ""
        fc = (p.get("fund_code") or "").strip()
        if fc and fc != scope_label:
            return ""
        if not fc and not bool(row.get("_m61_in_scope")):
            return ""
        eff = str(row.get("effective_date_key") or "")
        return f"{did}|{eff}|sub"
    if nc != "Fin":
        return ""
    did = normalise_deal_id_key(p.get("deal_id") or "")
    if not did:
        return ""
    fc = (p.get("fund_code") or "").strip()
    if fc and fc != scope_label:
        return ""
    if not fc and not bool(row.get("_m61_in_scope")):
        return ""
    eff = str(row.get("effective_date_key") or "")
    sx = _normalize_ln_suffix_token(p.get("source_suffix") or "")
    # Liability Note may omit bank/repo while **Liability Name** carries ``ACPIII-MS-Repo`` / ``…-TBD-TBD``.
    if not sx:
        sx = _primary_facility_match_token(
            row.get("Financial Institution"),
            row.get("Liability Name"),
        )
        # Only use the FI/name fallback when it is a recognised bank token (jpm/gs/ms/boa/tbd).
        # Generic liability names such as ``AOCII-3rd Party-Sale`` produce unstable join keys
        # that diverge from the ACORE facility token and block valid Deal-ID matches.
        if sx not in FACILITY_NORM_MAP.values():
            sx = ""
    # When the note has no suffix and name/FI yield no token, match on deal id + effective date only.
    if sx:
        return f"{did}|{eff}|{sx}"
    return f"{did}|{eff}"


def build_match_key(df, deal_col, facility_col, note_col, effective_date_col):
    """Join key = normalized deal + facility + effective date (``note_col`` is unused; kept for call-site clarity)."""
    out = df.copy()
    out["deal_norm"] = out[deal_col].apply(normalise_text)
    out["facility_norm"] = out[facility_col].apply(normalise_facility)
    out["effective_date_key"] = date_key(out[effective_date_col])
    out["match_key"] = (
        out["deal_norm"]
        + " | "
        + out["facility_norm"]
        + " | "
        + out["effective_date_key"]
    )
    return out


def _apply_deal_id_suffix_fallback_matches(
    merged: pd.DataFrame, df_b: pd.DataFrame, df_a: pd.DataFrame, *, label_a: str
) -> tuple[pd.DataFrame, int]:
    """
    Fallback aid only for currently-unmatched rows:
    pair one-to-one rows where primary Deal ID == extracted Liability Note suffix.
    """
    need_b = {"match_key", "deal_id_key"}
    need_a = {"match_key", "liability_note_suffix_key"}
    if not need_b.issubset(df_b.columns) or not need_a.issubset(df_a.columns):
        return merged, 0
    if "_merge" not in merged.columns:
        return merged, 0

    mk_a = f"{label_a}_match_key"
    right_only_keys = set(merged.loc[merged["_merge"] == "right_only", mk_a].dropna().tolist())
    left_only_keys = set(merged.loc[merged["_merge"] == "left_only", "match_key"].dropna().tolist())
    if not right_only_keys or not left_only_keys:
        return merged, 0

    b_pool = df_b[df_b["match_key"].isin(left_only_keys)].copy()
    a_pool = df_a[df_a["match_key"].isin(right_only_keys)].copy()
    b_pool = b_pool[b_pool["deal_id_key"].ne("")].copy()
    a_pool = a_pool[a_pool["liability_note_suffix_key"].ne("")].copy()
    if b_pool.empty or a_pool.empty:
        return merged, 0

    b_counts = b_pool["deal_id_key"].value_counts(dropna=False)
    a_counts = a_pool["liability_note_suffix_key"].value_counts(dropna=False)
    unique_keys = set(b_counts[b_counts == 1].index).intersection(set(a_counts[a_counts == 1].index))
    if not unique_keys:
        return merged, 0

    b_pick = b_pool[b_pool["deal_id_key"].isin(unique_keys)].copy()
    a_pick = a_pool[a_pool["liability_note_suffix_key"].isin(unique_keys)].copy()
    fb = pd.merge(
        b_pick,
        a_pick.add_prefix(f"{label_a}_"),
        left_on="deal_id_key",
        right_on=f"{label_a}_liability_note_suffix_key",
        how="inner",
    )
    if fb.empty:
        return merged, 0

    fb["_merge"] = "both"
    fb["_fallback_match_method"] = "deal_id_suffix"

    used_b = set(fb["match_key"].dropna().tolist())
    used_a = set(fb[mk_a].dropna().tolist())
    rm_left = (merged["_merge"] == "left_only") & merged["match_key"].isin(used_b)
    rm_right = (merged["_merge"] == "right_only") & merged[mk_a].isin(used_a)
    merged_keep = merged.loc[~(rm_left | rm_right)].copy()
    merged_out = pd.concat([merged_keep, fb], ignore_index=True, sort=False)
    return merged_out, int(len(fb))


def _merged_liab_col(row: pd.Series, label_a: str, source_col: str):
    key = f"{label_a}_{source_col}"
    if key not in row.index:
        return pd.NA
    return row[key]


def _first_liab_recourse_value(row: pd.Series, label_a: str):
    # Liability export uses "Recourse" in practice; keep fallbacks for other templates.
    for src in ("Recourse", "Recourse %", "RecoursePct"):
        key = f"{label_a}_{src}"
        if key not in row.index:
            continue
        v = row[key]
        if pd.isna(v):
            continue
        if isinstance(v, str) and not str(v).strip():
            continue
        return v
    return pd.NA


def _stripped_nonempty_str(v) -> str | None:
    """Return a stripped display string, or None if absent/blank (for coalescing recon columns)."""
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        return None
    s = str(v).strip()
    if not s or s.lower() in ("nan", "<na>", "nat", "none"):
        return None
    return s


def load_liability_cre_mapping(path: str) -> pd.DataFrame:
    """
    Loads LiabilityNoteID -> CRENoteID mapping workbook.
    Update sheet_name/column names if the template differs.
    """
    df = pd.read_excel(path, sheet_name=0)
    return df


def reconcile(
    file_a_path,
    file_b_path,
    primary_file_type: str = "ACORE",
    mapping_path: str | None = None,
    uploaded_primary_filename: str | None = None,
    return_diagnostics: bool = False,
    match_diagnostics: bool = False,
):
    global LAST_RECON_DIAGNOSTICS, LAST_RECON_CONTEXT
    label_a = "in_liability"
    p_cfg = get_primary_config(primary_file_type)
    # One fund identity per primary workbook template (no row-level Fund on Fin Inpt).
    business_fund_label = detect_fund_label(uploaded_primary_filename, primary_file_type)
    fund_fallback_display = _fund_cfg(primary_file_type).get("recon_fund_display") or business_fund_label

    raw_m61, excluded_by_type_df, m61_load_diag = load_file_a(
        file_a_path,
        primary_file_type=primary_file_type,
        return_excluded=True,
        return_diagnostics=True,
    )
    m61_row_count_after_load = len(raw_m61)
    _debug_rows(f"Reconcile input: M61 rows after load_file_a={len(raw_m61)}")
    _debug_rows(f"TEMP DEBUG: M61 row count immediately after load_file_a = {len(raw_m61)}")
    _debug_m61_load_preview(raw_m61)

    if primary_file_type in ("AOC II", "AOC I") and mapping_path:
        df_map = load_liability_cre_mapping(mapping_path)
        raw_m61 = raw_m61.merge(
            df_map,
            how="left",
            left_on="Liability Note",
            right_on="LiabilityNoteID",
        )
        # Keep distinct M61 liability lines (Fund vs Repo vs Subline, FI variants) from
        # collapsing onto the same CRE note id in key-based stages.
        _cre = raw_m61.get("CRENoteID", pd.Series([""] * len(raw_m61), index=raw_m61.index))
        _lt = raw_m61.get("Liability Type", pd.Series([""] * len(raw_m61), index=raw_m61.index))
        _fi = raw_m61.get("Financial Institution", pd.Series([""] * len(raw_m61), index=raw_m61.index))
        _ln = raw_m61.get("Liability Note", pd.Series([""] * len(raw_m61), index=raw_m61.index))
        raw_m61["_m61_match_facility"] = (
            _cre.fillna("").astype(str).str.strip()
            + " | "
            + _lt.fillna("").astype(str).str.strip()
            + " | "
            + _fi.fillna("").astype(str).str.strip()
            + " | "
            + _ln.fillna("").astype(str).str.strip()
        )
        _debug_trace_uncommons_m61_load(
            raw_m61,
            "reconcile:raw_m61_after_CRE_mapping_merge",
            primary_file_type=primary_file_type,
        )
        df_a = build_match_key(
            raw_m61,
            "Deal Name",
            "_m61_match_facility",
            "Liability Note",
            "Effective Date",
        )
        df_pri_raw = load_primary_file(file_b_path, primary_file_type)
        _debug_rows(
            f"TEMP DEBUG: Primary workbook row count immediately after load_primary_file = {len(df_pri_raw)}"
        )
        df_b = build_match_key(
            df_pri_raw,
            "Deal Name",
            "Note Name",
            "Note Name",
            "Effective Date",
        )
        _debug_match_key_sample_rows(
            "primary (Fin Inpt)",
            df_b,
            deal_col="Deal Name",
            facility_col="Note Name",
            note_col="Note Name",
            eff_col="Effective Date",
        )
        _debug_match_key_sample_rows(
            "M61",
            df_a,
            deal_col="Deal Name",
            facility_col="CRENoteID",
            note_col="Liability Note",
            eff_col="Effective Date",
        )
    else:
        # Match key: Deal Name + facility token + Effective Date (legacy V7 parity). Primary uses
        # ``Facility``; M61 uses ``Liability Name`` in that slot — same pairing as Financing Line
        # Reconciliation Tool V7 (FI + Liability Type was not aligning to Fin Inpt Facility strings).
        df_a = build_match_key(
            raw_m61,
            "Deal Name",
            "Liability Name",
            "Liability Note",
            "Effective Date",
        )
        df_pri_raw = load_primary_file(file_b_path, primary_file_type)
        _debug_rows(
            f"TEMP DEBUG: Primary workbook row count immediately after load_primary_file = {len(df_pri_raw)}"
        )
        df_b = build_match_key(
            df_pri_raw,
            "Deal Name",
            "Facility",
            "Note Name",
            "Effective Date",
        )
        _debug_match_key_sample_rows(
            "primary (Fin Inpt)",
            df_b,
            deal_col="Deal Name",
            facility_col="Facility",
            note_col="Note Name",
            eff_col="Effective Date",
        )
        _debug_match_key_sample_rows(
            "M61",
            df_a,
            deal_col="Deal Name",
            facility_col="Liability Name",
            note_col="Liability Note",
            eff_col="Effective Date",
        )

    # TEMP DEBUG: spread source resolution on each side before merged assembly.
    _debug_rows(
        "TEMP DEBUG: spread column presence — "
        f"primary_has_spread={'Spread' in df_b.columns} m61_has_spread={'Spread' in df_a.columns}"
    )
    if "Spread" in df_b.columns:
        _debug_rows(
            "TEMP DEBUG: primary Spread sample head(10)="
            f"{df_b['Spread'].head(10).tolist()!r}"
        )
    if "Spread" in df_a.columns:
        _debug_rows(
            "TEMP DEBUG: M61 Spread sample head(10)="
            f"{df_a['Spread'].head(10).tolist()!r}"
        )
    _debug_rows(
        "TEMP DEBUG: index-floor column presence — "
        f"primary_has_floor={'Floor' in df_b.columns} m61_has_indexfloor={'IndexFloor' in df_a.columns}"
    )
    if "Floor" in df_b.columns:
        _debug_rows(
            "TEMP DEBUG: primary Floor sample head(10)="
            f"{df_b['Floor'].head(10).tolist()!r}"
        )
    if "IndexFloor" in df_a.columns:
        _debug_rows(
            "TEMP DEBUG: M61 IndexFloor sample head(10)="
            f"{df_a['IndexFloor'].head(10).tolist()!r}"
        )

    # Helper IDs for validation / fallback alignment aid.
    if "Deal ID" in df_b.columns:
        df_b["deal_id_value"] = df_b["Deal ID"].apply(lambda v: "" if pd.isna(v) else str(v).strip())
    else:
        df_b["deal_id_value"] = ""
    # Normalize ACP key via same token extractor used for M61 notes.
    df_b["acp_extracted_deal_id"] = df_b["deal_id_value"].apply(extract_deal_id_token)
    # Fin Inpt fallback: when Deal ID column is blank, recover id token from Note/Deal text if present.
    _acp_missing_id = df_b["acp_extracted_deal_id"].astype(str).str.strip().eq("")
    if _acp_missing_id.any():
        _note_fallback = (
            df_b.loc[_acp_missing_id, "Note Name"].apply(extract_deal_id_token)
            if "Note Name" in df_b.columns
            else pd.Series("", index=df_b.index[_acp_missing_id], dtype="object")
        )
        df_b.loc[_acp_missing_id, "acp_extracted_deal_id"] = _note_fallback
    _acp_missing_id2 = df_b["acp_extracted_deal_id"].astype(str).str.strip().eq("")
    if _acp_missing_id2.any():
        _deal_fallback = df_b.loc[_acp_missing_id2, "Deal Name"].apply(extract_deal_id_token)
        df_b.loc[_acp_missing_id2, "acp_extracted_deal_id"] = _deal_fallback
    df_b["acp_match_key"] = df_b["acp_extracted_deal_id"].apply(normalise_deal_id_key)
    df_b["deal_id_key"] = df_b["acp_match_key"]
    _debug_trace_uncommons_primary_fin_inpt(df_b, primary_file_type=primary_file_type)

    df_a["m61_extracted_deal_id"] = df_a["Liability Note"].apply(extract_liability_note_suffix)
    df_a["liability_note_suffix"] = df_a["m61_extracted_deal_id"]
    df_a["liability_note_suffix_key"] = df_a["liability_note_suffix"].apply(normalise_deal_id_key)
    df_a["m61_match_key"] = df_a["liability_note_suffix_key"].astype(str)
    _debug_trace_uncommons_m61_load(
        df_a,
        "reconcile:df_a_after_liability_note_suffix",
        primary_file_type=primary_file_type,
    )
    _debug_rows(
        "TEMP DEBUG: Deal ID helper readiness — "
        f"primary_nonblank_deal_id={int(df_b['acp_match_key'].ne('').sum())} "
        f"m61_nonblank_note_suffix={int(df_a['m61_match_key'].ne('').sum())}"
    )
    _debug_rows("TEMP DEBUG: Deal ID helper sample — primary head 10")
    for i, (_, r) in enumerate(
        df_b.loc[
            :,
            [
                c
                for c in [
                    "Deal Name",
                    "deal_id_value",
                    "acp_extracted_deal_id",
                    "acp_match_key",
                    "effective_date_key",
                    "match_key",
                ]
                if c in df_b.columns
            ],
        ]
        .head(10)
        .iterrows(),
        start=1,
    ):
        _debug_rows(
            "TEMP DEBUG:   "
            f"#{i} deal={r.get('Deal Name')!r} deal_id={r.get('deal_id_value')!r} "
            f"acp_extracted={r.get('acp_extracted_deal_id')!r} acp_match_key={r.get('acp_match_key')!r} "
            f"eff_key={r.get('effective_date_key')!r} match_key={r.get('match_key')!r}"
        )
    _debug_rows("TEMP DEBUG: Deal ID helper sample — M61 head 10")
    for i, (_, r) in enumerate(
        df_a.loc[
            :,
            [
                c
                for c in [
                    "Deal Name",
                    "Liability Note",
                    "m61_extracted_deal_id",
                    "m61_match_key",
                    "effective_date_key",
                    "match_key",
                ]
                if c in df_a.columns
            ],
        ]
        .head(10)
        .iterrows(),
        start=1,
    ):
        _debug_rows(
            "TEMP DEBUG:   "
            f"#{i} deal={r.get('Deal Name')!r} liability_note={r.get('Liability Note')!r} "
            f"m61_extracted_deal_id={r.get('m61_extracted_deal_id')!r} m61_match_key={r.get('m61_match_key')!r} "
            f"eff_key={r.get('effective_date_key')!r} "
            f"match_key={r.get('match_key')!r}"
        )
    _debug_rows(
        "TEMP DEBUG: requested key columns (ACP side) — "
        "ACP Deal ID | acp_match_key"
    )
    for i, (_, r) in enumerate(
        df_b.loc[:, [c for c in ["Deal Name", "deal_id_value", "acp_match_key"] if c in df_b.columns]]
        .head(50)
        .iterrows(),
        start=1,
    ):
        _debug_rows(
            f"TEMP DEBUG:   ACP[{i}] deal={r.get('Deal Name')!r} "
            f"ACP Deal ID={r.get('deal_id_value')!r} acp_match_key={r.get('acp_match_key')!r}"
        )
    _debug_rows(
        "TEMP DEBUG: requested key columns (M61 side) — "
        "M61 Liability Note | m61_extracted_deal_id | m61_match_key"
    )
    for i, (_, r) in enumerate(
        df_a.loc[:, [c for c in ["Deal Name", "Liability Note", "m61_extracted_deal_id", "m61_match_key"] if c in df_a.columns]]
        .head(50)
        .iterrows(),
        start=1,
    ):
        _debug_rows(
            f"TEMP DEBUG:   M61[{i}] deal={r.get('Deal Name')!r} "
            f"M61 Liability Note={r.get('Liability Note')!r} "
            f"m61_extracted_deal_id={r.get('m61_extracted_deal_id')!r} "
            f"m61_match_key={r.get('m61_match_key')!r}"
        )

    _debug_match_key_overlap_diagnosis(df_b, df_a, n=10)

    # Informational only: do not drop M61 rows here — outer merge below already carries
    # M61-only / primary-only sides; pre-filtering df_a hid M61 export rows absent from the model.
    primary_match_keys = set(df_b["match_key"].dropna().tolist())
    m61_in_primary = int(df_a["match_key"].isin(primary_match_keys).sum())
    _debug_rows(
        "M61 vs primary match_key overlap (no row drop): "
        f"m61_rows={len(df_a)} primary_rows={len(df_b)} "
        f"m61_rows_with_key_in_primary={m61_in_primary} distinct_primary_keys={len(primary_match_keys)}"
    )

    # ---- Row-level staged matching (strict -> source-aware -> fallback), then outer-preserving assembly ----
    b = df_b.copy()
    a = df_a.copy()
    b["_row_id_b"] = range(len(b))
    a["_row_id_a"] = range(len(a))
    b["strict_note_norm"] = b["Note Name"].apply(normalise_text) if "Note Name" in b.columns else ""
    a["strict_note_norm"] = a["Liability Note"].apply(normalise_text) if "Liability Note" in a.columns else ""
    b["strict_key"] = (
        b["deal_norm"] + " | " + b["facility_norm"] + " | " + b["strict_note_norm"] + " | " + b["effective_date_key"]
    )
    a["strict_key"] = (
        a["deal_norm"] + " | " + a["facility_norm"] + " | " + a["strict_note_norm"] + " | " + a["effective_date_key"]
    )
    b["deal_date_key"] = b["deal_norm"] + " | " + b["effective_date_key"]
    a["deal_date_key"] = a["deal_norm"] + " | " + a["effective_date_key"]
    b["source_bucket"] = (
        b["Source"].apply(lambda v: _source_bucket(v, primary_file_type=primary_file_type))
        if "Source" in b.columns
        else ""
    )
    a["source_bucket"] = (
        a.apply(
            lambda r: _source_bucket(
                r.get("Liability Type"),
                primary_file_type=primary_file_type,
                liability_note_raw=r.get("Liability Note"),
                is_m61=True,
            ),
            axis=1,
        )
        if "Liability Type" in a.columns
        else ""
    )
    b["source_aware_key"] = b["deal_date_key"] + " | " + b["source_bucket"]
    a["source_aware_key"] = a["deal_date_key"] + " | " + a["source_bucket"]
    b["source_aware_facility_key"] = b["source_aware_key"] + " | " + b["facility_norm"]
    a["source_aware_facility_key"] = a["source_aware_key"] + " | " + a["facility_norm"]
    fcfg = _fund_cfg(primary_file_type)
    fpattern = fcfg.get("fund_regex") or (
        re.escape(fcfg.get("fund_token")) if fcfg.get("fund_token") else None
    )
    if fpattern:
        a["_m61_in_scope"] = (
            a["Fund Name"].fillna("").astype(str).str.contains(fpattern, case=False, regex=True, na=False)
        )
    else:
        a["_m61_in_scope"] = True
    _debug_rows(
        "TEMP DEBUG: staged matcher fund-scope guard on M61 candidates — "
        f"in_scope={int(a['_m61_in_scope'].sum())}/{len(a)} "
        f"pattern={fpattern!r} primary_type={primary_file_type}"
    )
    primary_id_keys = set(b["acp_match_key"].astype(str).tolist())
    primary_id_keys.discard("")
    a["_id_in_primary"] = a["m61_match_key"].astype(str).isin(primary_id_keys)
    _debug_rows(
        "TEMP DEBUG: Deal ID / Liability Note suffix key overlap — "
        f"primary_nonblank_ids={len(primary_id_keys)} "
        f"m61_rows_with_id_in_primary={int(a['_id_in_primary'].sum())}/{len(a)}"
    )
    m61_rows_after_type = int(len(a))
    m61_rows_nonblank_note = int(a["Liability Note"].fillna("").astype(str).str.strip().ne("").sum())
    m61_rows_extracted_nonblank = int(a["m61_match_key"].astype(str).ne("").sum())
    m61_rows_extracted_in_acp = int(a["_id_in_primary"].sum())
    acp_total_rows = int(len(b))
    m61_total_rows_after_filter = int(len(a))
    _debug_rows(
        "TEMP DEBUG: requested pre-merge counts — "
        f"total_acp_rows={acp_total_rows} "
        f"total_m61_rows_after_liability_type_filter={m61_total_rows_after_filter} "
        f"m61_rows_with_extracted_deal_id={m61_rows_extracted_nonblank} "
        f"m61_extracted_ids_matching_acp={m61_rows_extracted_in_acp}"
    )
    _debug_rows(
        "TEMP DEBUG: M61 extraction counts after liability-type filter — "
        f"rows_after_type={m61_rows_after_type} "
        f"nonblank_liability_note={m61_rows_nonblank_note} "
        f"nonblank_extracted_deal_id={m61_rows_extracted_nonblank} "
        f"extracted_id_in_acp={m61_rows_extracted_in_acp}"
    )
    if m61_rows_extracted_in_acp == 0:
        _debug_rows(
            "TEMP DEBUG: extracted M61 deal IDs matching ACP = 0; printing first 50 raw Liability Note values."
        )
        for i, v in enumerate(a["Liability Note"].head(50).tolist(), start=1):
            _debug_rows(f"TEMP DEBUG: raw_m61_liability_note[{i}]={v!r}")
    _debug_rows("TEMP DEBUG: M61 first 50 rows after type filter (note + extracted id)")
    m61_dbg_cols = [
        c
        for c in (
            "Deal Name",
            "Liability Type",
            "Liability Note",
            "liability_note_suffix",
            "Target Advance Rate",
        )
        if c in a.columns
    ]
    for i, (_, rr) in enumerate(a.loc[:, m61_dbg_cols].head(50).iterrows(), start=1):
        _debug_rows(
            "TEMP DEBUG:   "
            f"#{i} deal={rr.get('Deal Name')!r} "
            f"type={rr.get('Liability Type')!r} "
            f"liability_note={rr.get('Liability Note')!r} "
            f"extracted_id={rr.get('liability_note_suffix')!r} "
            f"target_adv={rr.get('Target Advance Rate')!r}"
        )

    unmatched_b = set(b["_row_id_b"].tolist())
    unmatched_a = set(a["_row_id_a"].tolist())
    matchable_a = set(a.loc[a["_m61_in_scope"], "_row_id_a"].tolist())
    if primary_file_type in FIN_INPT_PRIMARY_TYPES:
        scope_lbl = scope_label_for_primary_type(primary_file_type)
        # Fin Inpt reconciliation is financing-only on the M61 side (no Eq/Fund fallback pairing).
        a["_m61_note_category"] = a["Liability Note"].apply(
            lambda v: parse_liability_note(v).get("note_category", "Other")
        )
        # ACP II: include LN_Sub (Subline) rows alongside LN_Fin so ACORE "Sub Debt" rows
        # can match their M61 counterparts.  All other Fin Inpt types remain Fin-only.
        if primary_file_type == "ACP II":
            fin_note_rows = set(
                a.loc[a["_m61_note_category"].isin({"Fin", "Sub"}), "_row_id_a"].tolist()
            )
        else:
            fin_note_rows = set(a.loc[a["_m61_note_category"].eq("Fin"), "_row_id_a"].tolist())
        matchable_a = matchable_a.intersection(fin_note_rows)
        _scope_ok = pd.Series(True, index=a.index)
        # Tight fund guard for ACP II / AOC II / AOC I where shared Deal IDs across funds are common.
        if primary_file_type in ("ACP II", "AOC II", "AOC I"):
            a["_m61_note_fund_code"] = a["Liability Note"].apply(
                lambda v: (parse_liability_note(v).get("fund_code") or "").strip()
            )
            _scope_ok = a["_m61_in_scope"] & a["_m61_note_fund_code"].eq(scope_lbl)
            matchable_a = matchable_a.intersection(set(a.loc[_scope_ok, "_row_id_a"].tolist()))
        # Fin Inpt runs (ACP III / ACP II / AOC II): anchor reconciliation candidates to
        # financing identifiers. Only restrict to ID-matched candidates when Deal IDs are
        # actually present on the primary side; if none exist, fall back to the fund-scope
        # candidate set so rows are not silently dropped.
        id_anchored = set(
            a.loc[a["_id_in_primary"] & _scope_ok, "_row_id_a"].tolist()
        ).intersection(fin_note_rows)
        if id_anchored:
            matchable_a = id_anchored
        else:
            _debug_rows(
                "TEMP DEBUG: Fin Inpt anchored candidate set is EMPTY "
                "(no primary Deal IDs found); falling back to fund-scope matchable_a"
            )
        _debug_rows(
            "TEMP DEBUG: Fin Inpt financing-note guard "
            f"fin_note_rows={len(fin_note_rows)}/{len(a)} matchable_after_guard={len(matchable_a)}"
        )
        _debug_rows(
            "TEMP DEBUG: Fin Inpt anchored candidate set "
            f"(ID overlap)={len(matchable_a)}/{len(a)}"
        )

    b["fin_acp_key"] = ""
    a["fin_m61_key"] = ""
    if primary_file_type in FIN_INPT_PRIMARY_TYPES:
        b["_pri_fac_tok"] = b.apply(
            lambda r: _primary_facility_match_token(r.get("Source"), r.get("Facility")), axis=1
        )
        msk_b = b["deal_id_key"].astype(str).str.strip().ne("")
        sub_b = b.loc[msk_b]
        tok_stripped = sub_b["_pri_fac_tok"].fillna("").astype(str).str.strip()
        base_k = sub_b["deal_id_key"].astype(str) + "|" + sub_b["effective_date_key"].astype(str)
        fin_vals = base_k + "|" + sub_b["_pri_fac_tok"].fillna("").astype(str)
        _known_fac_toks = frozenset(FACILITY_NORM_MAP.values())
        bare_cond = tok_stripped.eq("") | ~tok_stripped.isin(_known_fac_toks)
        fin_vals.loc[bare_cond] = base_k[bare_cond]
        b.loc[msk_b, "fin_acp_key"] = fin_vals
        # ACP II: override fin_acp_key for "Sub Debt" source rows to use the ``|sub``
        # discriminator token, matching the M61 LN_Sub key.  This ensures Sub Debt rows
        # only pair with LN_Sub counterparts and never compete with Repo/NoN/CLO rows
        # whose bare deal+date key would otherwise collide.
        if primary_file_type == "ACP II":
            sub_debt_mask = msk_b & b["Source"].fillna("").astype(str).str.lower().str.strip().eq("sub debt")
            if sub_debt_mask.any():
                sub_b2 = b.loc[sub_debt_mask]
                b.loc[sub_debt_mask, "fin_acp_key"] = (
                    sub_b2["deal_id_key"].astype(str) + "|" +
                    sub_b2["effective_date_key"].astype(str) + "|sub"
                )
        a["fin_m61_key"] = a.apply(lambda r: _fin_m61_key_from_row(r, scope_lbl), axis=1)

    pair_rows: list[dict] = []
    b_by_id = b.set_index("_row_id_b", drop=False)
    a_by_id = a.set_index("_row_id_a", drop=False)

    def _pair_by_key(key_b: str, key_a: str, stage: str) -> int:
        lb = b[b["_row_id_b"].isin(unmatched_b)].copy()
        la = a[a["_row_id_a"].isin(unmatched_a.intersection(matchable_a))].copy()
        lb = lb[lb[key_b].astype(str).str.strip().ne("")]
        la = la[la[key_a].astype(str).str.strip().ne("")]
        if lb.empty or la.empty:
            return 0
        lb = lb.sort_values(["Deal Name", "effective_date_key", "_row_id_b"]).copy()
        la = la.sort_values(["Deal Name", "effective_date_key", "_row_id_a"]).copy()
        lb["_rk"] = lb.groupby(key_b).cumcount()
        la["_rk"] = la.groupby(key_a).cumcount()
        pairs = lb[["_row_id_b", key_b, "_rk"]].merge(
            la[["_row_id_a", key_a, "_rk"]],
            left_on=[key_b, "_rk"],
            right_on=[key_a, "_rk"],
            how="inner",
        )
        if pairs.empty:
            return 0
        for _, pr in pairs.iterrows():
            rid_b = int(pr["_row_id_b"])
            rid_a = int(pr["_row_id_a"])
            if rid_b not in unmatched_b or rid_a not in unmatched_a:
                continue
            unmatched_b.remove(rid_b)
            unmatched_a.remove(rid_a)
            pair_rows.append({"_row_id_b": rid_b, "_row_id_a": rid_a, "_match_stage": stage, "_merge": "both"})
            br = b_by_id.loc[rid_b]
            ar = a_by_id.loc[rid_a]
            reason = {
                "financing_note": "Financing: primary Deal ID + effective date + facility/suffix aligned to parsed M61 liability note",
                "strict": "all strict components aligned (deal, facility, note, effective date)",
                "source_aware_facility": "same deal+effective date+source with facility alignment",
                "source_aware": "same deal+effective date+source",
                "fallback": "fallback on deal+effective date only",
            }.get(stage, stage)
            _debug_rows(
                "TEMP DEBUG: selected pair "
                f"stage={stage} acp_id={rid_b} m61_id={rid_a} "
                f"deal={br.get('Deal Name')!r} eff_key={br.get('effective_date_key')!r} "
                f"acp_source={br.get('Source')!r} m61_type={ar.get('Liability Type')!r} "
                f"m61_undrawn={ar.get('Undrawn Capacity')!r} reason={reason}"
            )
        return len(pairs)

    def _pair_note_deal_id_relaxed() -> int:
        """ACP II / AOC II / AOC I only: pair by parsed liability-note deal id ``==`` primary ``acp_match_key``
        after stricter stages fail. Disambiguate by effective date, then facility token; skip if ambiguous."""
        if primary_file_type not in ("ACP II", "AOC II", "AOC I"):
            return 0
        _fac_vals = set(FACILITY_NORM_MAP.values())
        added = 0

        def _m61_fac_tok_for_relaxed(ar: pd.Series) -> str:
            p = parse_liability_note(ar.get("Liability Note"))
            if p.get("note_category") != "Fin":
                return ""
            tok = _normalize_ln_suffix_token(p.get("source_suffix") or "")
            if not tok:
                tok = _primary_facility_match_token(
                    ar.get("Financial Institution"), ar.get("Liability Name")
                )
            if tok and tok not in _fac_vals:
                return ""
            return tok

        def _pick_primary_row(ar: pd.Series, cand: pd.DataFrame) -> int | None:
            if cand.empty:
                return None
            if len(cand) == 1:
                return int(cand.iloc[0]["_row_id_b"])
            eff_m = str(ar.get("effective_date_key") or "")
            sub_eff = cand[cand["effective_date_key"].astype(str) == eff_m]
            tok = _m61_fac_tok_for_relaxed(ar)

            if len(sub_eff) == 1:
                return int(sub_eff.iloc[0]["_row_id_b"])

            if len(sub_eff) > 1:
                if tok:
                    hit = sub_eff[sub_eff["_pri_fac_tok"].astype(str).str.strip().eq(tok)]
                    if len(hit) == 1:
                        return int(hit.iloc[0]["_row_id_b"])
                _debug_rows(
                    "TEMP DEBUG: note_deal_id_relaxed SKIP ambiguous "
                    "(multiple primary rows same deal id + eff date): "
                    f"m61_id={int(ar['_row_id_a'])} deal_id={ar.get('m61_match_key')!r} n={len(sub_eff)}"
                )
                return None

            # No primary row shares this effective date — try facility token across all candidates.
            if tok:
                hit = cand[cand["_pri_fac_tok"].astype(str).str.strip().eq(tok)]
                if len(hit) == 1:
                    return int(hit.iloc[0]["_row_id_b"])
                if len(hit) > 1:
                    _debug_rows(
                        "TEMP DEBUG: note_deal_id_relaxed SKIP ambiguous "
                        f"(facility token matches multiple primaries): m61_id={int(ar['_row_id_a'])} "
                        f"deal_id={ar.get('m61_match_key')!r} tok={tok!r} n={len(hit)}"
                    )
                    return None
            _debug_rows(
                "TEMP DEBUG: note_deal_id_relaxed SKIP no unique primary "
                f"m61_id={int(ar['_row_id_a'])} deal_id={ar.get('m61_match_key')!r} "
                f"eff_m={eff_m!r} cand={len(cand)} tok={tok!r}"
            )
            return None

        la_iter = a[
            a["_row_id_a"].isin(unmatched_a.intersection(matchable_a))
            & a["m61_match_key"].astype(str).str.strip().ne("")
        ].sort_values(["m61_match_key", "effective_date_key", "_row_id_a"])

        for _, ar in la_iter.iterrows():
            rid_a = int(ar["_row_id_a"])
            if rid_a not in unmatched_a:
                continue
            did = str(ar["m61_match_key"]).strip()
            if not did:
                continue
            cand = b[(b["_row_id_b"].isin(unmatched_b)) & (b["acp_match_key"].astype(str) == did)]
            rid_pick = _pick_primary_row(ar, cand)
            if rid_pick is None:
                continue
            if rid_pick not in unmatched_b or rid_a not in unmatched_a:
                continue
            unmatched_b.remove(rid_pick)
            unmatched_a.remove(rid_a)
            pair_rows.append(
                {
                    "_row_id_b": rid_pick,
                    "_row_id_a": rid_a,
                    "_match_stage": "note_deal_id_relaxed",
                    "_merge": "both",
                }
            )
            br = b_by_id.loc[rid_pick]
            _debug_rows(
                "TEMP DEBUG: selected pair "
                f"stage=note_deal_id_relaxed acp_id={rid_pick} m61_id={rid_a} "
                f"deal={br.get('Deal Name')!r} eff_key={br.get('effective_date_key')!r} "
                f"acp_source={br.get('Source')!r} m61_type={ar.get('Liability Type')!r} "
                f"reason=note deal id fallback after stricter Fin Inpt stages"
            )
            added += 1
        return added

    def _pair_deal_id_ignore_effective_date_fallback() -> int:
        """Pair still-unmatched rows when ``acp_match_key`` == ``m61_match_key`` (non-blank), ignoring
        effective date. Runs only after all exact / staged matchers above; does not alter prior pairs.

        Duplicate handling: for each shared deal id, unmatched primary row ids and unmatched M61 row ids
        are sorted by ``(effective_date_key, row_id)`` then paired in order (first with first, …) up to
        ``min(count_primary, count_m61)``. Any surplus rows on either side stay unmatched (existing
        Missing / M61-only behavior).
        """
        added = 0
        b_pool = b[
            b["_row_id_b"].isin(unmatched_b)
            & b["acp_match_key"].astype(str).str.strip().ne("")
        ]
        a_pool = a[
            a["_row_id_a"].isin(unmatched_a.intersection(matchable_a))
            & a["m61_match_key"].astype(str).str.strip().ne("")
        ]
        b_ids_by: dict[str, list[int]] = {}
        for _, r in b_pool.iterrows():
            k = str(r["acp_match_key"]).strip()
            b_ids_by.setdefault(k, []).append(int(r["_row_id_b"]))
        a_ids_by: dict[str, list[int]] = {}
        for _, r in a_pool.iterrows():
            k = str(r["m61_match_key"]).strip()
            a_ids_by.setdefault(k, []).append(int(r["_row_id_a"]))
        for deal_key in sorted(set(b_ids_by.keys()) & set(a_ids_by.keys())):
            b_ids = sorted(
                b_ids_by[deal_key],
                key=lambda rid: (str(b_by_id.loc[rid, "effective_date_key"]), rid),
            )
            a_ids = sorted(
                a_ids_by[deal_key],
                key=lambda rid: (str(a_by_id.loc[rid, "effective_date_key"]), rid),
            )
            n_pair = min(len(b_ids), len(a_ids))
            if len(b_ids) != len(a_ids):
                _debug_rows(
                    "TEMP DEBUG: deal_id_ignore_eff_date key="
                    f"{deal_key!r} surplus primary={max(0, len(b_ids) - n_pair)} "
                    f"surplus_m61={max(0, len(a_ids) - n_pair)} pairing={n_pair}"
                )
            for i in range(n_pair):
                rid_b, rid_a = b_ids[i], a_ids[i]
                if rid_b not in unmatched_b or rid_a not in unmatched_a:
                    continue
                unmatched_b.remove(rid_b)
                unmatched_a.remove(rid_a)
                pair_rows.append(
                    {
                        "_row_id_b": rid_b,
                        "_row_id_a": rid_a,
                        "_match_stage": "deal_id_ignore_eff_date",
                        "_merge": "both",
                    }
                )
                br = b_by_id.loc[rid_b]
                ar = a_by_id.loc[rid_a]
                _debug_rows(
                    "TEMP DEBUG: selected pair "
                    f"stage=deal_id_ignore_eff_date acp_id={rid_b} m61_id={rid_a} "
                    f"deal={br.get('Deal Name')!r} acp_eff={br.get('effective_date_key')!r} "
                    f"m61_eff={ar.get('effective_date_key')!r} "
                    f"reason=deal id / note suffix match; effective date may differ"
                )
                added += 1
        return added

    b["id_match_key"] = b["acp_match_key"].astype(str)
    a["id_match_key"] = a["m61_match_key"].astype(str)
    # AOC II / AOC I name variants are common; pair by Deal ID + Effective Date for ID-stage matching.
    if primary_file_type in ("AOC II", "AOC I"):
        b["id_match_key"] = b["acp_match_key"].astype(str) + "|" + b["effective_date_key"].astype(str)
        a["id_match_key"] = a["m61_match_key"].astype(str) + "|" + a["effective_date_key"].astype(str)
        _debug_trace_uncommons_m61_match_state(
            a,
            primary_file_type=primary_file_type,
            matchable_a=matchable_a,
            fin_note_rows=fin_note_rows,
            id_anchored=id_anchored,
            scope_lbl=scope_lbl,
            _scope_ok=_scope_ok,
        )
        _debug_trace_uncommons_pairing_keys(a, b)

    if primary_file_type in FIN_INPT_PRIMARY_TYPES:
        # Financing-first: parsed M61 liability note (Fin + fund + deal id + eff date + suffix)
        # aligned to primary Deal ID + eff date + facility/source token.
        fin_note_n = _pair_by_key("fin_acp_key", "fin_m61_key", "financing_note")
        _debug_rows(f"TEMP DEBUG: Fin Inpt financing-note composite matches={fin_note_n}")

        # ACP II: after financing_note stage, withdraw any still-unmatched LN_Sub rows from
        # matchable_a so they cannot pair with Repo/NoN/CLO ACORE rows at later stages.
        # Unmatched LN_Sub rows remain in unmatched_a and appear as "MISSING FROM ACORE",
        # which is correct — their deal has no Sub Debt counterpart on the ACORE side.
        if primary_file_type == "ACP II":
            sub_row_ids = set(
                a.loc[a["_m61_note_category"].eq("Sub"), "_row_id_a"].tolist()
            )
            matchable_a.difference_update(sub_row_ids)
            _debug_rows(
                f"TEMP DEBUG: ACP II post-financing_note — removed {len(sub_row_ids)} LN_Sub rows from matchable_a; "
                f"matchable_a now={len(matchable_a)}"
            )

        # Deal ID key merge (when Deal IDs are present); then staged matchers for remaining rows.
        lb = b[b["_row_id_b"].isin(unmatched_b)].copy()
        la = a[a["_row_id_a"].isin(unmatched_a.intersection(matchable_a))].copy()
        lb_id = lb[lb["id_match_key"].astype(str).str.strip().ne("")]
        la_id = la[la["id_match_key"].astype(str).str.strip().ne("")]
        id_n = 0
        if not lb_id.empty and not la_id.empty:
            lb_id = lb_id.sort_values(["id_match_key", "effective_date_key", "_row_id_b"]).copy()
            la_id = la_id.sort_values(["id_match_key", "effective_date_key", "_row_id_a"]).copy()
            lb_id["_rk"] = lb_id.groupby("id_match_key").cumcount()
            la_id["_rk"] = la_id.groupby("id_match_key").cumcount()
            id_pairs = lb_id[["_row_id_b", "id_match_key", "_rk"]].merge(
                la_id[["_row_id_a", "id_match_key", "_rk"]],
                on=["id_match_key", "_rk"],
                how="inner",
            )
            for _, pr in id_pairs.iterrows():
                rid_b = int(pr["_row_id_b"])
                rid_a = int(pr["_row_id_a"])
                if rid_b not in unmatched_b or rid_a not in unmatched_a:
                    continue
                unmatched_b.remove(rid_b)
                unmatched_a.remove(rid_a)
                pair_rows.append(
                    {
                        "_row_id_b": rid_b,
                        "_row_id_a": rid_a,
                        "_match_stage": "deal_id",
                        "_merge": "both",
                    }
                )
                id_n += 1
        _debug_rows(
            "TEMP DEBUG: Fin Inpt key merge acp_match_key -> m61_match_key "
            f"matched_rows={id_n}"
        )
        strict_n = _pair_by_key("strict_key", "strict_key", "strict")
        _debug_rows(f"TEMP DEBUG: Fin Inpt staged matcher strict matches={strict_n}")
        source_fac_n = _pair_by_key(
            "source_aware_facility_key", "source_aware_facility_key", "source_aware_facility"
        )
        _debug_rows(f"TEMP DEBUG: Fin Inpt staged matcher source-aware+facility matches={source_fac_n}")
        source_n = _pair_by_key("source_aware_key", "source_aware_key", "source_aware")
        _debug_rows(f"TEMP DEBUG: Fin Inpt staged matcher source-aware matches={source_n}")
        fallback_n = _pair_by_key("deal_date_key", "deal_date_key", "fallback")
        _debug_rows(f"TEMP DEBUG: Fin Inpt staged matcher fallback matches={fallback_n}")
        relaxed_ndid_n = _pair_note_deal_id_relaxed()
        _debug_rows(f"TEMP DEBUG: Fin Inpt note-deal-id relaxed matches={relaxed_ndid_n}")
    else:
        strict_n = _pair_by_key("strict_key", "strict_key", "strict")
        _debug_rows(f"TEMP DEBUG: staged matcher strict matches={strict_n}")

        # Required debugging: candidate ACP rows, source-filter match, selected row.
        cand_logged = 0
        lb_dbg = b[b["_row_id_b"].isin(unmatched_b)].copy()
        la_dbg = a[a["_row_id_a"].isin(unmatched_a.intersection(matchable_a))].copy()
        common_dd = sorted(set(lb_dbg["deal_date_key"]).intersection(set(la_dbg["deal_date_key"])))
        for dd in common_dd:
            p_cand = lb_dbg[lb_dbg["deal_date_key"] == dd]
            m_cand = la_dbg[la_dbg["deal_date_key"] == dd]
            if p_cand.empty or m_cand.empty:
                continue
            deal_name_dbg = str(p_cand.iloc[0].get("Deal Name", ""))
            if cand_logged < 40 or "geoffrey drive" in normalise_text(deal_name_dbg):
                _debug_rows(
                    "TEMP DEBUG: candidate set (source-aware) "
                    f"deal_date={dd!r} acp_candidates={len(p_cand)} m61_candidates={len(m_cand)}"
                )
                for _, pr in p_cand.iterrows():
                    _debug_rows(
                        "TEMP DEBUG:   ACP cand "
                        f"id={int(pr['_row_id_b'])} deal={pr.get('Deal Name')!r} src={pr.get('Source')!r} "
                        f"source_bucket={pr.get('source_bucket')!r} eff_key={pr.get('effective_date_key')!r}"
                    )
                for _, ar in m_cand.iterrows():
                    _debug_rows(
                        "TEMP DEBUG:   M61 cand "
                        f"id={int(ar['_row_id_a'])} deal={ar.get('Deal Name')!r} liab_type={ar.get('Liability Type')!r} "
                        f"source_bucket={ar.get('source_bucket')!r} eff_key={ar.get('effective_date_key')!r} "
                        f"facility_norm={ar.get('facility_norm')!r} undrawn={ar.get('Undrawn Capacity')!r}"
                    )
                src_match_n = len(
                    p_cand[["_row_id_b", "source_bucket"]]
                    .merge(
                        m_cand[["_row_id_a", "source_bucket"]],
                        on="source_bucket",
                        how="inner",
                    )
                )
                _debug_rows(f"TEMP DEBUG:   source-aware candidate links={src_match_n}")
                cand_logged += 1

        source_fac_n = _pair_by_key(
            "source_aware_facility_key", "source_aware_facility_key", "source_aware_facility"
        )
        _debug_rows(f"TEMP DEBUG: staged matcher source-aware+facility matches={source_fac_n}")

        source_n = _pair_by_key("source_aware_key", "source_aware_key", "source_aware")
        _debug_rows(f"TEMP DEBUG: staged matcher source-aware matches={source_n}")

        fallback_n = _pair_by_key("deal_date_key", "deal_date_key", "fallback")
        _debug_rows(f"TEMP DEBUG: staged matcher fallback matches={fallback_n}")

    deal_id_eff_fb_n = _pair_deal_id_ignore_effective_date_fallback()
    _debug_rows(
        f"TEMP DEBUG: deal_id_ignore_eff_date fallback (same deal id, date may differ) pairs={deal_id_eff_fb_n}"
    )

    # Build merged-like frame while preserving unmatched rows (outer behavior).
    for rid_b in sorted(unmatched_b):
        pair_rows.append({"_row_id_b": int(rid_b), "_row_id_a": pd.NA, "_match_stage": "none", "_merge": "left_only"})
    for rid_a in sorted(unmatched_a):
        pair_rows.append(
            {
                "_row_id_b": pd.NA,
                "_row_id_a": int(rid_a),
                "_match_stage": "none",
                "_merge": "right_only",
            }
        )

    paired_row_ids_m61: set[int] = set()
    for _pr in pair_rows:
        if _pr.get("_merge") != "both":
            continue
        _ra = _pr.get("_row_id_a")
        try:
            if _ra is None or pd.isna(_ra):
                continue
            paired_row_ids_m61.add(int(_ra))
        except (TypeError, ValueError):
            continue

    map_df = pd.DataFrame(pair_rows)
    # Keep merge key dtypes aligned even when unmatched side uses pd.NA.
    if "_row_id_b" in map_df.columns:
        map_df["_row_id_b"] = pd.to_numeric(map_df["_row_id_b"], errors="coerce").astype("Int64")
    if "_row_id_a" in map_df.columns:
        map_df["_row_id_a"] = pd.to_numeric(map_df["_row_id_a"], errors="coerce").astype("Int64")
    b["_row_id_b"] = pd.to_numeric(b["_row_id_b"], errors="coerce").astype("Int64")
    a["_row_id_a"] = pd.to_numeric(a["_row_id_a"], errors="coerce").astype("Int64")
    merged = map_df.merge(b, on="_row_id_b", how="left")
    a_pref = a.add_prefix(f"{label_a}_")
    merged = merged.merge(
        a_pref,
        left_on="_row_id_a",
        right_on=f"{label_a}__row_id_a",
        how="left",
    )
    _debug_unmatched_after_merge(merged, label_a=label_a, n=10)
    if primary_file_type == "AOC II":
        _debug_trace_uncommons_merged(merged, label_a=label_a)
    matched_rows_after_merge = int((merged["_merge"] == "both").sum()) if "_merge" in merged.columns else 0
    _debug_rows(f"TEMP DEBUG: matched rows after merge={matched_rows_after_merge}")
    if "_merge" in merged.columns:
        _debug_rows(
            "TEMP DEBUG: post-merge indicator counts "
            f"{merged['_merge'].value_counts(dropna=False).to_dict()}"
        )
    if primary_file_type in FIN_INPT_PRIMARY_TYPES:
        _debug_rows(
            "TEMP DEBUG: Fin Inpt final assembly uses liability-note parse + Deal ID / staged keys "
            "(full outer merge preserves both sides)."
        )
    else:
        _debug_rows(
            "TEMP DEBUG: Non–Fin Inpt primary: staged matchers on match_key; full outer merge."
        )
    _debug_rows(
        "Reconcile merge rows: "
        f"primary_rows={len(df_b)} m61_rows={len(df_a)} merged_rows={len(merged)} "
        f"merge_indicator={merged['_merge'].value_counts(dropna=False).to_dict()}"
    )
    _debug_rows(f"TEMP DEBUG: row count immediately after pd.merge = {len(merged)}")
    _vc = merged["_merge"].value_counts(dropna=False)
    _debug_rows(
        "TEMP DEBUG: merged['_merge'].value_counts(dropna=False) — "
        "left_only=Fin Inpt only, right_only=M61 only, both=matched key"
    )
    for _k, _v in _vc.items():
        _debug_rows(f"TEMP DEBUG:   {_k} = {_v}")
    _debug_rows(
        "[TEMP VALIDATION] primary read + M61 read + merge: "
        f"primary_rows_after_load={len(df_pri_raw)} "
        f"m61_rows_after_load={m61_row_count_after_load} "
        f"merged_rows={len(merged)} "
        f"merged['_merge'].value_counts(dropna=False)={merged['_merge'].value_counts(dropna=False).to_dict()}"
    )

    liability_extra = [
        "Current Balance",
        "Liability Note",
        "Financial Institution",
        "Maturity Date",
        "Fund Name",
        "Liability Name",
        "target",
        "Status",
    ]

    rows = []

    for _, row in merged.iterrows():
        a_deal = row.get(f"{label_a}_Deal Name")
        b_deal = row.get("Deal Name")
        in_a = not pd.isna(a_deal)
        in_b = not pd.isna(b_deal)
        spread_primary = row.get("Spread") if in_b else pd.NA
        spread_m61 = row.get(f"{label_a}_Spread") if in_a else pd.NA
        floor_primary = row.get("Floor") if in_b else pd.NA

        record = {col: row.get(col) for col in ACP_SHEET_COLS}

        if not in_b and in_a:
            record["Deal Name"] = row.get(f"{label_a}_Deal Name")
            record["Facility"] = row.get(f"{label_a}_Liability Name")
            record["Note Name"] = row.get(f"{label_a}_Liability Note")
            record["Pledge"] = row.get(f"{label_a}_Pledge")
            record["Pledge Date"] = row.get(f"{label_a}_Pledge Date")
            record["Effective Date"] = row.get(f"{label_a}_Effective Date")
            record["Maturity Date"] = row.get(f"{label_a}_Maturity Date")

        facility_raw = record.get("Facility") or row.get(f"{label_a}_Liability Name")
        facility_norm = normalise_facility(facility_raw)
        record["Financial Line"] = f"{record.get('Deal Name', '')} & {str(facility_norm).upper()}"
        record["match_key"] = (
            row.get("match_key")
            if not pd.isna(row.get("match_key"))
            else row.get(f"{label_a}_match_key")
        )
        record["Liability Name (M61)"] = (
            row.get(f"{label_a}_Liability Name") if in_a else pd.NA
        )
        record["Facility Norm (ACP)"] = (
            _facility_norm_for_debug_cell(row.get("Facility")) if in_b else pd.NA
        )
        record["Facility Norm (M61)"] = (
            _facility_norm_for_debug_cell(row.get(f"{label_a}_Liability Name"))
            if in_a
            else pd.NA
        )

        # Filter-friendly Source: only for M61-only rows, derive from M61 fields.
        # Keep matched/primary rows strictly primary-driven for Source Type (ACORE).
        src_pri = _stripped_nonempty_str(record.get("Source"))
        if not src_pri and in_a and not in_b:
            lt = _stripped_nonempty_str(row.get(f"{label_a}_Liability Type"))
            fi = _stripped_nonempty_str(row.get(f"{label_a}_Financial Institution"))
            parts = [p for p in (lt, fi) if p]
            if parts:
                record["Source"] = " | ".join(parts)

        deal_id_acp = row.get("deal_id_value") if in_b else ""
        note_suffix_m61 = row.get(f"{label_a}_liability_note_suffix") if in_a else ""
        deal_id_key = normalise_deal_id_key(deal_id_acp)
        note_suffix_key = normalise_deal_id_key(note_suffix_m61)
        record["Deal ID (ACP)"] = deal_id_acp if deal_id_key else pd.NA
        record["Deal ID Match Key (ACP)"] = deal_id_key if deal_id_key else pd.NA
        record["Liability Note Suffix (M61)"] = note_suffix_m61 if note_suffix_key else pd.NA
        record["M61 Extracted Deal ID"] = note_suffix_key if note_suffix_key else pd.NA
        record["ID Match Result"] = row.get("_merge")
        record["Match Stage"] = row.get("_match_stage")
        if match_diagnostics:
            record.update(
                _match_diagnostic_fields(row, label_a, in_a=in_a, in_b=in_b)
            )

        record["Target Advance Rate (M61)"] = (
            row.get(f"{label_a}_Target Advance Rate") if in_a else pd.NA
        )
        record["Current Advance Rate (M61 Raw)"] = (
            row.get(f"{label_a}_Current Advance Rate") if in_a else pd.NA
        )
        record["Deal Level Advance Rate (M61 Raw)"] = (
            _m61_deal_level_advance_rate(row, label_a) if in_a else pd.NA
        )
        # Temporary trace columns for Advance Rate source debugging.
        record["Advance Rate (ACP) Debug"] = record.get("Advance Rate")
        record["Raw Target Advance Rate from M61"] = record["Target Advance Rate (M61)"]
        record["Raw Current Advance Rate from M61"] = record["Current Advance Rate (M61 Raw)"]
        record["Raw Deal Level Advance Rate from M61"] = record["Deal Level Advance Rate (M61 Raw)"]
        record["Liability Type (M61 Raw)"] = (
            row.get(f"{label_a}_Liability Type") if in_a else pd.NA
        )

        in_liability_raw = row.get(f"{label_a}_in_liability")
        in_liability_value = "" if pd.isna(in_liability_raw) else str(in_liability_raw).strip().lower()
        only_target_from_invis = in_liability_value == "invis"

        key_field_statuses = []

        for b_field, a_field, ctype in COMPARE_FIELDS:
            val_a = row.get(f"{label_a}_{a_field}")
            val_b = row.get(b_field)
            liability_label = LIABILITY_VALUE_LABELS.get(a_field, f"{a_field} (M61)")
            if b_field == "Spread":
                # Explicit source-scoped values to avoid same-name Spread collisions after merge.
                val_a = spread_m61
                val_b = spread_primary

            if only_target_from_invis and a_field != "target":
                record[liability_label] = pd.NA
                record[f"{b_field} Status"] = FIELD_STATUS_MISSING_M61
                if b_field in RECON_STATUS_FIELDS:
                    key_field_statuses.append("MISSING")
                continue

            # Advance Rate comparison basis:
            # - ACP II / ACP I / AOC II / AOC I: Sale-type liabilities → Deal Level advance; else Target.
            # - ACORE (ACP III): keep Target-only behavior.
            # - Other flows retain legacy sale-type fallback to deal-level advance.
            if b_field == "Advance Rate":
                if primary_file_type in SALE_DEALLEVEL_PRIMARY_TYPES:
                    fund_name = row.get(f"{label_a}_Fund Name") if in_a else ""
                    liab_type = row.get(f"{label_a}_Liability Type") if in_a else ""
                    use_deal_level_adv = _is_sale_type_fund_or_deal(
                        fund_name=fund_name,
                        liability_type=liab_type,
                    )
                    if use_deal_level_adv:
                        compare_val = _m61_deal_level_advance_rate(row, label_a) if in_a else pd.NA
                        record["Advance Rate (M61)"] = _normalize_aoc_ii_m61_adv_value(compare_val)
                        record["Advance Rate Source (M61)"] = "Deal Level Advance Rate"
                    else:
                        compare_val = row.get(f"{label_a}_Target Advance Rate") if in_a else pd.NA
                        record["Advance Rate (M61)"] = _normalize_aoc_ii_m61_adv_value(compare_val)
                        record["Advance Rate Source (M61)"] = "Target Advance Rate"
                    val_a = record["Advance Rate (M61)"]
                    record["Final Advance Rate (M61)"] = record.get("Advance Rate (M61)")
                elif primary_file_type in FIN_INPT_PRIMARY_TYPES:
                    compare_val = row.get(f"{label_a}_Target Advance Rate") if in_a else pd.NA
                    record["Advance Rate (M61)"] = compare_val
                    record["Advance Rate Source (M61)"] = "Target Advance Rate"
                    val_a = compare_val
                    record["Final Advance Rate (M61)"] = record.get("Advance Rate (M61)")
                else:
                    fund_name = row.get(f"{label_a}_Fund Name") if in_a else ""
                    liab_type = row.get(f"{label_a}_Liability Type") if in_a else ""
                    use_deal_level_adv = _is_sale_type_fund_or_deal(
                        fund_name=fund_name,
                        liability_type=liab_type,
                    )
                    if use_deal_level_adv:
                        compare_val = _m61_deal_level_advance_rate(row, label_a) if in_a else pd.NA
                        record["Advance Rate (M61)"] = compare_val
                        record["Advance Rate Source (M61)"] = "Deal Level Advance Rate"
                        val_a = compare_val
                    else:
                        compare_val = row.get(f"{label_a}_Target Advance Rate") if in_a else pd.NA
                        record["Advance Rate (M61)"] = compare_val
                        record["Advance Rate Source (M61)"] = "Target Advance Rate"
                        val_a = compare_val
                    record["Final Advance Rate (M61)"] = record.get("Advance Rate (M61)")
            else:
                record[liability_label] = val_a

            if b_field == "Effective Date":
                if not in_b:
                    ed_status = FIELD_STATUS_MISSING_ACORE
                elif not in_a:
                    ed_status = FIELD_STATUS_MISSING_M61
                else:
                    ed_status = compare_effective_date_status(val_a, val_b)
                record[f"{b_field} Status"] = ed_status
                if b_field in RECON_STATUS_FIELDS:
                    key_field_statuses.append(_recon_token_for_effective_date_status(ed_status))
                continue

            if not in_a and not in_b:
                status = FIELD_STATUS_MISSING_BOTH
            elif not in_b:
                status = FIELD_STATUS_MISSING_ACORE
            elif not in_a:
                status = FIELD_STATUS_MISSING_M61
            else:
                if b_field == "Spread":
                    l_ok = _has_compare_value(val_a)
                    p_ok = _has_compare_value(val_b)
                    if not l_ok and not p_ok:
                        status = FIELD_STATUS_MISSING_BOTH
                    elif not p_ok and l_ok:
                        status = FIELD_STATUS_MISSING_ACORE
                    elif p_ok and not l_ok:
                        status = FIELD_STATUS_MISSING_M61
                    else:
                        q_a = _spread_percent_quantized_m61_compare(val_a)
                        q_b = _spread_percent_quantized_m61_compare(val_b)
                        if q_a is not None and q_b is not None:
                            status = (
                                FIELD_STATUS_MATCH
                                if q_a == q_b
                                else FIELD_STATUS_MISMATCH
                            )
                        else:
                            status = compare_liability_primary_status(val_a, val_b, ctype)
                elif b_field == "Advance Rate":
                    n_a = _coerce_rate_fraction(val_a)
                    n_b = _coerce_rate_fraction(val_b)
                    if n_a is not None and n_b is not None:
                        status = (
                            FIELD_STATUS_MATCH
                            if abs(n_a - n_b) <= RATE_STATUS_TOLERANCE
                            else FIELD_STATUS_MISMATCH
                        )
                    else:
                        status = compare_liability_primary_status(val_a, val_b, ctype)
                else:
                    status = compare_liability_primary_status(val_a, val_b, ctype)

            record[f"{b_field} Status"] = status
            if b_field in RECON_STATUS_FIELDS:
                key_field_statuses.append(_recon_token_for_compare_status(status))

        for a_col in liability_extra:
            business_col = LIABILITY_VALUE_LABELS.get(a_col, f"{a_col} (M61)")
            if business_col in record:
                continue
            record[business_col] = (
                pd.NA if (only_target_from_invis and a_col != "target") else row.get(f"{label_a}_{a_col}")
            )

        # M61 Note Category: prefer parsed Liability Note (LN_Fin/LN_Sub/LN_Eq), fallback to Liability Type.
        record["M61 Note Category"] = categorize_m61_note_category(
            row.get(f"{label_a}_Liability Note"),
            record.get("Liability Type (M61 Raw)"),
            record.get("Source"),
            primary_file_type=primary_file_type,
        )

        pliab = pd.NA if only_target_from_invis else (row.get(f"{label_a}_Pledge") if in_a else pd.NA)
        pacp = row.get("Pledge") if in_b else pd.NA
        pdt_liab = pd.NA if only_target_from_invis else (row.get(f"{label_a}_Pledge Date") if in_a else pd.NA)
        pdt_acp = row.get("Pledge Date") if in_b else pd.NA

        record["Pledge (ACP)"] = pacp
        record["Pledge (M61)"] = pliab
        record["Pledge Date (ACP)"] = pdt_acp
        record["Pledge Date (M61)"] = pdt_liab
        record["Pledge Date Status"] = compare_pledge_date_status(
            val_liability=pdt_liab,
            val_acp=pdt_acp,
        )

        m61_fund_s = _stripped_nonempty_str(row.get(f"{label_a}_Fund Name")) if in_a else None
        # Prefer M61 export Fund Name; when missing, use M61-style display name (not export_label).
        if m61_fund_s:
            record["Fund"] = m61_fund_s
        else:
            record["Fund"] = fund_fallback_display
        record["Effective Date (ACP)"] = record.get("Effective Date") if in_b else pd.NA
        record["Advance Rate (ACP)"] = record.get("Advance Rate") if in_b else pd.NA
        record["Spread (ACP)"] = spread_primary
        record["Spread (M61)"] = (
            pd.NA if only_target_from_invis else spread_m61
        )
        record["Undrawn Capacity (ACP)"] = record.get("Current Undrawn Capacity") if in_b else pd.NA
        record["Undrawn Capacity (M61)"] = record.get("Current Undrawn Capacity (M61)")

        # ACP-side values
        record["Index Floor (ACP)"] = _normalize_index_floor_value(floor_primary)
        record["Index Name (ACP)"] = pd.NA
        record["Recourse % (ACP)"] = record.get("Recourse %") if in_b else pd.NA

        record["Undrawn Capacity Status"] = compare_liability_primary_status(
            record.get("Undrawn Capacity (M61)"),
            record.get("Undrawn Capacity (ACP)"),
            "numeric",
        )

        # Liability-side values
        if only_target_from_invis or not in_a:
            record["Index Floor (M61)"] = pd.NA
            record["Index Name (M61)"] = pd.NA
            record["Recourse % (M61)"] = pd.NA
        else:
            ix_fl = _normalize_index_floor_value(_merged_liab_col(row, label_a, "IndexFloor"))
            ix_nm = _merged_liab_col(row, label_a, "IndexName")
            if pd.notna(ix_nm):
                ix_nm = str(ix_nm).strip() or pd.NA

            record["Index Floor (M61)"] = ix_fl
            record["Index Name (M61)"] = ix_nm
            record["Recourse % (M61)"] = _first_liab_recourse_value(row, label_a)

        record["Index Floor Status"] = compare_optional(
            record.get("Index Floor (M61)"),
            record.get("Index Floor (ACP)"),
            "text",
        )
        record["Index Name Status"] = compare_optional(
            record.get("Index Name (M61)"),
            record.get("Index Name (ACP)"),
            "text",
        )
        rec_m61 = record.get("Recourse % (M61)")
        rec_acp = record.get("Recourse % (ACP)")
        record["Recourse % Status"] = compare_optional(
            rec_m61,
            rec_acp,
            "numeric",
        )

        record["recon_status"] = _derive_business_recon_status(
            record,
            in_a=in_a,
            in_b=in_b,
            primary_file_type=primary_file_type,
        )
        # Set last so nothing in the row payload overwrites; drives UI "File Source" column.
        record["File Source"] = _file_source_label_from_sides(in_a=in_a, in_b=in_b)
        rows.append(record)

    _out_cols = list(RECON_ORDERED_COLS)
    if match_diagnostics:
        _out_cols = _out_cols + list(MATCH_DIAGNOSTIC_COLUMNS)

    def _target_22203_snapshot(df: pd.DataFrame) -> dict[str, object]:
        if df is None or df.empty:
            return {"count": 0, "rows": []}
        m = pd.Series(True, index=df.index)
        if "Deal Name" in df.columns:
            m &= df["Deal Name"].fillna("").astype(str).str.strip().str.lower().eq("block 21 san mateo")
        if "Facility" in df.columns:
            m &= df["Facility"].fillna("").astype(str).str.strip().str.lower().eq("tbk bank")
        if "Deal ID Match Key (ACP)" in df.columns:
            did = df["Deal ID Match Key (ACP)"].fillna("").astype(str).str.strip()
            if did.ne("").any():
                m &= did.eq("222203")
        elif "Deal ID (ACP)" in df.columns:
            did2 = df["Deal ID (ACP)"].map(lambda v: normalise_deal_id_key(v))
            if did2.fillna("").ne("").any():
                m &= did2.eq("222203")
        if "Effective Date (ACP)" in df.columns:
            m &= pd.to_datetime(df["Effective Date (ACP)"], errors="coerce").dt.strftime("%Y-%m-%d").eq("2022-08-22")
        if "Pledge Date (ACP)" in df.columns:
            m &= pd.to_datetime(df["Pledge Date (ACP)"], errors="coerce").dt.strftime("%Y-%m-%d").eq("2023-08-31")
        if "Source" in df.columns:
            m &= df["Source"].fillna("").astype(str).str.lower().str.contains(r"\bsale\b", regex=True, na=False)
        cols = [
            c
            for c in (
                "Deal ID (ACP)",
                "Deal ID Match Key (ACP)",
                "Deal Name",
                "Source",
                "File Source",
                "Effective Date (ACP)",
                "Effective Date (M61)",
                "Pledge Date (ACP)",
                "Pledge Date (M61)",
                "Undrawn Capacity (M61)",
                "Advance Rate (M61)",
                "Spread (M61)",
                "Index Floor (M61)",
            )
            if c in df.columns
        ]
        hit = df.loc[m, cols].copy() if cols else df.loc[m].copy()
        return {"count": int(len(hit)), "rows": hit.astype(str).to_dict("records")}

    df_out = pd.DataFrame(rows).reindex(columns=_out_cols).reset_index(drop=True)
    _target_stage1 = _target_22203_snapshot(df_out)
    df_out = _surface_related_m61_for_acore_only_rows(
        df_out,
        a,
        paired_row_ids_m61=paired_row_ids_m61,
        primary_file_type=primary_file_type,
    )
    _target_stage2 = _target_22203_snapshot(df_out)
    # Hard guard for Fin Inpt runs that remain Target-only (ACORE / ACP III).
    # SALE_DEALLEVEL_PRIMARY_TYPES pick Target vs Deal Level per row upstream.
    if (
        primary_file_type in FIN_INPT_PRIMARY_TYPES
        and primary_file_type not in SALE_DEALLEVEL_PRIMARY_TYPES
        and "Advance Rate (M61)" in df_out.columns
        and "Target Advance Rate (M61)" in df_out.columns
    ):
        df_out["Advance Rate (M61)"] = df_out["Target Advance Rate (M61)"]
        if "Advance Rate Source (M61)" in df_out.columns:
            has_target = df_out["Target Advance Rate (M61)"].notna()
            df_out.loc[has_target, "Advance Rate Source (M61)"] = "Target Advance Rate"

    df_out = _ensure_file_source_populated(df_out)

    _debug_cols = [
        "Deal Name",
        "Fund",
        "Fund (M61)",
        "Deal ID (ACP)",
        "Facility",
        "Liability Name (M61)",
        "Effective Date (ACP)",
        "Effective Date (M61)",
        "Match Stage",
        "File Source",
        "Target Advance Rate (M61)",
        "Advance Rate (ACP)",
        "recon_status",
    ]
    _present_dbg = [c for c in _debug_cols if c in df_out.columns]
    if _present_dbg:
        _debug_rows(f"TEMP DEBUG: post-assembly key output snapshot cols={_present_dbg!r}")
        for i, (_, rr) in enumerate(df_out.loc[:, _present_dbg].head(80).iterrows(), start=1):
            _debug_rows("TEMP DEBUG: OUT " + f"#{i} " + " | ".join(f"{c}={rr.get(c)!r}" for c in _present_dbg))
    _debug_rows(f"Reconciliation output rows (df_out)={len(df_out)}")

    # ── Temporary per-row diagnostic for FIN_INPT runs ──────────────────────────
    # Prints to stderr so it appears in the Streamlit server log regardless of
    # RECON_DEBUG.  Shows every ACORE-driven row with match metadata so mismatches
    # between ACORE and M61 sides are immediately visible.  Remove once stable.
    if primary_file_type in FIN_INPT_PRIMARY_TYPES:
        import sys as _sys
        _dbg_cols = [
            "Deal ID (ACP)", "Deal Name", "Effective Date",
            "M61 Extracted Deal ID", "Liability Name (M61)", "Effective Date (M61)",
            "Liability Type (M61 Raw)", "File Source", "Match Stage", "match_key",
            "Target Advance Rate (M61)", "Spread (M61)",
        ]
        _avail = [c for c in _dbg_cols if c in df_out.columns]
        _acore_mask = (
            df_out["File Source"].isin([FILE_SOURCE_BOTH, FILE_SOURCE_ACORE_ONLY])
            if "File Source" in df_out.columns
            else pd.Series([True] * len(df_out))
        )
        _dbg_df = df_out.loc[_acore_mask, _avail].copy()
        print(
            f"\n{'='*70}\n"
            f"[{primary_file_type} ROW DEBUG] ACORE-driven rows: {len(_dbg_df)}\n"
            f"{_dbg_df.to_string(index=True)}\n"
            f"{'='*70}",
            file=_sys.stderr,
        )
    # ── End temporary diagnostic ─────────────────────────────────────────────────

    adv_rate_col = f"{p_cfg['display_name']} Advance Rate"
    adv_rows = []
    for _, row in merged.iterrows():
        a_deal = row.get(f"{label_a}_Deal Name")
        b_deal = row.get("Deal Name")
        in_a = not pd.isna(a_deal)
        in_b = not pd.isna(b_deal)
        deal = b_deal if in_b else (a_deal if in_a else pd.NA)
        acp_adv = row["Advance Rate"] if in_b and not pd.isna(row.get("Advance Rate")) else pd.NA

        for col in LIABILITY_ADVANCE_RATE_COLUMNS:
            pk = f"{label_a}_{col}"
            liab_val = row.get(pk) if pk in row.index else pd.NA
            if not in_a and not in_b:
                result = FIELD_STATUS_MISSING_BOTH
            elif not in_b:
                result = FIELD_STATUS_MISSING_ACORE
            elif not in_a:
                result = FIELD_STATUS_MISSING_M61
            else:
                result = compare_liability_primary_status(liab_val, acp_adv, "numeric")
            adv_rows.append(
                {
                    "Deal": deal,
                    adv_rate_col: acp_adv,
                    "M61 Column": col,
                    "M61 Value": liab_val,
                    "Result": result,
                }
            )

    df_adv = pd.DataFrame(adv_rows).reset_index(drop=True)
    _target_stage_final = _target_22203_snapshot(df_out)
    _debug_rows(
        "UNDRAWN TRACE 4 (final reconciliation df before UI): "
        f"records={_target_stage_final.get('rows', [])!r}"
    )

    LAST_RECON_DIAGNOSTICS = {
        "primary_file_type": primary_file_type,
        "raw_primary_rows_loaded": int(len(df_pri_raw)),
        "raw_m61_rows_loaded": int(m61_load_diag.get("m61_raw_rows_loaded", len(raw_m61))),
        "m61_rows_after_fund_filter_for_primary": int(a["_m61_in_scope"].sum()) if "_m61_in_scope" in a.columns else int(len(a)),
        "m61_rows_after_filters": int(m61_load_diag.get("m61_rows_after_liability_type_filter", len(raw_m61))),
        "m61_rows_matching_acp_identifiers": int(
            a["_id_in_primary"].sum()
        ) if "_id_in_primary" in a.columns else 0,
        "m61_rows_nonblank_liability_note": m61_rows_nonblank_note,
        "m61_rows_extracted_deal_id_nonblank": m61_rows_extracted_nonblank,
        "m61_rows_extracted_deal_id_in_acp": m61_rows_extracted_in_acp,
        "matched_rows_after_merge": matched_rows_after_merge,
        "m61_rows_matching_target_fund_labels": int(
            m61_load_diag.get("m61_rows_matching_target_fund_labels", len(raw_m61))
        ),
        "m61_rows_excluded_by_type_filter": int(
            m61_load_diag.get("m61_rows_excluded_by_liability_type_filter", len(excluded_by_type_df))
        ),
        "primary_rows_after_exclusions": int(len(df_b)),
        "final_reconciliation_rows": int(len(df_out)),
        "reconciliation_basis": (
            "fin_inpt_left_anchored_by_liability_note_and_deal_id"
            if primary_file_type in FIN_INPT_PRIMARY_TYPES
            else "outer_merge_preserving_both_files"
        ),
        "m61_id_extraction_preview": a.loc[:, m61_dbg_cols].head(50).to_dict("records"),
        # Note category and liability type breakdowns for UI debug panel.
        "m61_note_category_counts": (
            df_out["M61 Note Category"]
            .fillna("Other")
            .astype(str)
            .value_counts(dropna=False)
            .to_dict()
            if "M61 Note Category" in df_out.columns
            else {}
        ),
        "m61_liability_type_counts": (
            df_out["Liability Type (M61 Raw)"]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", "Blank/Unknown")
            .value_counts(dropna=False)
            .to_dict()
            if "Liability Type (M61 Raw)" in df_out.columns
            else {}
        ),
        "target_22203_stage1_count": int(_target_stage1.get("count", 0)),
        "target_22203_stage1_rows": _target_stage1.get("rows", []),
        "target_22203_stage2_count": int(_target_stage2.get("count", 0)),
        "target_22203_stage2_rows": _target_stage2.get("rows", []),
        "target_22203_stage_final_count": int(_target_stage_final.get("count", 0)),
        "target_22203_stage_final_rows": _target_stage_final.get("rows", []),
    }
    LAST_RECON_CONTEXT = {
        "primary_file_type": primary_file_type,
        "df_primary_raw": df_pri_raw.copy(),
        "df_m61_raw": raw_m61.copy(),
        "df_primary_matchable": df_b.copy(),
        "df_m61_matchable": df_a.copy(),
    }
    _debug_rows(f"Diagnostics snapshot: {LAST_RECON_DIAGNOSTICS}")
    if return_diagnostics:
        return df_out, df_adv, excluded_by_type_df, dict(LAST_RECON_DIAGNOSTICS)
    return df_out, df_adv, excluded_by_type_df


# --------------------------------------------------
# 8. EXCEL STYLING HELPERS
# --------------------------------------------------
HEADER_BG = "1F3864"
SUBHDR_BG = "2F5597"
MATCH_BG = "C6EFCE"
MISMATCH_BG = "FFC7CE"
MISSING_BG = "FFEB9C"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# Data rows (reconciliation grid from row 5): taller rows + wrap so status text is visible.
EXCEL_RECON_DATA_ROW_HEIGHT = 44


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _hdr_font(size=9, color=WHITE, bold=True):
    return Font(name="Arial", size=size, bold=bold, color=color)


def _body_font(size=9, color="000000", bold=False):
    return Font(name="Arial", size=size, bold=bold, color=color)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


GROUP_COLORS = {"ID": "2E4057", "ACP": "1A5276", "LIB": "1B4F72", "STATUS": "4A235A"}

COL_DEFS = [
    ("Fund", 18, False, False, "ID"),
    ("Deal Name", 22, False, False, "ID"),
    ("Facility", 14, False, False, "ID"),
    ("Financial Line", 32, False, False, "ID"),
    ("Note Name", 22, False, False, "ID"),
    ("Liability Note (M61)", 26, False, False, "LIB"),
    ("Deal ID (ACP)", 14, False, False, "ACP"),
    ("Liability Note Suffix (M61)", 20, False, False, "LIB"),
    ("Source", 9, False, False, "ID"),
    ("File Source", 24, False, False, "ID"),
    ("Effective Date (ACP)", 14, False, True, "ACP"),
    ("Effective Date (M61)", 16, False, True, "LIB"),
    ("Pledge Date (ACP)", 14, False, True, "ACP"),
    ("Pledge Date (M61)", 16, False, True, "LIB"),
    ("Advance Rate (ACP)", 13, True, False, "ACP"),
    ("Advance Rate (M61)", 16, True, False, "LIB"),
    ("Spread (ACP)", 11, True, False, "ACP"),
    ("Spread (M61)", 14, True, False, "LIB"),
    ("Undrawn Capacity (ACP)", 17, False, False, "ACP"),
    ("Undrawn Capacity (M61)", 20, False, False, "LIB"),
    ("Index Floor (ACP)", 15, False, False, "ACP"),
    ("Index Floor (M61)", 15, False, False, "LIB"),
    ("Index Name (ACP)", 22, False, False, "ACP"),
    ("Index Name (M61)", 22, False, False, "LIB"),
    ("Recourse % (ACP)", 12, True, False, "ACP"),
    ("Recourse % (M61)", 12, True, False, "LIB"),
    ("Effective Date Status", 24, False, False, "STATUS"),
    ("Pledge Date Status", 24, False, False, "STATUS"),
    ("Advance Rate Status", 24, False, False, "STATUS"),
    ("Spread Status", 22, False, False, "STATUS"),
    ("Undrawn Capacity Status", 28, False, False, "STATUS"),
    ("Index Floor Status", 22, False, False, "STATUS"),
    ("Index Name Status", 22, False, False, "STATUS"),
    ("Recourse % Status", 22, False, False, "STATUS"),
    ("Overall Recon Status", 80, False, False, "STATUS"),
]

RECON_COL_MAP = list(RECON_ORDERED_COLS)


def _status_cell(cell, val):
    v = str(val).upper()
    if v in ("N/A", "—", "-"):
        cell.fill = _fill(LIGHT_GRAY)
        cell.font = _body_font(color="888888")
    elif "DIFFERENT" in v or "MISMATCH" in v or "NO MATCH" in v:
        cell.fill = _fill(MISMATCH_BG)
        cell.font = _body_font(bold=True, color="9C0006")
    elif "BOTH MISSING" in v or "MISSING" in v:
        cell.fill = _fill(MISSING_BG)
        cell.font = _body_font(bold=True, color="7D6608")
    elif "MATCH" in v and "MIS" not in v:
        cell.fill = _fill(MATCH_BG)
        cell.font = _body_font(bold=True, color="375623")
    else:
        cell.fill = _fill(LIGHT_GRAY)
        cell.font = _body_font(color="888888")
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = BORDER


EXCEL_WIDE_READABILITY_COLUMNS = frozenset(
    {
        "File Source",
        "Effective Date Status",
        "Pledge Date Status",
        "Advance Rate Status",
        "Spread Status",
        "Undrawn Capacity Status",
        "Index Floor Status",
        "Index Name Status",
        "Recourse % Status",
        "Overall Recon Status",
    }
)


def _autofit_recon_columns_for_readability(ws, *, data_start_row: int, data_end_row: int) -> None:
    """Widen key columns based on content length while preserving readable minimums."""
    for col_idx, (hdr, base_w, _pct, _dt, _grp) in enumerate(COL_DEFS, start=1):
        if hdr not in EXCEL_WIDE_READABILITY_COLUMNS:
            continue
        max_len = len(str(_excel_header_for_primary(hdr, "")))
        for rr in range(data_start_row, data_end_row + 1):
            v = ws.cell(row=rr, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        # Keep widths in a practical range: wide enough to read, not so wide sheet becomes unusable.
        computed = min(90, max(base_w, int(max_len * 0.9) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = computed


def _fmt_date(v):
    if pd.isna(v) or str(v) in ("NaT", "nan", ""):
        return None
    try:
        return pd.to_datetime(v).date()
    except Exception:
        return None


def _fmt_num(v):
    return _coerce_numeric_value(v)


def _fmt_str_cell(v):
    if pd.isna(v) or str(v).strip().lower() in ("", "nan", "<na>", "nat"):
        return ""
    return str(v).strip()


def _fmt_index_floor_cell(v):
    if pd.isna(v) or str(v).strip().lower() in ("", "nan", "<na>"):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        s = str(v).strip()
        return s if s else None


def _fmt_index_name_cell(v):
    if pd.isna(v) or str(v).strip().lower() in ("", "nan", "<na>"):
        return None
    s = str(v).strip()
    return s if s else None


def _fmt_status(v):
    if pd.isna(v) or str(v) in ("nan", "NaN", ""):
        return "N/A"
    return str(v)


def _row_bg(recon_status):
    rs = str(recon_status).upper()
    if "MATCH WITH DIFFERENCES" in rs or "MISMATCH" in rs or "DIFFERENCE" in rs:
        return MISMATCH_BG
    if "MATCH WITH MISSING FIELDS" in rs or "MISSING" in rs:
        return MISSING_BG
    if "MATCH" in rs and "MIS" not in rs:
        return MATCH_BG
    return WHITE


def primary_workbook_context(primary_file_type: str = "ACORE") -> dict:
    cfg = get_primary_config(primary_file_type)
    dn = cfg["display_name"]
    as_of = datetime.now().strftime("%m/%d/%Y")
    return {
        "title": f"Financing Line Reconciliation  —  {dn}  |  As of {as_of}",
        "subtitle": (
            f"Primary Source: {cfg['model_descriptor']}  |  "
            "Comparison Source: M61  |  "
            "Target from M61 file only"
        ),
        "group_primary_header": cfg["primary_group_header"],
        "legend_match_detail": f"All key fields align between {dn} and M61",
        "legend_primary_only_label": cfg["primary_only_legend_label"],
        "legend_primary_only_detail": f"Record found only in {dn} model; not yet in M61",
        "excel_primary_column_suffix": cfg["excel_primary_column_suffix"],
    }


def _excel_header_for_primary(hdr: str, column_suffix: str) -> str:
    if column_suffix == "ACP":
        return hdr
    return hdr.replace("(ACP)", f"({column_suffix})")


def build_workbook(df_recon, primary_file_type: str = "ACORE"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"
    wb_ctx = primary_workbook_context(primary_file_type)
    col_suffix = wb_ctx["excel_primary_column_suffix"]

    total_cols = len(COL_DEFS)
    last_col = get_column_letter(total_cols)

    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = wb_ctx["title"]
    c.font = Font(name="Arial", size=13, bold=True, color=WHITE)
    c.fill = _fill(HEADER_BG)
    c.alignment = _center()
    ws.row_dimensions[1].height = 24

    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = wb_ctx["subtitle"]
    c.font = Font(name="Arial", size=9, italic=True, color=WHITE)
    c.fill = _fill(SUBHDR_BG)
    c.alignment = _center()
    ws.row_dimensions[2].height = 16

    grp_labels = {
        "ID": "IDENTIFICATION",
        "ACP": wb_ctx["group_primary_header"],
        "LIB": "M61 — COMPARISON DATA",
        "STATUS": "RECONCILIATION STATUS",
    }

    runs = []
    run_start = 1
    current_grp = COL_DEFS[0][4]

    for col_idx in range(2, len(COL_DEFS) + 1):
        grp = COL_DEFS[col_idx - 1][4]
        if grp != current_grp:
            runs.append((current_grp, run_start, col_idx - 1))
            run_start = col_idx
            current_grp = grp
    runs.append((current_grp, run_start, len(COL_DEFS)))

    for grp, c_start, c_end in runs:
        if c_start != c_end:
            ws.merge_cells(start_row=3, start_column=c_start, end_row=3, end_column=c_end)

        cell = ws.cell(row=3, column=c_start)
        cell.value = grp_labels[grp]
        cell.font = _hdr_font()
        cell.fill = _fill(GROUP_COLORS[grp])
        cell.alignment = _center()
        cell.border = BORDER

        for cnum in range(c_start, c_end + 1):
            ws.cell(row=3, column=cnum).fill = _fill(GROUP_COLORS[grp])
            ws.cell(row=3, column=cnum).border = BORDER

    ws.row_dimensions[3].height = 35

    for i, (hdr, w, _, _, grp) in enumerate(COL_DEFS, 1):
        cell = ws.cell(row=4, column=i)
        cell.value = _excel_header_for_primary(hdr, col_suffix)
        cell.font = _hdr_font()
        cell.fill = _fill(GROUP_COLORS[grp])
        cell.alignment = _center()
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[4].height = 36

    for data_row_idx, (_, row) in enumerate(df_recon.iterrows(), 5):
        row_bg = _row_bg(row.get("recon_status", ""))

        vals = [
            _fmt_str_cell(row.get("Fund")),
            _fmt_str_cell(row.get("Deal Name", "")),
            _fmt_str_cell(row.get("Facility", "")),
            _fmt_str_cell(row.get("Financial Line", "")),
            _fmt_str_cell(row.get("Note Name", "")),
            _fmt_str_cell(row.get("Liability Note (M61)", "")),
            _fmt_str_cell(row.get("Deal ID (ACP)", "")),
            _fmt_str_cell(row.get("Liability Note Suffix (M61)", "")),
            _fmt_str_cell(row.get("Source", "")),
            _fmt_str_cell(row.get("File Source", "")),
            _fmt_date(row.get("Effective Date (ACP)")),
            _fmt_date(row.get("Effective Date (M61)")),
            _fmt_date(row.get("Pledge Date (ACP)")),
            _fmt_date(row.get("Pledge Date (M61)")),
            _fmt_num(row.get("Advance Rate (ACP)")),
            _fmt_num(row.get("Advance Rate (M61)")),
            _fmt_num(row.get("Spread (ACP)")),
            _fmt_num(row.get("Spread (M61)")),
            _fmt_num(row.get("Undrawn Capacity (ACP)")),
            _fmt_num(row.get("Undrawn Capacity (M61)")),
            _fmt_index_floor_cell(row.get("Index Floor (ACP)")),
            _fmt_index_floor_cell(row.get("Index Floor (M61)")),
            _fmt_index_name_cell(row.get("Index Name (ACP)")),
            _fmt_index_name_cell(row.get("Index Name (M61)")),
            _fmt_num(row.get("Recourse % (ACP)")),
            _fmt_num(row.get("Recourse % (M61)")),
            _fmt_status(row.get("Effective Date Status")),
            _fmt_status(row.get("Pledge Date Status")),
            _fmt_status(row.get("Advance Rate Status")),
            _fmt_status(row.get("Spread Status")),
            _fmt_status(row.get("Undrawn Capacity Status")),
            _fmt_status(row.get("Index Floor Status")),
            _fmt_status(row.get("Index Name Status")),
            _fmt_status(row.get("Recourse % Status")),
            _fmt_status(row.get("recon_status")),  # written under "Overall Recon Status" header
        ]

        for col_idx, (val, (hdr, _w, pct, dt, grp)) in enumerate(zip(vals, COL_DEFS), 1):
            cell = ws.cell(row=data_row_idx, column=col_idx)
            cell.value = val
            cell.border = BORDER

            if grp == "STATUS":
                _status_cell(cell, val)
            else:
                cell.fill = _fill(row_bg)
                cell.font = _body_font()
                if dt and val is not None:
                    cell.number_format = "m/d/yy"
                    cell.alignment = _center()
                elif pct and isinstance(val, float):
                    cell.number_format = "0.00%"
                    cell.alignment = _center()
                elif isinstance(val, str):
                    cell.alignment = _left()
                elif grp in ("ACP", "LIB") and not dt and not pct and isinstance(val, (int, float)):
                    cell.alignment = _center()
                else:
                    cell.alignment = _left()
                if hdr in EXCEL_WIDE_READABILITY_COLUMNS:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        ws.row_dimensions[data_row_idx].height = EXCEL_RECON_DATA_ROW_HEIGHT

    _autofit_recon_columns_for_readability(ws, data_start_row=5, data_end_row=4 + len(df_recon))
    ws.freeze_panes = "A5"

    ws_leg = wb.create_sheet("Legend")
    ws_leg.column_dimensions["A"].width = 22
    ws_leg.column_dimensions["B"].width = 50
    ws_leg.cell(1, 1).value = "Legend — Reconciliation Status"
    ws_leg.cell(1, 1).font = Font(name="Arial", size=12, bold=True, color=HEADER_BG)
    ws_leg.cell(1, 1).alignment = _left()

    legends = [
        ("MATCH", MATCH_BG, "375623", wb_ctx["legend_match_detail"]),
        ("MISMATCH", MISMATCH_BG, "9C0006", "Key field values differ between primary model and M61"),
        (
            "MISSING",
            MISSING_BG,
            "7D6608",
            "Record exists in one file only, red-flagged in M61 export, or a key field is "
            "MISSING FROM ACORE / MISSING FROM M61 / MISSING FROM BOTH.",
        ),
        (FILE_SOURCE_ACORE_ONLY, "D9E1F2", "1F3864", wb_ctx["legend_primary_only_detail"]),
        (FILE_SOURCE_BOTH, "E2EFDA", "375623", "Record found in primary model and M61 — basis for comparison"),
    ]

    for r, (lbl, bg_hex, fc, desc) in enumerate(legends, 3):
        c1 = ws_leg.cell(r, 1)
        c1.value = lbl
        c1.fill = _fill(bg_hex)
        c1.font = Font(name="Arial", size=9, bold=True, color=fc)
        c1.alignment = _center()
        c1.border = BORDER

        c2 = ws_leg.cell(r, 2)
        c2.value = desc
        c2.font = _body_font()
        c2.alignment = _left()
        ws_leg.row_dimensions[r].height = 18

    return wb


def run(
    file_a_path=DEFAULT_FILE_A_PATH,
    file_b_path=DEFAULT_FILE_B_PATH,
    out_path: str | None = None,
    primary_file_type: str = "ACORE",
    mapping_path: str | None = None,
):
    if out_path is None:
        out_path = default_recon_output_path(
            primary_file_type,
            uploaded_filename=os.path.basename(file_b_path),
        )
    print(f"Loading M61 export     : {file_a_path}")
    print(f"Loading primary file   : {file_b_path}  ({primary_file_type})")

    df_recon, df_adv, df_excluded = reconcile(
        file_a_path,
        file_b_path,
        primary_file_type=primary_file_type,
        mapping_path=mapping_path,
        uploaded_primary_filename=os.path.basename(file_b_path),
    )
    df_recon = normalize_recon_fund_for_output(df_recon)

    print("\nRECONCILIATION SUMMARY")
    print("=" * 40)
    for status, count in df_recon["recon_status"].value_counts().items():
        print(f"  {status:<15} {count:>4}")
    print(f"  {'TOTAL':<15} {len(df_recon):>4}")
    print("=" * 40)

    wb = build_workbook(df_recon, primary_file_type=primary_file_type)
    wb.save(out_path)
    print(f"\nSaved → {out_path}")
    print(
        "Excluded M61 rows by Liability Type filter "
        f"(visible in Streamlit diagnostics): {len(df_excluded)}"
    )
    return df_recon, df_adv


def main():
    parser = argparse.ArgumentParser(description="Financing Line Reconciliation — Enhanced Output")
    parser.add_argument("--file-a", dest="file_a", default=DEFAULT_FILE_A_PATH, help="Path to (In) M61 export")
    parser.add_argument("--file-b", dest="file_b", default=DEFAULT_FILE_B_PATH, help="Path to primary business file")
    parser.add_argument(
        "--primary-type",
        default="ACORE",
        choices=tuple(sorted(PRIMARY_FILE_CONFIG.keys())),
        help="Primary workbook template (column mapping)",
    )
    parser.add_argument(
        "--out",
        default=None,
        help="Output Excel path (default: <primary-type> Finance Recon - YYYY-MM-DD.xlsx in script dir)",
    )
    parser.add_argument(
        "--inspect-primary",
        action="store_true",
        help="Print configured primary sheet and column headers for --file-b and --primary-type, then exit",
    )
    parser.add_argument(
        "--mapping",
        default=None,
        help="Path to LiabilityNoteID -> CRENoteID mapping workbook (required for --primary-type AOC II or AOC I)",
    )
    args = parser.parse_args()
    if args.inspect_primary:
        inspect_primary_workbook(args.file_b, args.primary_type)
        return
    out_path = args.out or default_recon_output_path(
        args.primary_type,
        uploaded_filename=os.path.basename(args.file_b),
    )
    run(
        args.file_a,
        args.file_b,
        out_path,
        primary_file_type=args.primary_type,
        mapping_path=args.mapping,
    )


if __name__ == "__main__":
    main()